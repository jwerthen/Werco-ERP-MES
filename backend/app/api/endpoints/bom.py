import logging
from datetime import datetime
from typing import Any, Dict, Iterable, List, NoReturn, Optional, Sequence, Set, Tuple

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.concurrency import run_in_threadpool
from sqlalchemy import String, cast, func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, aliased, joinedload, selectinload

from app.api.deps import get_audit_service, get_current_company_id, get_current_user, require_role
from app.core.time_utils import to_utc_iso
from app.db.database import get_db
from app.models.bom import BOM, BOMItem, BOMItemType, BOMLineType
from app.models.part import Part, PartType, UnitOfMeasure, uom_disagrees, uom_label
from app.models.user import User, UserRole
from app.models.work_center import WorkCenter
from app.schemas.bom import (
    BOMCreate,
    BOMExploded,
    BOMFlatItem,
    BOMFlattened,
    BOMItemCreate,
    BOMItemResponse,
    BOMItemUpdate,
    BOMItemWithChildren,
    BOMLineUomMismatch,
    BOMResponse,
    BOMUomMismatchReport,
    BOMUpdate,
    ComponentPartInfo,
    PartInfo,
)
from app.schemas.bom_import import (
    BOMImportAssembly,
    BOMImportCommitRequest,
    BOMImportItem,
    BOMImportPreviewResponse,
    BOMImportResponse,
)
from app.services.audit_service import AuditService
from app.services.completion_inventory_service import (
    armed_parts_affected_by_bom,
    backflush_armed_edit_warning,
    bom_line_is_backflush_consumed,
)
from app.services.import_service import (
    MAX_CONSECUTIVE_BLANK_ROWS,
    MAX_IMPORT_COLUMNS,
    MAX_IMPORT_ROWS,
    MAX_SCANNED_ROWS,
    ImportFileError,
)
from app.services.llm_service import extract_bom_data_with_llm
from app.services.part_number_service import generate_werco_part_number
from app.services.pdf_service import SUPPORTED_EXTENSIONS, extract_text_from_document, save_uploaded_document

logger = logging.getLogger(__name__)

router = APIRouter()


def parts_with_active_bom(db: Session, part_ids: Iterable[int], company_id: int) -> Set[int]:
    """Which of ``part_ids`` have an active, NON-DELETED BOM in this company.

    THE single ``has_bom`` probe for this module. It drives the expand/drill-down
    affordance in the BOM tree, so it was quietly wrong in two ways at all four sites that
    open-coded it: no ``company_id`` predicate (the residual-foreign-key shape invariant #1
    exists to close), and no ``is_deleted`` predicate on a model that carries
    ``SoftDeleteMixin`` -- so a soft-deleted BOM made the response claim the component is
    an assembly and offered a drill-down into a structure the shop had deleted.
    """
    ids = {pid for pid in part_ids if pid}
    if not ids:
        return set()
    rows = (
        db.query(BOM.part_id)
        .filter(
            BOM.part_id.in_(ids),
            BOM.company_id == company_id,
            BOM.is_active == True,  # noqa: E712
            BOM.is_deleted == False,  # noqa: E712
        )
        .all()
    )
    return {row.part_id for row in rows}


def tenant_parts_by_id(db: Session, part_ids: Iterable[Optional[int]], company_id: int) -> Dict[int, Part]:
    """``{part_id: Part}`` for a BOM's parts, resolved TENANT-SCOPED, in ONE query.

    THE way this module turns a part id on a BOM row into a renderable ``Part`` -- both the
    LINE's ``component_part_id`` and the HEADER's ``part_id``. Never ``BOMItem.
    component_part`` and never ``BOM.part``: both relationships join on the foreign key
    alone (``models/bom.py``) and apply no ``company_id`` predicate, so on a mis-parented
    row either one happily materialises ANOTHER COMPANY's part and every response builder
    here rendered its part number, name and revision straight back to the caller.

    The WRITE side of that hole is already closed -- all four line-write paths scope
    ``component_part_id`` to the active company, and ``create_bom`` / the importers scope
    the header's ``part_id`` -- but the READ side must not lean on that: it renders rows
    written before those fixes, rows a future door might write, and rows a residual foreign
    key left behind. Scoping the LOOKUP means the foreign object is never materialised at
    all, rather than materialised and then carefully not printed (the same reasoning, and
    the same shape, as ``completion_inventory_service._tenant_components``).

    Batched because the read paths are hot: a 40-line BOM must not become 40 round trips.
    A part that does not come back is simply absent from the map, and the callers render
    ``component_part: null`` / ``part: null`` -- exactly what they already did for a
    component whose part row had been hard-deleted.
    """
    ids = {int(pid) for pid in part_ids if pid}
    if not ids:
        return {}
    rows = db.query(Part).filter(Part.id.in_(ids), Part.company_id == company_id).all()
    return {row.id: row for row in rows}


def component_part_info(part: Part, *, has_bom: bool) -> ComponentPartInfo:
    """Render a component part into the response - handles NULL values defensively.

    THE one place the ``ComponentPartInfo`` shape is built; ``has_bom`` is supplied rather
    than probed, because this module knows it in three different ways and two of them
    already hold the answer:

    * PROBE it -- ``get_component_part_info`` (below), for a single part. The general path.
    * A BATCHED ``has_bom_by_part_id`` map, resolved once per page by ``list_boms`` /
      ``get_bom`` and passed down, so a 40-line BOM costs one probe instead of forty.
    * Straight off a BOM ROW already in hand: ``explode_bom_recursive`` has resolved the
      component's own BOM because it needs it to recurse into, and
      ``component_bom is not None`` is the same predicate ``parts_with_active_bom``
      evaluates (same ``part_id`` / ``company_id`` / ``is_active`` / ``is_deleted``
      filters), so probing again was a second identical query per line.

    Keeping the construction here is what stops those three ``has_bom`` sources drifting
    into subtly different ``ComponentPartInfo`` shapes -- three sites had already
    open-coded this constructor, which is how the drift starts.
    """
    return ComponentPartInfo(
        id=part.id,
        part_number=part.part_number or "",
        name=part.name or "",
        revision=part.revision or "A",
        part_type=part.part_type.value if part.part_type else "manufactured",
        has_bom=has_bom,
    )


def get_component_part_info(part: Part, db: Session, company_id: int) -> ComponentPartInfo:
    """``component_part_info`` for a single part, probing ``has_bom`` itself.

    ``company_id`` is the CALLER'S ACTIVE COMPANY, and it is a required argument rather
    than something read off ``part``. The previous version probed ``part.company_id``,
    which is only ever the right answer when ``part`` was itself resolved tenant-scoped --
    and the one caller that mattered handed it an object off the unscoped
    ``BOMItem.component_part`` relationship. Every caller now resolves the part through
    ``tenant_parts_by_id`` first, so this parameter is the active company by
    construction; taking it explicitly is what stops the next caller re-deriving it from
    data it does not control.

    One query per call: use it on single-item paths only. A path that already knows
    ``has_bom`` -- from a batched probe or from a BOM row it has in hand -- must call
    ``component_part_info`` directly rather than pay for the probe again.
    """
    return component_part_info(part, has_bom=bool(parts_with_active_bom(db, [part.id], company_id)))


# ---------------------------------------------------------------------------
# Released-BOM edit policy
# ---------------------------------------------------------------------------
# FROZEN <=> ``bom.status != "draft"``. Allowlist-shaped ON PURPOSE: a BOM whose status
# column holds junk (an unvalidated ``BOMUpdate.status`` used to be able to write any
# string -- that hole is closed by deleting the field, but rows it already wrote may exist)
# lands on the SAFE side, frozen, instead of being accidentally editable. The comparison is
# against the lowercase literals ``release_bom`` / ``unrelease_bom`` write, nothing else.
#
# Unlike ``routing.py``, which carves out an in-place lane for time standards on a released
# routing, a released BOM is refused OUTRIGHT: BOM has an ``unrelease`` verb (routing does
# not), so ``unrelease -> edit -> re-release`` strands nobody, costs one gated click, and
# leaves a BETTER AS9100D record than an in-place edit would -- withdrawal, changes, and
# re-approval are three separate rows on the chain instead of one silent mutation.
#
# 400, not 409, for every status-state refusal: same class of guard as routing's, and no
# frontend path branches on 409 for BOM. Role refusals stay 403 via ``require_role``.
_BOM_FROZEN_LINE_MSG = (
    "Released BOM: lines cannot be added, changed or removed — unrelease the BOM to edit it, then release it again."
)
_BOM_NOT_DRAFT_MSG = "Cannot modify a BOM with status '{status}'."

# The ONLY field editable on a released BOM header. ``revision`` relabels an approved
# controlled document without re-approval (the revision IS the document's identity),
# ``bom_type`` changes how the BOM explodes into demand (phantom vs standard), and
# ``effective_date`` is AS9100D effectivity stamped by ``release_bom`` -- re-dating it
# rewrites when the approved configuration took effect. A description is metadata ABOUT the
# document, not the configuration.
_BOM_RELEASED_EDITABLE_FIELDS = {"description"}


def _assert_bom_lines_editable(bom: BOM) -> None:
    """Refuse a BOM-LINE write unless the parent BOM is a draft (400).

    Called by all three line verbs (``add_bom_item`` / ``update_bom_item`` /
    ``delete_bom_item``) BEFORE the first ``setattr`` / ``db.add`` / ``db.delete``, so a
    refusal leaves the row untouched -- the same "evaluate before mutating" placement
    ``routing.update_operation`` and ``parts.assert_backflush_change_allowed`` use.

    On ``add_bom_item`` it must also sit AFTER the parent 404, so a foreign ``bom_id``
    cannot be probed by the shape of the error it comes back with.
    """
    if bom.status == "draft":
        return
    if bom.status == "released":
        raise HTTPException(status_code=400, detail=_BOM_FROZEN_LINE_MSG)
    raise HTTPException(status_code=400, detail=_BOM_NOT_DRAFT_MSG.format(status=bom.status))


def bom_identifier(part_number: Optional[str], revision: Optional[str]) -> str:
    """The human handle for a BOM HEADER, used as the audit row's ``resource_identifier``.

    Same reasoning as ``bom_line_identifier``: an auditor reads the trail by part number
    and revision, not by row id. The part number is resolved TENANT-SCOPED by the callers
    (``_tenant_part_number``) -- never off ``BOM.part``, which carries no ``company_id``
    predicate and on a mis-parented row would write a foreign part number into this
    tenant's audit chain.
    """
    return f"{part_number or '?'} BOM rev {revision or '?'}"


def bom_line_identifier(
    parent_part_number: Optional[str], item_number: Optional[int], component_part_number: Optional[str]
) -> str:
    """The human handle for a BOM line, used as the audit row's ``resource_identifier``.

    An auditor reads the trail by part number, not by row id, so the identifier names the
    assembly, the line and the component. Both part numbers are resolved TENANT-SCOPED by
    the callers -- never off ``BOMItem.component_part`` / ``BOM.part``, which carry no
    ``company_id`` predicate and on a mis-parented row would write a foreign part number
    into this tenant's audit chain.
    """
    parent = parent_part_number or "?"
    component = f" ({component_part_number})" if component_part_number else ""
    return f"{parent} line {item_number if item_number is not None else '?'}{component}"


def _audit_values(bom: BOM, fields: Iterable[str]) -> Dict[str, Any]:
    """``{field: value}`` for an audit diff, with datetimes normalised to UTC ISO-8601.

    ``old_values`` is read off the persisted row, where a ``DateTime`` column comes back
    NAIVE; ``new_values`` is read back after the setattr, where an ``effective_date`` that
    arrived on the request body is OFFSET-AWARE. Serialised raw, the same instant appears on
    the two halves of one audit row in two different formats ("2026-01-02T03:04:05" vs
    "...+00:00"), which reads as a change of shape rather than of value to anyone diffing
    the chain. Both halves go through ``to_utc_iso`` so the comparison is like-for-like --
    the same "store UTC, serve UTC (Z)" rule the response schemas follow.
    """
    values: Dict[str, Any] = {}
    for field in fields:
        value = getattr(bom, field)
        values[field] = to_utc_iso(value) if isinstance(value, datetime) else value
    return values


def _tenant_part_number(db: Session, part_id: Optional[int], company_id: int) -> Optional[str]:
    """A part number resolved TENANT-SCOPED, or None. Display/audit text only."""
    if part_id is None:
        return None
    return db.query(Part.part_number).filter(Part.id == part_id, Part.company_id == company_id).scalar()


def _assert_work_center_owned(db: Session, work_center_id: Optional[int], company_id: int) -> None:
    """A BOM line's optional ``work_center_id`` must name a machine in THIS company.

    THE guard for every BOM-line write path that accepts the field (invariant #1). The
    component part and the parent part were always scoped; ``work_center_id`` rode in
    unchecked, so a caller in company B could point a line at company A's machine and leak
    that machine's identity back through routing / explosion reports.

    There are FOUR BOM-line write paths (``grep "BOMItem("``). Two accept a caller-supplied
    ``work_center_id`` -- ``add_bom_item`` and ``create_bom``'s inline ``items`` -- and both
    call this. ``update_bom_item`` is the third door onto an existing row and calls it too.
    The two IMPORT paths never set the field at all, so they have nothing to check.

    404 rather than 403, matching the component check, so a foreign id cannot be probed.
    ``None`` is always allowed: it means "no work center", and on the update path it CLEARS
    the reference, which needs no ownership to do.
    """
    if work_center_id is None:
        return
    owned = db.query(WorkCenter.id).filter(WorkCenter.id == work_center_id, WorkCenter.company_id == company_id).first()
    if not owned:
        raise HTTPException(status_code=404, detail="Work center not found")


def _armed_extra_data(armed_parts: Sequence[Part]) -> Optional[Dict[str, Any]]:
    """``extra_data`` naming the backflush-armed parts this BOM edit affects, or None.

    Only when the list is NON-EMPTY -- an empty key on every BOM-line row would be noise on
    a chain that is read by hand. This is what makes the arming verdict and the later edit
    correlatable on ONE chain: the flip is queried as ``resource_type='part' AND
    action='UPDATE' AND extra_data->>'backflush_readiness' IS NOT NULL`` (see
    ``parts.assert_backflush_change_allowed``), and this is its counterpart on the edit
    side. The list is "at least these" -- see ``armed_parts_affected_by_bom`` for the bound.
    """
    if not armed_parts:
        return None
    return {"backflush_armed_parts": [p.part_number or f"part {p.id}" for p in armed_parts]}


def build_bom_item_response(
    item: BOMItem,
    db: Session,
    has_bom_by_part_id: Optional[dict] = None,
    *,
    company_id: int,
    components_by_id: Optional[Dict[int, Part]] = None,
    backflush_armed_warning: Optional[str] = None,
) -> BOMItemResponse:
    """Build BOM item response with part info - handles NULL values defensively.

    ``company_id`` is REQUIRED (keyword-only) because this function renders the component
    part into the response: it resolves ``item.component_part_id`` through
    ``tenant_parts_by_id`` rather than reading the unscoped
    ``BOMItem.component_part`` relationship. ``components_by_id`` is that same map,
    pre-resolved by a list caller so a page of BOMs costs one component query instead of
    one per line; omit it on a single-item path and this does the one scoped read itself.
    """
    # Handle component_part safely - it might be None if the part was deleted, hard-deleted,
    # or (the case this scoping exists for) owned by another tenant.
    component_info = None
    component = (
        components_by_id.get(item.component_part_id)
        if components_by_id is not None
        else tenant_parts_by_id(db, [item.component_part_id], company_id).get(item.component_part_id)
    )
    if component:
        try:
            if has_bom_by_part_id is not None:
                component_info = component_part_info(component, has_bom=has_bom_by_part_id.get(component.id, False))
            else:
                component_info = get_component_part_info(component, db, company_id)
        except Exception as e:
            logger.warning("Failed to get component part info for BOM item %s: %s", item.id, e)

    return BOMItemResponse(
        id=item.id,
        bom_id=item.bom_id,
        component_part_id=item.component_part_id,
        item_number=item.item_number if item.item_number is not None else 10,
        quantity=item.quantity if item.quantity is not None else 1.0,
        item_type=item.item_type if item.item_type else BOMItemType.MAKE,
        line_type=item.line_type if item.line_type else BOMLineType.COMPONENT,
        unit_of_measure=item.unit_of_measure or "each",
        reference_designator=item.reference_designator,
        find_number=item.find_number,
        notes=item.notes,
        torque_spec=item.torque_spec,
        installation_notes=item.installation_notes,
        work_center_id=item.work_center_id,
        operation_sequence=item.operation_sequence if item.operation_sequence is not None else 10,
        scrap_factor=item.scrap_factor if item.scrap_factor is not None else 0.0,
        lead_time_offset=item.lead_time_offset if item.lead_time_offset is not None else 0,
        is_optional=item.is_optional if item.is_optional is not None else False,
        is_alternate=item.is_alternate if item.is_alternate is not None else False,
        alternate_group=item.alternate_group,
        component_part=component_info,
        backflush_armed_warning=backflush_armed_warning,
        created_at=item.created_at,
        updated_at=item.updated_at,
    )


def _normalize_uom(value: Optional[str]) -> str:
    if not value:
        return UnitOfMeasure.EACH.value
    val = value.strip().lower()
    mapping = {
        "ea": "each",
        "each": "each",
        "pcs": "each",
        "pc": "each",
        "lb": "pounds",
        "lbs": "pounds",
        "pound": "pounds",
        "ft": "feet",
        "feet": "feet",
        "in": "inches",
        "inch": "inches",
        "inches": "inches",
        "gal": "gallons",
        "gallon": "gallons",
        "l": "liters",
        "liter": "liters",
    }
    return mapping.get(val, val)


def _component_default_uom(component: Optional[Part]) -> str:
    """The unit a BOM line INHERITS when its author stated none: the component part's own.

    Owner decision, 2026-07-27. The literal ``"each"`` this replaced was a default nobody
    chose, and the BLOCKING ``unit_of_measure_mismatch`` diagnostic
    (``completion_inventory_service``) reads a stored unit as a STATED CLAIM -- so on a
    sheet-metal shop's data, where components are stocked in sheets / lbs / ft, the
    automatic-backflush opt-in refused nearly every part over a value the shop never typed.

    ``"each"`` survives only as the last resort: a component that cannot be resolved, or
    one whose own ``unit_of_measure`` is NULL. Falling back to the column default keeps a
    line's unit non-null the way it has always been, and a component with no stocking unit
    is silent in ``uom_disagrees`` anyway, so this cannot manufacture a mismatch.
    """
    return uom_label(getattr(component, "unit_of_measure", None)) or UnitOfMeasure.EACH.value


def _resolve_line_uom(stated: Optional[str], component: Optional[Part], *, normalize_stated: bool = False) -> str:
    """The unit of measure to STORE on a BOM line: what the caller said, else the part's.

    Used by all four BOM-line write paths (see the module note on ``add_bom_item``). A
    stated value always wins -- this resolves an ABSENCE, it does not second-guess a human.

    ``normalize_stated`` is the one difference between the doors and is intentional: the
    two importer paths already ran free-text spreadsheet/LLM values through
    ``_normalize_uom`` (``ea`` -> ``each``, ``lbs`` -> ``pounds``) before this change and
    still do, while the two JSON API paths have always stored the client's string verbatim.
    Turning normalisation on for the API paths would silently rewrite a value a client sent
    on purpose, which is a bigger behaviour change than the default this fixes.
    """
    if stated is not None and str(stated).strip():
        return _normalize_uom(stated) if normalize_stated else str(stated)
    return _component_default_uom(component)


def _coerce_item_type(value: Optional[str]) -> str:
    if not value:
        return BOMItemType.BUY.value
    val = value.strip().lower()
    if val in {BOMItemType.MAKE.value, BOMItemType.BUY.value, BOMItemType.PHANTOM.value}:
        return val
    return BOMItemType.BUY.value


def _coerce_line_type(value: Optional[str]) -> str:
    if not value:
        return BOMLineType.COMPONENT.value
    val = value.strip().lower()
    if val in {
        BOMLineType.COMPONENT.value,
        BOMLineType.HARDWARE.value,
        BOMLineType.CONSUMABLE.value,
        BOMLineType.REFERENCE.value,
    }:
        return val
    return BOMLineType.COMPONENT.value


def _classify_line_type(description: str, explicit: Optional[str]) -> str:
    if explicit:
        return _coerce_line_type(explicit)
    text = (description or "").lower()
    hardware_keywords = [
        "bolt",
        "screw",
        "washer",
        "nut",
        "fastener",
        "pin",
        "rivet",
        "clip",
        "stud",
        "standoff",
        "spacer",
    ]
    consumable_keywords = [
        "adhesive",
        "loctite",
        "glue",
        "epoxy",
        "tape",
        "oil",
        "grease",
        "lubricant",
        "paint",
        "primer",
        "sealant",
    ]
    reference_keywords = ["reference", "ref only", "for reference", "ref."]
    if any(k in text for k in hardware_keywords):
        return BOMLineType.HARDWARE.value
    if any(k in text for k in consumable_keywords):
        return BOMLineType.CONSUMABLE.value
    if any(k in text for k in reference_keywords):
        return BOMLineType.REFERENCE.value
    return BOMLineType.COMPONENT.value


def _infer_part_type(line_type: str, item_type: str, description: str) -> str:
    if line_type == BOMLineType.HARDWARE.value:
        return PartType.HARDWARE.value
    if line_type == BOMLineType.CONSUMABLE.value:
        return PartType.CONSUMABLE.value
    if line_type == BOMLineType.REFERENCE.value:
        return PartType.PURCHASED.value
    text = (description or "").lower()
    if item_type == BOMItemType.PHANTOM.value:
        return PartType.ASSEMBLY.value
    if item_type == BOMItemType.MAKE.value:
        if "assembly" in text or "assy" in text:
            return PartType.ASSEMBLY.value
        return PartType.MANUFACTURED.value
    return PartType.PURCHASED.value


def _part_type_value(value: Optional[Any]) -> Optional[str]:
    if value is None:
        return None
    if hasattr(value, "value"):
        return str(value.value).strip().lower()
    return str(value).strip().lower()


def _resolve_import_parent_part_type(doc_type: str, extracted_type: Optional[str]) -> str:
    if doc_type == "bom":
        return PartType.ASSEMBLY.value

    normalized = (extracted_type or PartType.MANUFACTURED.value).strip().lower()
    if normalized not in {p.value for p in PartType}:
        return PartType.MANUFACTURED.value
    return normalized


def _safe_part_number(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    return value.strip()


def _generate_fallback_part_number(prefix: str, index: int) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    return f"{prefix}-{timestamp}-{index:03d}"


def _reject_deleted_part(db: Session, part_number: str) -> NoReturn:
    """Fail the import when a part number collides with a soft-deleted part.

    ``uq_parts_company_part_number`` has no soft-delete carve-out, so the
    deleted row still owns the number: silently reusing it would resurrect
    deleted data, and creating fresh would raise IntegrityError. Mirror the
    POST /parts precedent (400 + the /parts/{id}/restore recovery path). The
    import is a single transaction — roll back anything already staged so a
    partial BOM never persists.
    """
    db.rollback()
    raise HTTPException(
        status_code=400,
        detail=(
            f"Part '{part_number}' matches a deleted part. Restore it from Parts "
            "(or use a different part number) and re-import."
        ),
    )


def _existing_bom_conflict_detail(
    part_number: Optional[str], *, bom_is_deleted: bool, bom_is_active: bool, retry: str
) -> str:
    """The 400 body for "this assembly part already owns a BOM row", branched on its state.

    Shared by the importer (``_reject_existing_bom``) and ``POST /bom/`` because they hit
    the SAME wall: ``BOM.part_id`` is UNIQUE with no soft-delete/active carve-out, so a
    probe that only looks for ``is_active == True`` rows cannot see the row that owns the
    slot and the insert dies on the constraint with an IntegrityError 500. Now that
    ``delete_bom`` is a SOFT delete that leaves the row in place, the deleted branch is the
    common case rather than the theoretical one -- and both doors name the same recovery,
    ``POST /bom/{bom_id}/restore``.
    """
    if bom_is_deleted:
        return f"A deleted BOM exists for part '{part_number}' — restore it before {retry}."
    if not bom_is_active:
        return f"An inactive BOM exists for part '{part_number}' — reactivate or delete it before {retry}."
    return f"A BOM already exists for assembly part '{part_number}'"


def _reject_existing_bom(db: Session, assembly_part: Part, company_id: int) -> None:
    """Fail the import if ANY BOM row already occupies the assembly part.

    ``BOM.part_id`` is unique with no soft-delete/active carve-out. The old
    ``is_active == True`` lookup made soft-deleted or inactive BOM rows
    invisible, so the import tried to create a second BOM and died with an
    IntegrityError 500. Branch on the row's state instead and return an
    actionable 400.

    This is also why NO import path needs a released-BOM guard: it refuses *any*
    pre-existing BOM row for the assembly part, so an importer can never reach a released
    BOM's lines in the first place.
    """
    existing_bom = db.query(BOM).filter(BOM.part_id == assembly_part.id, BOM.company_id == company_id).first()
    if existing_bom is None:
        return
    # Capture state before rollback expires the instances.
    detail = _existing_bom_conflict_detail(
        assembly_part.part_number,
        bom_is_deleted=bool(existing_bom.is_deleted),
        bom_is_active=bool(existing_bom.is_active),
        retry="importing",
    )
    # Single-transaction import: discard anything already staged (e.g. the
    # in-place part_type promotion and its audit row) before failing.
    db.rollback()
    raise HTTPException(status_code=400, detail=detail)


def _flush_new_bom(db: Session, part_number: Optional[str], *, retry: str) -> None:
    """Flush a freshly-added ``BOM``, turning the UNIQUE-slot collision into a 400.

    The LAST line of defence behind ``_reject_existing_bom`` / ``create_bom``'s probe, not a
    replacement for them: those give a specific, actionable body ("restore it", "reactivate
    it") and this cannot, because the row it collided with is one this caller is not allowed
    to see. ``BOM.part_id`` is UNIQUE **globally**, with no ``company_id`` in the
    constraint, while both probes are correctly company-scoped -- so a residual row
    belonging to ANOTHER tenant occupies the slot invisibly and the insert died on the
    constraint as an uncaught IntegrityError 500. The body is deliberately flat and says
    nothing about who owns the row or that one exists elsewhere (invariant #1: never
    confirm existence across a tenant boundary); 400 because the caller genuinely cannot
    create this BOM, and a 500 would page someone for a data condition no code path
    produces any more.
    """
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        logger.warning("BOM insert collided with the unique part_id slot for part %r", part_number)
        raise HTTPException(
            status_code=400,
            detail=f"Cannot create a BOM for part '{part_number}' — the part already has one. Contact an administrator "
            f"before {retry}.",
        )


def _ensure_part(
    db: Session,
    part_number: Optional[str],
    name: str,
    description: str,
    part_type: str,
    drawing_number: Optional[str],
    unit_of_measure: Optional[str],
    create_missing: bool,
    fallback_index: int,
    company_id: int,
    audit: AuditService,
    created_by: Optional[int] = None,
) -> Tuple[Optional[Part], Optional[str], bool]:
    if part_number:
        existing = db.query(Part).filter(Part.part_number == part_number, Part.company_id == company_id).first()
        if existing:
            if existing.is_deleted:
                _reject_deleted_part(db, part_number)
            return existing, None, False
    if not create_missing:
        return None, part_number or name, False

    normalized_type = part_type if part_type in {p.value for p in PartType} else PartType.PURCHASED.value
    candidate_number = part_number
    if not candidate_number:
        if normalized_type in {PartType.RAW_MATERIAL.value, PartType.HARDWARE.value, PartType.CONSUMABLE.value}:
            candidate_number = generate_werco_part_number(description or name, normalized_type)
        if not candidate_number:
            candidate_number = _generate_fallback_part_number("AUTO", fallback_index)

    part = Part(
        part_number=candidate_number,
        revision="A",
        name=name or candidate_number,
        description=description,
        part_type=normalized_type,
        unit_of_measure=_normalize_uom(unit_of_measure),
        drawing_number=drawing_number,
        company_id=company_id,
        created_by=created_by,
    )
    db.add(part)
    db.flush()
    # Before the terminal commit so the audit row persists atomically with the part.
    audit.log_create("part", part.id, part.part_number, new_values=part, extra_data={"source": "bom_import"})
    return part, None, True


def _build_preview(extracted: Dict[str, Any]) -> Tuple[BOMImportAssembly, List[BOMImportItem], List[str], str]:
    warnings: List[str] = []
    items: List[BOMImportItem] = []
    assembly_data = extracted.get("assembly", {}) or {}
    assembly = BOMImportAssembly(
        part_number=_safe_part_number(assembly_data.get("part_number"))
        or _safe_part_number(assembly_data.get("drawing_number")),
        name=assembly_data.get("name"),
        revision=assembly_data.get("revision") or "A",
        description=assembly_data.get("description"),
        drawing_number=_safe_part_number(assembly_data.get("drawing_number")),
        part_type=assembly_data.get("part_type"),
    )

    for idx, item in enumerate(extracted.get("items", []) or [], start=1):
        line_number = int(item.get("line_number") or (idx * 10))
        description = (item.get("description") or "").strip()
        part_number = _safe_part_number(item.get("part_number"))
        line_type = _classify_line_type(description, item.get("line_type"))
        item_type = _coerce_item_type(item.get("item_type"))
        if not part_number:
            warnings.append(f"Line {line_number}: missing part number; will be generated if created.")
        if not description:
            warnings.append(f"Line {line_number}: missing description.")
        quantity = item.get("quantity")
        if quantity is None or float(quantity) <= 0:
            warnings.append(f"Line {line_number}: quantity not found or invalid; defaulting to 1.")
        items.append(
            BOMImportItem(
                line_number=line_number,
                part_number=part_number,
                description=description,
                quantity=float(quantity) if quantity and float(quantity) > 0 else 1.0,
                unit_of_measure=_normalize_uom(item.get("unit_of_measure")),
                item_type=item_type,
                line_type=line_type,
                reference_designator=item.get("reference_designator"),
                find_number=item.get("find_number"),
                notes=item.get("notes"),
            )
        )

    extraction_confidence = extracted.get("extraction_confidence", "low")
    return assembly, items, warnings, extraction_confidence


def _normalize_header(value: str) -> str:
    return "".join(ch for ch in value.strip().lower() if ch.isalnum() or ch.isspace()).replace(" ", "")


def _extract_excel_table(file_path: str, ext: str) -> Tuple[List[str], List[List[str]]]:
    """Extract ``(columns, data rows)`` from an uploaded Excel BOM with a bounded scan.

    Semantics are unchanged from the original implementation: the first
    non-empty row found anywhere in the workbook becomes ``columns``; every
    later non-empty row across ALL sheets becomes a data row; cell values are
    stringified and stripped.

    The scan is bounded the same way as the Import Center's shared parser
    (:func:`app.services.import_service.parse_import_file`), which fixed this
    exact bug class — one stray formatted/whitespace cell at XFD1048576 used to
    make openpyxl iterate the full 16,384 x 1,048,576 declared grid (minutes of
    CPU for a KB-sized file):

    * at most :data:`MAX_IMPORT_COLUMNS` columns are read per row;
    * a run of more than :data:`MAX_CONSECUTIVE_BLANK_ROWS` blank rows ends the
      scan of THAT sheet only — sheets are independent documents here, so the
      cutoff unit is the sheet and scanning continues with the next one.
      Unlike ``parse_import_file`` there is deliberately NO loud-refusal
      look-ahead past the gap: BOM spreadsheets legitimately scatter data
      blocks down a sheet, and the preview flow shows users exactly which rows
      parsed before anything is committed, so a quiet per-sheet cutoff after a
      1,000-row gap is the right trade-off;
    * a workbook-wide counter of raw rows scanned refuses the file past
      :data:`MAX_SCANNED_ROWS`;
    * collected data rows are capped at :data:`MAX_IMPORT_ROWS`.

    Raises :class:`ImportFileError` (mapped to HTTP 400 by the import
    endpoints) for corrupt/unreadable files and for workbooks that exceed the
    caps.
    """
    columns: List[str] = []
    rows: List[List[str]] = []
    scanned_rows = 0

    def _consume_sheet(sheet_rows: Iterable[Sequence[Any]]) -> None:
        """Fold one sheet's raw rows into ``columns``/``rows`` under the shared caps."""
        nonlocal columns, scanned_rows
        consecutive_blank_rows = 0
        for raw_row in sheet_rows:
            scanned_rows += 1
            if scanned_rows > MAX_SCANNED_ROWS:
                raise ImportFileError(
                    f"The spreadsheet's used range is enormous (over {MAX_SCANNED_ROWS:,} rows scanned). "
                    "Delete trailing empty rows/columns or re-save the data as CSV, then try again."
                )
            row_vals = ["" if cell is None else str(cell).strip() for cell in raw_row]
            # Read-only iteration pads every row to max_col; drop the trailing
            # padding so columns/rows keep the used range's natural width (the
            # shape the mapping UI and _items_from_table indexing expect).
            while row_vals and not row_vals[-1]:
                row_vals.pop()
            if not row_vals:
                consecutive_blank_rows += 1
                if consecutive_blank_rows > MAX_CONSECUTIVE_BLANK_ROWS:
                    return  # used-range bloat on this sheet — move on to the next sheet
                continue
            consecutive_blank_rows = 0
            if not columns:
                columns = row_vals
                continue
            if len(rows) >= MAX_IMPORT_ROWS:
                raise ImportFileError(f"Too many rows (max {MAX_IMPORT_ROWS}). Split the file and import in batches.")
            rows.append(row_vals)

    if ext == ".xlsx":
        from openpyxl import load_workbook

        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ImportFileError("Could not read the Excel file. Re-save it as a standard Excel workbook.") from exc
        try:
            for ws in wb.worksheets:
                _consume_sheet(ws.iter_rows(values_only=True, max_col=MAX_IMPORT_COLUMNS))
        except ImportFileError:
            raise
        except Exception as exc:  # read-only mode parses lazily; corruption can surface mid-iteration
            raise ImportFileError("Could not read the Excel file. Re-save it as a standard Excel workbook.") from exc
        finally:
            wb.close()
    else:
        import xlrd

        try:
            book = xlrd.open_workbook(file_path)
        except Exception as exc:
            raise ImportFileError("Could not read the Excel file. Re-save it as a standard Excel workbook.") from exc
        # The .xls grid is natively capped at 65,536 x 256; the same per-sheet
        # structure, column slice, and workbook-wide counter apply for uniformity.
        for sheet in book.sheets():
            _consume_sheet(
                [sheet.cell_value(r, c) for c in range(min(sheet.ncols, MAX_IMPORT_COLUMNS))]
                for r in range(sheet.nrows)
            )

    # Pad the header out to the widest data row: an unheadered trailing column
    # (notes/vendor exports) must still appear in the mapping UI as "Col N" and
    # stay manually mappable, as it did before the trailing-blank trim.
    max_width = max((len(r) for r in rows), default=0)
    if len(columns) < max_width:
        columns = columns + [""] * (max_width - len(columns))

    return columns, rows


def _suggest_mapping(columns: List[str]) -> Dict[str, Optional[int]]:
    synonyms = {
        "line_number": ["itemno", "itemnumber", "item", "lineno", "linenumber", "line", "item#", "itemno."],
        "part_number": ["part#", "partnumber", "partno", "pn", "p/n", "part"],
        "description": ["description", "desc", "partname", "name", "itemdescription", "material", "sheet"],
        "quantity": ["qty", "quantity", "quantityrequired", "reqqty", "q'ty"],
        "unit_of_measure": ["uom", "unit", "unitofmeasure", "units"],
        "item_type": ["itemtype", "makebuy", "make/buy", "mb"],
        "line_type": ["linetype", "type", "componenttype", "category"],
    }
    normalized = [_normalize_header(c) for c in columns]
    mapping: Dict[str, Optional[int]] = {k: None for k in synonyms.keys()}
    for field, keys in synonyms.items():
        for idx, name in enumerate(normalized):
            if any(key.replace("/", "") in name for key in keys):
                mapping[field] = idx
                break
    return mapping


def _items_from_table(
    columns: List[str], rows: List[List[str]], mapping: Dict[str, Optional[int]]
) -> List[BOMImportItem]:
    items: List[BOMImportItem] = []
    next_line = 10
    for row in rows:
        if not any(cell.strip() for cell in row):
            continue

        def get_val(field: str) -> str:
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return ""
            return str(row[idx]).strip()

        line_val = get_val("line_number")
        try:
            line_number = int(float(line_val)) if line_val else next_line
        except Exception:
            line_number = next_line
        next_line = line_number + 10

        description = get_val("description")
        part_number = get_val("part_number")
        qty_val = get_val("quantity")
        try:
            quantity = float(qty_val) if qty_val else 1.0
        except Exception:
            quantity = 1.0
        uom = get_val("unit_of_measure")
        item_type = get_val("item_type")
        line_type = get_val("line_type")
        line_type = _classify_line_type(description, line_type)
        item_type = _coerce_item_type(item_type)

        items.append(
            BOMImportItem(
                line_number=line_number,
                part_number=part_number or None,
                description=description or None,
                quantity=quantity if quantity > 0 else 1.0,
                unit_of_measure=_normalize_uom(uom),
                item_type=item_type,
                line_type=line_type,
            )
        )
    return items


def _create_from_import_payload(
    payload: BOMImportCommitRequest, db: Session, current_user: User, company_id: int, audit: AuditService
) -> BOMImportResponse:
    items = payload.items or []
    doc_type = (payload.document_type or ("bom" if items else "part")).lower()
    if items and doc_type != "bom":
        doc_type = "bom"

    warnings: List[str] = []
    missing_parts: List[str] = []

    assembly = payload.assembly
    assembly_number = _safe_part_number(assembly.part_number) or _safe_part_number(assembly.drawing_number)
    if not assembly_number:
        assembly_number = (
            _generate_fallback_part_number("ASSY", 1)
            if doc_type == "bom"
            else _generate_fallback_part_number("PART", 1)
        )
        warnings.append("Assembly/part number not found; generated a temporary number.")

    assembly_name = (assembly.name or assembly.description or assembly_number).strip()
    assembly_description = (assembly.description or assembly.name or "").strip()
    assembly_revision = (assembly.revision or "A").strip()
    assembly_drawing = _safe_part_number(assembly.drawing_number)
    assembly_part_type = _resolve_import_parent_part_type(doc_type, assembly.part_type)

    existing_part = db.query(Part).filter(Part.part_number == assembly_number, Part.company_id == company_id).first()
    if existing_part is not None and existing_part.is_deleted:
        _reject_deleted_part(db, assembly_number)
    if existing_part:
        assembly_part = existing_part
        if doc_type == "bom" and _part_type_value(assembly_part.part_type) != PartType.ASSEMBLY.value:
            old_part_type = _part_type_value(assembly_part.part_type)
            assembly_part.part_type = PartType.ASSEMBLY.value
            audit.log_update(
                "part",
                assembly_part.id,
                assembly_part.part_number,
                old_values={"part_type": old_part_type},
                new_values={"part_type": PartType.ASSEMBLY.value},
                extra_data={"source": "bom_import"},
            )
    else:
        assembly_part = Part(
            part_number=assembly_number,
            revision=assembly_revision,
            name=assembly_name,
            description=assembly_description,
            part_type=assembly_part_type,
            unit_of_measure=UnitOfMeasure.EACH.value,
            drawing_number=assembly_drawing,
            created_by=current_user.id,
            company_id=company_id,
        )
        db.add(assembly_part)
        db.flush()
        audit.log_create(
            "part",
            assembly_part.id,
            assembly_part.part_number,
            new_values=assembly_part,
            extra_data={"source": "bom_import"},
        )

    created_parts = 0 if existing_part else 1
    created_bom_items = 0
    bom_id: Optional[int] = None

    if doc_type == "part" and not items:
        db.commit()
        return BOMImportResponse(
            document_type="part",
            assembly_part_id=assembly_part.id,
            assembly_part_number=assembly_part.part_number,
            bom_id=None,
            created_parts=created_parts,
            created_bom_items=0,
            extraction_confidence="medium",
            warnings=warnings,
        )

    _reject_existing_bom(db, assembly_part, company_id)

    bom = BOM(
        part_id=assembly_part.id,
        revision=assembly_revision or "A",
        description=assembly_description,
        status="draft",
        bom_type="standard",
        created_by=current_user.id,
        company_id=company_id,
    )
    db.add(bom)
    _flush_new_bom(db, assembly_part.part_number, retry="retrying the import")
    bom_id = bom.id

    next_line = 10
    component_part_numbers: List[str] = []
    for idx, item in enumerate(items, start=1):
        item_number = int(item.line_number or next_line)
        next_line = item_number + 10
        description = (item.description or "").strip()
        item_part_number = _safe_part_number(item.part_number)
        line_type = _classify_line_type(description, item.line_type)
        item_type = _coerce_item_type(item.item_type)
        part_type = _infer_part_type(line_type, item_type, description)
        uom = item.unit_of_measure

        part_name = description or item_part_number or f"Item {item_number}"
        if not item_part_number:
            warnings.append(f"Line {item_number}: missing part number; generated automatically.")

        component_part, missing, was_created = _ensure_part(
            db,
            item_part_number,
            part_name,
            description,
            part_type,
            None,
            uom,
            payload.create_missing_parts,
            idx,
            company_id=company_id,
            audit=audit,
            created_by=current_user.id,
        )
        if missing:
            missing_parts.append(missing)
            continue
        if was_created:
            created_parts += 1
        component_part_numbers.append(component_part.part_number)

        quantity = float(item.quantity or 1)
        bom_item = BOMItem(
            bom_id=bom.id,
            component_part_id=component_part.id,
            item_number=item_number,
            quantity=quantity if quantity > 0 else 1.0,
            item_type=item_type,
            line_type=line_type,
            # BOM-LINE WRITE PATH 1 of 4 (review-commit importer). A row whose UOM column
            # was blank / unmapped inherits the component part's unit, not "each".
            unit_of_measure=_resolve_line_uom(uom, component_part, normalize_stated=True),
            reference_designator=item.reference_designator,
            find_number=item.find_number,
            notes=item.notes,
            company_id=company_id,
        )
        db.add(bom_item)
        created_bom_items += 1

    if missing_parts:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Missing parts: {', '.join(missing_parts)}")

    # One audit row for the BOM with the items summarized (house pattern: the
    # WO import logs the parent and summarizes children in extra_data), before
    # the terminal commit so it persists atomically with the import.
    #
    # ``bom_identifier``, NOT the bare part number -- see the note on the same call in
    # ``import_bom_or_part``.
    audit.log_create(
        "bom",
        bom.id,
        bom_identifier(assembly_part.part_number, bom.revision),
        new_values=bom,
        extra_data={
            "source": "bom_import",
            "item_count": created_bom_items,
            "component_part_numbers": component_part_numbers,
        },
    )

    db.commit()

    return BOMImportResponse(
        document_type="bom",
        assembly_part_id=assembly_part.id,
        assembly_part_number=assembly_part.part_number,
        bom_id=bom_id,
        created_parts=created_parts,
        created_bom_items=created_bom_items,
        extraction_confidence="medium",
        warnings=warnings,
    )


@router.post("/import/preview", response_model=BOMImportPreviewResponse)
async def import_bom_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
):
    """
    Upload a BOM or single-part document (PDF/DOC/DOCX/XLSX/XLS) and return a preview for review.

    Excel uploads are parsed directly into a raw table plus a suggested column mapping (no LLM
    call) with a bounded scan: all sheets are read, at most 256 columns per row, more than
    1,000 consecutive blank rows ends that sheet's scan (later sheets are still read — review
    the preview rows), and a file is refused with 400 if it yields more than 10,000 data rows,
    scans more than 100,000 raw rows workbook-wide, or cannot be read as an Excel workbook.
    PDF/Word uploads go through text extraction + LLM extraction.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    ext = f".{file.filename.split('.')[-1]}".lower() if "." in file.filename else ""
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use PDF, Word, or Excel documents.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    filename = file.filename

    # Save + extraction + table parse + LLM call are CPU-bound sync work; run
    # them in the threadpool so a pathological upload can't stall the event
    # loop (the request-scoped Session is used sequentially from one worker
    # thread — same as a sync endpoint).
    def _build_preview_response() -> BOMImportPreviewResponse:
        doc_path = save_uploaded_document(content, filename)

        if ext in [".xlsx", ".xls"]:
            # Excel goes straight to the bounded table parser; the generic
            # text extraction below only feeds the LLM path, so running it
            # here would pay a full-workbook scan for output nobody reads.
            columns, rows = _extract_excel_table(doc_path, ext)
            if not rows:
                raise HTTPException(status_code=400, detail="No data rows found in Excel file.")
            mapping = _suggest_mapping(columns)
            items = _items_from_table(columns, rows, mapping)
            warnings: List[str] = []
            if not mapping.get("part_number"):
                warnings.append("Part number column not detected. Map it in the preview.")
            if not mapping.get("quantity"):
                warnings.append("Quantity column not detected. Map it in the preview.")
            assembly = BOMImportAssembly()
            return BOMImportPreviewResponse(
                document_type="bom",
                assembly=assembly,
                items=items,
                extraction_confidence="medium",
                warnings=warnings,
                raw_columns=columns,
                raw_rows=rows,
                suggested_mapping=mapping,
                source_format="excel",
            )

        extraction_result = extract_text_from_document(doc_path)
        if not extraction_result.text or len(extraction_result.text.strip()) < 50:
            raise HTTPException(status_code=400, detail="Could not extract text from document")

        extracted = extract_bom_data_with_llm(
            extraction_result.text, is_ocr=extraction_result.is_ocr, company_id=company_id
        )
        if extracted.get("_error"):
            raise HTTPException(status_code=400, detail=extracted.get("_error"))

        llm_assembly, llm_items, llm_warnings, confidence = _build_preview(extracted)
        doc_type = (extracted.get("document_type") or ("bom" if llm_items else "part")).lower()
        if llm_items and doc_type != "bom":
            doc_type = "bom"

        return BOMImportPreviewResponse(
            document_type=doc_type,
            assembly=llm_assembly,
            items=llm_items,
            extraction_confidence=confidence,
            warnings=llm_warnings,
        )

    try:
        return await run_in_threadpool(_build_preview_response)
    except ImportFileError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/import/commit", response_model=BOMImportResponse, status_code=status.HTTP_201_CREATED)
def import_bom_commit(
    payload: BOMImportCommitRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """
    Commit a reviewed BOM/part import payload.

    Writes tamper-evident audit_log entries (extra_data.source = "bom_import"): one CREATE per
    created part, an UPDATE when an existing part is promoted to part_type=assembly, and one
    CREATE for the BOM with item_count + component part numbers summarized on the parent row.
    Conflicts are refused with actionable 400s and the whole import is rolled back: a part number
    matching a soft-deleted part, or a deleted / inactive / active BOM already occupying the
    assembly part.
    """
    return _create_from_import_payload(payload, db, current_user, company_id, audit)


@router.post("/import", response_model=BOMImportResponse, status_code=status.HTTP_201_CREATED)
async def import_bom_or_part(
    file: UploadFile = File(...),
    create_missing_parts: bool = Form(True),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """
    Upload a BOM or single-part document (PDF/DOC/DOCX/XLSX/XLS) and create parts/BOM items.

    The document is text-extracted and parsed by the LLM in one shot (no review step — prefer
    /bom/import/preview + /bom/import/commit for a reviewable flow). Excel text extraction is
    scan-bounded and degrades gracefully at the cap (partial text, "medium" confidence).

    Writes the same tamper-evident audit_log entries as /bom/import/commit
    (extra_data.source = "bom_import") and refuses the same conflicts with actionable 400s
    (soft-deleted part number collision; deleted / inactive / active BOM on the assembly part),
    rolling back the whole import.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="File name is required")

    ext = f".{file.filename.split('.')[-1]}".lower() if "." in file.filename else ""
    if ext not in SUPPORTED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use PDF, Word, or Excel documents.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    filename = file.filename

    # Save + extraction + LLM call + DB writes are CPU/DB-bound sync work; run
    # them in the threadpool so a large document can't stall the event loop
    # (the request-scoped Session and audit service are used sequentially from
    # one worker thread — same as a sync endpoint).
    def _run_import() -> BOMImportResponse:
        doc_path = save_uploaded_document(content, filename)
        extraction_result = extract_text_from_document(doc_path)
        if not extraction_result.text or len(extraction_result.text.strip()) < 50:
            raise HTTPException(status_code=400, detail="Could not extract text from document")

        extracted = extract_bom_data_with_llm(
            extraction_result.text, is_ocr=extraction_result.is_ocr, company_id=company_id
        )
        if extracted.get("_error"):
            raise HTTPException(status_code=400, detail=extracted.get("_error"))

        items = extracted.get("items", []) or []
        assembly = extracted.get("assembly", {}) or {}
        doc_type = (extracted.get("document_type") or ("bom" if items else "part")).lower()
        if items and doc_type != "bom":
            doc_type = "bom"

        warnings: List[str] = []
        missing_parts: List[str] = []

        assembly_number = _safe_part_number(assembly.get("part_number")) or _safe_part_number(
            assembly.get("drawing_number")
        )
        if not assembly_number:
            assembly_number = (
                _generate_fallback_part_number("ASSY", 1)
                if doc_type == "bom"
                else _generate_fallback_part_number("PART", 1)
            )
            warnings.append("Assembly/part number not found; generated a temporary number.")

        assembly_name = (assembly.get("name") or assembly.get("description") or assembly_number).strip()
        assembly_description = (assembly.get("description") or assembly.get("name") or "").strip()
        assembly_revision = (assembly.get("revision") or "A").strip()
        assembly_drawing = _safe_part_number(assembly.get("drawing_number"))
        assembly_part_type = _resolve_import_parent_part_type(doc_type, assembly.get("part_type"))

        # Get or create assembly part
        existing_part = (
            db.query(Part).filter(Part.part_number == assembly_number, Part.company_id == company_id).first()
        )
        if existing_part is not None and existing_part.is_deleted:
            _reject_deleted_part(db, assembly_number)
        if existing_part:
            assembly_part = existing_part
            if doc_type == "bom" and _part_type_value(assembly_part.part_type) != PartType.ASSEMBLY.value:
                old_part_type = _part_type_value(assembly_part.part_type)
                assembly_part.part_type = PartType.ASSEMBLY.value
                audit.log_update(
                    "part",
                    assembly_part.id,
                    assembly_part.part_number,
                    old_values={"part_type": old_part_type},
                    new_values={"part_type": PartType.ASSEMBLY.value},
                    extra_data={"source": "bom_import"},
                )
        else:
            assembly_part = Part(
                part_number=assembly_number,
                revision=assembly_revision,
                name=assembly_name,
                description=assembly_description,
                part_type=assembly_part_type,
                unit_of_measure=UnitOfMeasure.EACH.value,
                drawing_number=assembly_drawing,
                created_by=current_user.id,
                company_id=company_id,
            )
            db.add(assembly_part)
            db.flush()
            audit.log_create(
                "part",
                assembly_part.id,
                assembly_part.part_number,
                new_values=assembly_part,
                extra_data={"source": "bom_import"},
            )

        created_parts = 0 if existing_part else 1
        created_bom_items = 0
        bom_id: Optional[int] = None

        if doc_type == "part" and not items:
            db.commit()
            return BOMImportResponse(
                document_type="part",
                assembly_part_id=assembly_part.id,
                assembly_part_number=assembly_part.part_number,
                bom_id=None,
                created_parts=created_parts,
                created_bom_items=0,
                extraction_confidence=extracted.get("extraction_confidence", "low"),
                warnings=warnings,
            )

        # If any BOM row already occupies the assembly part, block the import.
        _reject_existing_bom(db, assembly_part, company_id)

        bom = BOM(
            part_id=assembly_part.id,
            revision=assembly_revision or "A",
            description=assembly_description,
            status="draft",
            bom_type="standard",
            created_by=current_user.id,
            company_id=company_id,
        )
        db.add(bom)
        _flush_new_bom(db, assembly_part.part_number, retry="retrying the import")
        bom_id = bom.id

        next_line = 10
        component_part_numbers: List[str] = []
        for idx, item in enumerate(items, start=1):
            item_number = int(item.get("line_number") or next_line)
            next_line = item_number + 10
            description = (item.get("description") or "").strip()
            item_part_number = _safe_part_number(item.get("part_number"))
            line_type = _classify_line_type(description, item.get("line_type"))
            item_type = _coerce_item_type(item.get("item_type"))
            part_type = _infer_part_type(line_type, item_type, description)
            uom = item.get("unit_of_measure")

            part_name = description or item_part_number or f"Item {item_number}"
            if not item_part_number:
                warnings.append(f"Line {item_number}: missing part number; generated automatically.")

            component_part, missing, was_created = _ensure_part(
                db,
                item_part_number,
                part_name,
                description,
                part_type,
                None,
                uom,
                create_missing_parts,
                idx,
                company_id=company_id,
                audit=audit,
                created_by=current_user.id,
            )
            if missing:
                missing_parts.append(missing)
                continue

            if was_created:
                created_parts += 1
            component_part_numbers.append(component_part.part_number)

            quantity = float(item.get("quantity") or 1)
            bom_item = BOMItem(
                bom_id=bom.id,
                component_part_id=component_part.id,
                item_number=item_number,
                quantity=quantity if quantity > 0 else 1.0,
                item_type=item_type,
                line_type=line_type,
                # BOM-LINE WRITE PATH 2 of 4 (one-shot upload+commit importer). Same rule
                # as path 1: an unstated unit inherits the component part's.
                unit_of_measure=_resolve_line_uom(uom, component_part, normalize_stated=True),
                reference_designator=item.get("reference_designator"),
                find_number=item.get("find_number"),
                notes=item.get("notes"),
                company_id=company_id,
            )
            db.add(bom_item)
            created_bom_items += 1

        if missing_parts:
            db.rollback()
            raise HTTPException(status_code=400, detail=f"Missing parts: {', '.join(missing_parts)}")

        # One audit row for the BOM with the items summarized (house pattern:
        # the WO import logs the parent and summarizes children in extra_data),
        # before the terminal commit so it persists atomically with the import.
        #
        # ``bom_identifier``, NOT the bare part number. Both importers used to write
        # ``"WRC-1001"`` here while the six header verbs write ``"WRC-1001 BOM rev A"`` for
        # the SAME ``resource_type="bom"`` chain -- so an import-born BOM's CREATE row was
        # the one row about that document that did not match the shape of every row that
        # follows it, and the revision (which the importer knows: it just wrote it) appeared
        # nowhere on the row. The bare number also collides visually with the
        # ``resource_type="part"`` rows this same import writes.
        #
        # Forward-only, deliberately: audit rows are immutable and prod already holds bare-
        # shaped rows, so this is a documented discontinuity rather than a rewritten history
        # -- the house "correct-forward, never backfill" posture, and invariant 2 forbids the
        # alternative outright. The discontinuity is one-directional and mild: the new form
        # CONTAINS the old one, and the audit search filters on
        # ``resource_identifier.ilike("%q%")`` (``audit.py``), so an auditor searching the
        # part number still matches rows of both shapes. Only a search for the full new
        # string misses the pre-existing ones.
        audit.log_create(
            "bom",
            bom.id,
            bom_identifier(assembly_part.part_number, bom.revision),
            new_values=bom,
            extra_data={
                "source": "bom_import",
                "item_count": created_bom_items,
                "component_part_numbers": component_part_numbers,
            },
        )

        db.commit()

        return BOMImportResponse(
            document_type="bom",
            assembly_part_id=assembly_part.id,
            assembly_part_number=assembly_part.part_number,
            bom_id=bom_id,
            created_parts=created_parts,
            created_bom_items=created_bom_items,
            extraction_confidence=extracted.get("extraction_confidence", "low"),
            warnings=warnings,
        )

    try:
        return await run_in_threadpool(_run_import)
    except ImportFileError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/", response_model=List[BOMResponse])
def list_boms(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=10000),
    status: Optional[str] = None,
    active_only: bool = True,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """List all BOMs"""
    # Use selectinload to avoid N+1 queries for parts and items
    query = (
        db.query(BOM)
        .filter(
            BOM.company_id == company_id,
            # Invariant 3: ``BOM`` carries ``SoftDeleteMixin`` and ``delete_bom`` is now a
            # SOFT delete, so every read path has to say so or a deleted BOM keeps showing
            # up in the list exactly as it did before it was deleted.
            BOM.is_deleted == False,  # noqa: E712
        )
        .options(selectinload(BOM.items))
    )

    if active_only:
        query = query.filter(BOM.is_active == True)

    if status:
        query = query.filter(BOM.status == status)

    boms = query.offset(skip).limit(limit).all()

    # Preload BOM existence for component parts to avoid per-item queries
    component_ids = {item.component_part_id for bom in boms for item in (bom.items or []) if item.component_part_id}
    has_bom_by_part_id = {pid: True for pid in parts_with_active_bom(db, component_ids, company_id)}
    # One TENANT-SCOPED read for every component on the page. This replaced
    # ``selectinload(BOMItem.component_part)``: that relationship has no ``company_id``
    # predicate, so on a mis-parented line it rendered a FOREIGN part's number and name
    # into this tenant's list response. See ``tenant_parts_by_id``.
    components_by_id = tenant_parts_by_id(db, component_ids, company_id)
    # And the same read for the HEADERS' parent parts, which came off the equally unscoped
    # ``selectinload(BOM.part)``: a mis-parented header rendered a foreign part number and
    # name as the assembly this BOM builds.
    parents_by_id = tenant_parts_by_id(db, {bom.part_id for bom in boms}, company_id)

    result = []
    for bom in boms:
        try:
            # Resolved TENANT-SCOPED above, never off ``bom.part``.
            part = parents_by_id.get(bom.part_id)

            # Build part info safely
            part_info = None
            if part:
                part_info = PartInfo(
                    id=part.id,
                    part_number=part.part_number or "",
                    name=part.name or "",
                    revision=part.revision or "A",
                    part_type=part.part_type.value if part.part_type else "manufactured",
                )

            # Items are already loaded via selectinload
            items = bom.items or []
            items_list = []
            for item in items:
                try:
                    # Resolved TENANT-SCOPED above, never off ``item.component_part``.
                    component = components_by_id.get(item.component_part_id)

                    component_info = None
                    if component:
                        # ``has_bom`` from the ONE batched probe above -- a page of BOMs
                        # must not become one probe per line.
                        component_info = component_part_info(
                            component, has_bom=has_bom_by_part_id.get(component.id, False)
                        )

                    items_list.append(
                        BOMItemResponse(
                            id=item.id,
                            bom_id=item.bom_id,
                            component_part_id=item.component_part_id,
                            item_number=item.item_number if item.item_number is not None else 10,
                            quantity=item.quantity if item.quantity is not None else 1.0,
                            item_type=item.item_type if item.item_type else BOMItemType.MAKE,
                            line_type=item.line_type if item.line_type else BOMLineType.COMPONENT,
                            unit_of_measure=item.unit_of_measure or "each",
                            reference_designator=item.reference_designator,
                            find_number=item.find_number,
                            notes=item.notes,
                            torque_spec=item.torque_spec,
                            installation_notes=item.installation_notes,
                            work_center_id=item.work_center_id,
                            operation_sequence=item.operation_sequence if item.operation_sequence is not None else 10,
                            scrap_factor=item.scrap_factor if item.scrap_factor is not None else 0.0,
                            lead_time_offset=item.lead_time_offset if item.lead_time_offset is not None else 0,
                            is_optional=item.is_optional if item.is_optional is not None else False,
                            is_alternate=item.is_alternate if item.is_alternate is not None else False,
                            alternate_group=item.alternate_group,
                            component_part=component_info,
                            created_at=item.created_at,
                            updated_at=item.updated_at,
                        )
                    )
                except Exception:
                    pass  # Skip items that fail

            bom_response = BOMResponse(
                id=bom.id,
                part_id=bom.part_id,
                revision=bom.revision or "A",
                description=bom.description or "",
                bom_type=bom.bom_type or "standard",
                status=bom.status or "draft",
                is_active=bom.is_active if bom.is_active is not None else True,
                effective_date=bom.effective_date,
                created_at=bom.created_at,
                updated_at=bom.updated_at,
                part=part_info,
                items=items_list,
            )
            result.append(bom_response)
        except Exception:
            pass  # Skip BOMs that fail

    return result


@router.post("/", response_model=BOMResponse)
def create_bom(
    bom_in: BOMCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Create a new BOM for a part.

    Audited as ``bom`` (CREATE) for the header **plus** one ``bom_line`` (CREATE) per inline
    line. Recording a line added later while not recording the lines the document was born
    with is not a record -- this is BOM-line write path 3 of 4, and PR #194 audited the
    other three.
    """
    # Check if part exists
    part = db.query(Part).filter(Part.id == bom_in.part_id, Part.company_id == company_id).first()
    if not part:
        raise HTTPException(status_code=404, detail="Part not found")

    # Check if ANY BOM row already occupies this part -- not just an active one.
    # ``BOM.part_id`` is UNIQUE with no soft-delete/active carve-out, so the old
    # ``is_active == True`` probe could not see a soft-deleted (or merely deactivated) row
    # that still owns the slot: it passed its own guard and then died on the constraint
    # with an IntegrityError 500. Now that ``delete_bom`` soft-deletes, that is the
    # ordinary path, not a corner. Scoped to the active company, which the old probe also
    # was not.
    existing = db.query(BOM).filter(BOM.part_id == bom_in.part_id, BOM.company_id == company_id).first()
    if existing:
        raise HTTPException(
            status_code=400,
            detail=_existing_bom_conflict_detail(
                part.part_number,
                bom_is_deleted=bool(existing.is_deleted),
                bom_is_active=bool(existing.is_active),
                retry="creating a new one",
            ),
        )

    # Create BOM
    bom = BOM(
        part_id=bom_in.part_id,
        revision=bom_in.revision,
        description=bom_in.description,
        bom_type=bom_in.bom_type,
        created_by=current_user.id,
    )
    bom.company_id = company_id
    db.add(bom)
    # Last line of defence behind the probe above -- see ``_flush_new_bom``.
    _flush_new_bom(db, part.part_number, retry="creating it")

    # (line, component) pairs, kept so the per-line audit rows below can name the component
    # by the part number this request already resolved TENANT-SCOPED.
    created_lines: List[Tuple[BOMItem, Part]] = []

    # Add items
    for item_data in bom_in.items:
        # Validate component part exists IN THIS COMPANY (invariant #1). Unscoped, this
        # resolved another tenant's Part -- which since the UoM change is not merely a
        # disclosure through ``build_bom_item_response`` but a WRITE: ``_resolve_line_uom``
        # would stamp that foreign part's stocking unit onto this tenant's bom_items row.
        component = db.query(Part).filter(Part.id == item_data.component_part_id, Part.company_id == company_id).first()
        if not component:
            raise HTTPException(status_code=400, detail=f"Component part ID {item_data.component_part_id} not found")

        # Same ownership rule as the single-add path: this is BOM-line write path 3 of 4
        # and it splats ``model_dump()`` straight into ``BOMItem``, so an unchecked
        # ``work_center_id`` here reopens the hole the other paths close.
        _assert_work_center_owned(db, item_data.work_center_id, company_id)

        # Check for circular reference
        if item_data.component_part_id == bom_in.part_id:
            raise HTTPException(status_code=400, detail="BOM cannot contain itself as a component")

        # BOM-LINE WRITE PATH 3 of 4 (BOM created with its lines in one call). The schema
        # no longer supplies a literal "each", so an omitted unit arrives as ``None`` here
        # and inherits the component part's. Resolved BEFORE the splat, because the splat
        # would otherwise hand ``BOMItem`` an explicit ``None`` and leave whether the
        # column default fires up to SQLAlchemy's insert-time treatment of it.
        line_values = item_data.model_dump()
        line_values["unit_of_measure"] = _resolve_line_uom(line_values.get("unit_of_measure"), component)

        item = BOMItem(bom_id=bom.id, company_id=company_id, **line_values)
        db.add(item)
        created_lines.append((item, component))

    db.flush()  # assigns every item.id for the audit rows below

    # Resolved ONCE, after every line is in place, so the walk sees the finished document
    # rather than a prefix of it -- and so a 40-line assembly does not pay for 40 walks.
    # Same ``extra_data`` shape as the three line verbs, which is what makes an arming
    # verdict and the edits around it correlatable on one chain (see ``_armed_extra_data``).
    armed = armed_parts_affected_by_bom(db, bom, company_id=company_id)
    armed_extra = _armed_extra_data(armed)
    identifier = bom_identifier(part.part_number, bom.revision)

    # BEFORE the terminal commit. AuditService.log() only FLUSHES, and the request session
    # never commits on teardown -- a call placed after db.commit() opens a fresh transaction
    # that get_db teardown rolls back, silently discarding the row. Same rule and same
    # comment as ``add_bom_item``.
    audit.log_create("bom", bom.id, identifier, new_values=bom, extra_data=armed_extra)
    for item, component in created_lines:
        audit.log_create(
            "bom_line",
            item.id,
            bom_line_identifier(part.part_number, item.item_number, component.part_number),
            new_values=item,
            extra_data=armed_extra,
        )

    db.commit()
    db.refresh(bom)

    # Return with full response
    return get_bom(bom.id, db, current_user, company_id)


# How many CANDIDATE rows the mismatch scan will pull before it stops and says so. The
# SQL predicate has already narrowed to disagreeing lines by the time this bites, so a
# tenant would need thousands of genuinely-wrong lines to reach it; the ceiling exists so
# that a pathological BOM set degrades into a truthful "there are at least this many"
# rather than into an unbounded query behind a synchronous request.
_UOM_MISMATCH_SCAN_CEILING = 5000


# DECLARED BEFORE ``@router.get("/{bom_id}")`` ON PURPOSE. FastAPI matches routes in
# declaration order, and ``/{bom_id}`` is a single-segment path parameter that would
# happily swallow the literal ``/uom-mismatches`` and 422 on the int conversion. Moving
# this below that route silently breaks the endpoint. Keep it here.
@router.get(
    "/uom-mismatches",
    response_model=BOMUomMismatchReport,
    summary="BOM lines whose stated unit of measure disagrees with the component part's",
)
def list_bom_uom_mismatches(
    part_id: Optional[int] = Query(None, description="Only lines on this assembly part's own BOM(s)"),
    bom_id: Optional[int] = Query(None, description="Only lines on this BOM"),
    component_part_id: Optional[int] = Query(None, description="Only lines naming this component part"),
    active_only: bool = Query(True, description="Only active BOMs — the ones a backflush actually reads"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
):
    """Every BOM line in this company whose stated unit contradicts its component part's.

    **This is the gate on arming a real part for automatic backflush.** The BLOCKING
    ``unit_of_measure_mismatch`` diagnostic refuses ``Part.backflush_components`` at opt-in
    and refuses that component at completion, and nothing in the platform converts units —
    a line stating ``each`` against a part stocked in ``sheets`` would issue the wrong
    quantity of the right material. New lines now inherit the component's unit
    (``_resolve_line_uom``), but this series is **correct-forward and does not backfill**:
    lines written before that change keep what they have. This endpoint is how a human
    finds and corrects them, deliberately, one at a time.

    **Pure read — writes nothing.** No ledger row, no audit row, no operational event.

    The comparison is ``models.part.uom_disagrees``, the SAME predicate the diagnostic
    uses, so the two cannot list different rows. In particular ``ea`` does NOT satisfy
    ``each`` here — that is not a bug to fix by teaching this synonyms, it is a stored
    value a human should normalise, because the gate will keep refusing it either way.

    Scope notes, so the numbers are read correctly:

    * ``part_id`` narrows to lines on that assembly's OWN BOM. It does NOT follow nested
      sub-assembly BOMs, which a readiness check for that part DOES reach. **The
      unfiltered report is the authoritative pre-arming worklist**; the filter is for
      working one assembly at a time.
    * ``blocks_backflush`` is False on alternate / optional / reference lines — the
      backflush never issues those, so they raise no diagnostic and refuse nothing.
    * Soft-deleted component parts are INCLUDED (with ``component_is_deleted``), because
      the readiness explosion resolves them on purpose; filtering them out here would hide
      a row that still blocks.

    Gated to ADMIN / MANAGER / SUPERVISOR: it is a remediation worklist, and that is
    exactly the set of roles that can edit a BOM line (``PUT /bom/items/{id}``) or arm the
    flag (``PUT /parts/{id}``). Handing it to someone who cannot act on it buys nothing.
    """
    assembly_part = aliased(Part)
    component_part = aliased(Part)

    # SQL-side narrowing only. It reproduces ``uom_disagrees`` closely enough to be a cheap
    # pre-filter, but it is NOT the authority: ``Part.unit_of_measure`` is a native enum
    # (hence the cast — Postgres will not apply string functions to it), and SQL ``trim``
    # strips spaces where Python's ``strip`` strips all whitespace. That gap can only ever
    # let a row THROUGH to the Python check below, never hide one from it, so the
    # authoritative predicate runs in Python on every candidate.
    line_label = func.lower(func.trim(func.coalesce(BOMItem.unit_of_measure, "")))
    part_label = func.lower(func.trim(func.coalesce(cast(component_part.unit_of_measure, String), "")))

    query = (
        db.query(BOMItem, BOM, assembly_part, component_part)
        .join(BOM, BOM.id == BOMItem.bom_id)
        .join(assembly_part, assembly_part.id == BOM.part_id)
        .join(component_part, component_part.id == BOMItem.component_part_id)
        .filter(
            # Tenant isolation on every table the row is assembled from. ``BOMItem`` is
            # scoped explicitly rather than inherited through its BOM because the backflush
            # explosion scopes it the same way (``_explode_backflush_bom``), and a report
            # that walked a wider set than the gate does would list rows nobody can act on.
            BOM.company_id == company_id,
            BOMItem.company_id == company_id,
            assembly_part.company_id == company_id,
            component_part.company_id == company_id,
            # A soft-deleted BOM is not a remediation target: nothing explodes it any more
            # (invariant 3), so a line on one blocks nothing and listing it would send a
            # human to fix a document the shop has deleted.
            BOM.is_deleted == False,  # noqa: E712
            line_label != "",
            part_label != "",
            line_label != part_label,
        )
    )

    if active_only:
        query = query.filter(BOM.is_active == True)  # noqa: E712
    if bom_id is not None:
        query = query.filter(BOM.id == bom_id)
    if part_id is not None:
        query = query.filter(BOM.part_id == part_id)
    if component_part_id is not None:
        query = query.filter(BOMItem.component_part_id == component_part_id)

    candidates = (
        query.order_by(
            assembly_part.part_number.asc(),
            BOM.id.asc(),
            BOMItem.item_number.asc(),
            BOMItem.id.asc(),
        )
        .limit(_UOM_MISMATCH_SCAN_CEILING + 1)
        .all()
    )
    truncated = len(candidates) > _UOM_MISMATCH_SCAN_CEILING
    candidates = candidates[:_UOM_MISMATCH_SCAN_CEILING]

    # Two passes on purpose. ``uom_disagrees`` is the AUTHORITY (the SQL predicate above is
    # only a narrowing pre-filter), so an accurate ``total`` genuinely requires walking every
    # candidate -- but building a response model for each one does not. A tenant mid-
    # remediation can hold thousands of legacy ``each`` lines, and constructing all of them to
    # return a page of 100 wasted the work on every page the client asked for.
    passing = [row for row in candidates if uom_disagrees(row[0].unit_of_measure, row[3].unit_of_measure)]

    rows: List[BOMLineUomMismatch] = []
    for item, bom, assembly, component in passing[skip : skip + limit]:
        rows.append(
            BOMLineUomMismatch(
                bom_id=bom.id,
                bom_revision=bom.revision,
                bom_status=bom.status,
                bom_is_active=bool(bom.is_active),
                part_id=assembly.id,
                part_number=assembly.part_number or "",
                bom_item_id=item.id,
                item_number=item.item_number,
                component_part_id=component.id,
                component_part_number=component.part_number or "",
                component_part_name=component.name,
                component_is_deleted=bool(getattr(component, "is_deleted", False)),
                # The NORMALISED labels the comparison actually used, not the raw column
                # text: what the row shows has to be what the gate compared.
                line_unit_of_measure=uom_label(item.unit_of_measure),
                component_unit_of_measure=uom_label(component.unit_of_measure),
                blocks_backflush=bom_line_is_backflush_consumed(item),
            )
        )

    return BOMUomMismatchReport(total=len(passing), returned=len(rows), truncated=truncated, items=rows)


@router.get("/{bom_id}", response_model=BOMResponse)
def get_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """Get a specific BOM with all items"""
    try:
        bom = (
            db.query(BOM)
            .options(joinedload(BOM.items))
            .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
            .first()
        )

        if not bom:
            raise HTTPException(status_code=404, detail="BOM not found")

        # The parent part TENANT-SCOPED, never ``bom.part`` -- that relationship carries no
        # ``company_id`` predicate, so on a mis-parented header it rendered another
        # company's part number and name as the assembly. See ``tenant_parts_by_id``.
        parent_part = tenant_parts_by_id(db, [bom.part_id], company_id).get(bom.part_id)

        # Safely get part_type - handle both enum and string values
        part_type_val = "manufactured"
        if parent_part and parent_part.part_type:
            if hasattr(parent_part.part_type, 'value'):
                part_type_val = parent_part.part_type.value
            else:
                part_type_val = str(parent_part.part_type)

        component_ids = {item.component_part_id for item in bom.items if item.component_part_id}
        has_bom_by_part_id = {pid: True for pid in parts_with_active_bom(db, component_ids, company_id)}
        # TENANT-SCOPED component resolution, batched, replacing
        # ``joinedload(BOMItem.component_part)`` -- see ``tenant_parts_by_id``.
        components_by_id = tenant_parts_by_id(db, component_ids, company_id)

        return BOMResponse(
            id=bom.id,
            part_id=bom.part_id,
            revision=bom.revision,
            description=bom.description,
            bom_type=bom.bom_type or "standard",
            status=bom.status,
            is_active=bom.is_active,
            effective_date=bom.effective_date,
            created_at=bom.created_at,
            updated_at=bom.updated_at,
            part=(
                PartInfo(
                    id=parent_part.id,
                    part_number=parent_part.part_number or "",
                    name=parent_part.name or "",
                    revision=parent_part.revision or "A",
                    part_type=part_type_val,
                )
                if parent_part
                else None
            ),
            items=[
                build_bom_item_response(
                    item, db, has_bom_by_part_id, company_id=company_id, components_by_id=components_by_id
                )
                for item in bom.items
            ],
        )
    except HTTPException:
        raise
    except Exception:
        # The traceback goes to the log (and Sentry, via main.py), never to the client:
        # a 500 body is readable by every authenticated caller and a formatted traceback
        # discloses absolute paths, ORM internals and library versions. See the guard at
        # tests/test_no_traceback_in_error_response_guard.py.
        logger.exception("Error getting BOM %s", bom_id)
        raise HTTPException(status_code=500, detail="Error getting BOM")


@router.get("/by-part/{part_id}", response_model=BOMResponse)
def get_bom_by_part(
    part_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """Get the active BOM for a part"""
    bom = (
        db.query(BOM)
        .filter(
            BOM.part_id == part_id,
            BOM.company_id == company_id,
            BOM.is_active == True,  # noqa: E712
            BOM.is_deleted == False,  # noqa: E712
        )
        .first()
    )

    if not bom:
        raise HTTPException(status_code=404, detail="No active BOM found for this part")

    return get_bom(bom.id, db, current_user, company_id)


@router.put("/{bom_id}", response_model=BOMResponse)
def update_bom(
    bom_id: int,
    bom_in: BOMUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Update a BOM header.

    On a DRAFT BOM every field in ``BOMUpdate`` is editable. On a RELEASED BOM only
    ``description`` is -- ``revision`` / ``bom_type`` / ``effective_date`` are the approved
    document's identity, its explosion semantics and its AS9100D effectivity, and changing
    any of them in place would relabel or re-date an approved configuration without
    re-approval. Unrelease the BOM to change those. Any other status is fully locked.

    ``status`` is NOT in ``BOMUpdate`` and cannot be written here at all -- both transitions
    belong to ``release_bom`` / ``unrelease_bom``, which are Admin/Manager and stamp (or
    clear) the approval evidence. Before that field was deleted a SUPERVISOR could
    ``PUT {"status": "released"}`` and produce a released controlled document with
    ``approved_by``/``approved_at`` NULL -- an approved document with no approver -- while
    bypassing both the Admin/Manager gate and the "no items" precondition on the release
    verb. An unset field on the request body is ignored (Pydantic ``extra="ignore"``), so a
    legacy client still gets 200 with the true status echoed back in ``BOMResponse``.

    A released ``description`` edit deliberately does NOT re-stamp ``approved_by`` /
    ``approved_at`` -- a divergence from ``routing.update_operation``, which re-stamps
    because its carved-out fields ARE the released production content. A BOM description is
    metadata about the document, not the configuration; re-stamping it would overwrite the
    record of who approved the actual configuration in order to record a typo fix.

    Audited as ``bom`` (UPDATE), and only when something actually changed.
    """
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    update_data = bom_in.model_dump(exclude_unset=True)

    # Only fields present in the payload AND different from the current value count as
    # changes (semantics copied from ``routing.update_operation``). Without this, a
    # form-shaped client that PUTs the whole record back would be refused for changing
    # nothing.
    #
    # The comparison is only sound while both sides are of the same shape, and
    # ``effective_date`` is the one field where that took work: the column is a NAIVE
    # ``DateTime`` while ``BOMResponse`` (a ``UTCModel``) serves the value with a trailing
    # ``Z``, so a round-tripped payload parsed to an AWARE datetime -- and ``naive !=
    # aware`` is ``True``. The field therefore counted as changed on EVERY no-op PUT, which
    # 400'd released BOMs (the freeze reads it as an edit to the AS9100D effectivity). The
    # chain itself stayed clean only because ``_audit_values`` below normalizes BOTH halves
    # of the diff and ``log_update`` drops an empty one -- do not remove that belt while
    # relying on this. ``BOMUpdate`` now normalizes the field to naive UTC in the schema
    # (``_normalize_to_naive_utc``); do not add a datetime field to that model without the
    # same validator.
    changed_fields = {f for f, v in update_data.items() if getattr(bom, f) != v}

    # Status gate -- evaluated BEFORE the first setattr, so a refusal leaves the row
    # untouched. See ``_assert_bom_lines_editable`` for the same placement rule on lines.
    if changed_fields:
        if bom.status == "released":
            frozen = changed_fields - _BOM_RELEASED_EDITABLE_FIELDS
            if frozen:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        "Released BOM: only the description can be edited — unrelease the BOM "
                        "to change revision, BOM type or effectivity."
                    ),
                )
        elif bom.status != "draft":
            raise HTTPException(status_code=400, detail=_BOM_NOT_DRAFT_MSG.format(status=bom.status))

    # Named as the document was BEFORE the edit, so the identifier matches what an auditor
    # searching the chain for the prior state would look for; the new revision is in
    # ``new_values``.
    identifier = bom_identifier(_tenant_part_number(db, bom.part_id, company_id), bom.revision)
    old_values = _audit_values(bom, changed_fields)

    for field, value in update_data.items():
        setattr(bom, field, value)

    if changed_fields:
        # BEFORE the terminal commit -- see ``add_bom_item`` for why that ordering is
        # load-bearing.
        db.flush()
        audit.log_update(
            "bom",
            bom.id,
            identifier,
            old_values=old_values,
            new_values=_audit_values(bom, changed_fields),
        )

    db.commit()
    db.refresh(bom)
    return get_bom(bom.id, db, current_user, company_id)


@router.post("/{bom_id}/release")
def release_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Release a BOM for production use.

    This is the APPROVAL of a controlled document and it is the ONLY door to
    ``status == "released"``. It was unaudited: a BOM went from draft to approved-for-
    production with nothing on the chain saying who approved it or when the state changed
    (``approved_by`` on the row is the current state, not a record of the transition).
    """
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    if bom.status == "released":
        raise HTTPException(status_code=400, detail="BOM is already released")

    # Only a DRAFT can be released. Previously anything that was not already ``released``
    # fell through -- including the junk statuses the old unvalidated ``BOMUpdate.status``
    # could write, and a terminal ``obsolete``.
    if bom.status != "draft":
        raise HTTPException(status_code=400, detail=f"Only a draft BOM can be released (this BOM is '{bom.status}').")

    if not bom.items:
        raise HTTPException(status_code=400, detail="Cannot release BOM with no items")

    # Read off the row rather than written as a literal -- but the draft-only guard above
    # pins it, so on this verb it is ALWAYS "draft". It is read anyway so the row and the
    # chain cannot disagree if that guard is ever widened, which is exactly what happened to
    # ``unrelease_bom``: the identical expression there IS general (it withdraws any
    # non-draft status back to draft, and the chain has to say which one it was).
    # ``routing.release_routing`` writes the literal; that is the shape not worth copying
    # onto a chain that is supposed to be evidence.
    previous_status = bom.status
    identifier = bom_identifier(_tenant_part_number(db, bom.part_id, company_id), bom.revision)

    bom.status = "released"
    bom.approved_by = current_user.id
    bom.approved_at = datetime.utcnow()
    bom.effective_date = datetime.utcnow()

    # BEFORE the terminal commit -- see ``add_bom_item``.
    audit.log_status_change(
        "bom",
        bom.id,
        identifier,
        old_status=previous_status,
        new_status="released",
        extra_data={"approved_by": current_user.id, "line_count": len(bom.items)},
    )

    db.commit()
    return {"message": "BOM released", "bom_id": bom.id}


@router.post("/{bom_id}/unrelease")
def unrelease_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Unrelease a BOM to allow editing.

    This is the WITHDRAWAL of an approved controlled document, and it DESTROYS the approval
    evidence on the row -- it NULLs ``approved_by`` / ``approved_at`` (and now
    ``effective_date``, which a draft has no business carrying: leaving it is the same
    defect class as leaving a stale approver). It wrote nothing anywhere, so the fact that a
    named approval had ever existed was simply gone. The pre-image goes on the chain in
    ``extra_data`` BEFORE the clear.

    **No free-text reason is required**, deliberately, against the NCR-void / receipt-void /
    vendor-delete precedent. Those are irreversible compensating actions whose "why" is
    unrecoverable from state. This is reversible, it is the routine FIRST STEP of the only
    revision workflow BOMs have (``unrelease -> edit -> bump revision -> release``), and its
    "why" is fully reconstructible from the chain: this row (who withdrew, when, what
    approval was cleared) -> the ``bom_line`` rows written while it was open (what changed)
    -> the re-release row (who re-approved). Forcing a reason on the everyday editing
    gesture yields ``"edit"`` and degrades the record.

    **This is also THE de-corruption door, and that is why it refuses only ``draft``
    rather than requiring ``released``.** A BOM's status column has exactly two legal values
    now, but the unvalidated ``BOMUpdate.status`` that was just deleted could write any
    string, so rows holding junk ("RELEASED", "obsolete", garbage) may exist. Every other
    verb refuses those -- ``release_bom`` wants a draft, ``update_bom`` and the three line
    verbs want a draft, ``delete_bom`` wants a draft -- and ``BOM.part_id`` is UNIQUE, so
    the part could never get a working BOM again: the document would be permanently
    unmanageable, readable and useless. Withdrawing anything that is not already a draft
    BACK to draft is both the correct semantic ("this is not a draft; withdraw it") and the
    one escape hatch, and it is a safe widening because nothing downstream treats a junk
    status as approved -- every consumer tests ``status == "released"`` exactly. The
    transition is audited with the ACTUAL prior status, so a normalisation is visible on the
    chain as exactly what it was.
    """
    try:
        bom = (
            db.query(BOM)
            .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
            .first()
        )
        if not bom:
            raise HTTPException(status_code=404, detail="BOM not found")
        if bom.status == "draft":
            # Message unchanged: for the only status a supported verb can produce besides
            # ``released``, "BOM is not released" is exactly what happened.
            raise HTTPException(status_code=400, detail="BOM is not released")

        previous_status = bom.status

        # BEFORE the clear, and before the terminal commit -- the whole point of this row is
        # the approval pre-image, which does not exist anywhere else once the columns are
        # NULLed.
        audit.log_status_change(
            "bom",
            bom.id,
            bom_identifier(_tenant_part_number(db, bom.part_id, company_id), bom.revision),
            old_status=previous_status,
            new_status="draft",
            extra_data={
                "cleared_approved_by": bom.approved_by,
                "cleared_approved_at": to_utc_iso(bom.approved_at),
                "cleared_effective_date": to_utc_iso(bom.effective_date),
            },
        )

        bom.status = "draft"
        bom.approved_by = None
        bom.approved_at = None
        bom.effective_date = None
        db.commit()
        return {"message": "BOM unreleased", "bom_id": bom.id}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        logger.exception("Error unreleasing BOM %s", bom_id)
        raise HTTPException(status_code=500, detail="Error unreleasing BOM")


@router.delete("/{bom_id}")
def delete_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    r"""Delete a BOM (only draft BOMs can be deleted). The delete is SOFT, and audited.

    ``BOM`` carries ``SoftDeleteMixin``, so the previous ``db.delete(bom)`` was a straight
    invariant-3 violation: it physically destroyed a controlled document's header and, with
    a bulk ``db.query(BOMItem)...delete()``, every one of its lines, leaving no record that
    any of it had ever existed.

    **The lines are KEPT, physically intact.** ``BOMItem`` carries only ``TenantMixin``, so
    a physical delete of them WOULD be legitimate under invariant 3 -- but it is wrong here
    for two independent reasons. (1) The header is a tombstone, not a grave: the lines are
    the document's content, and ``POST /bom/{bom_id}/restore`` can only mean something if
    they survive. Wiping them would make "restore" bring back an empty BOM, which is a
    worse outcome than the bug being fixed. (2) Invariant 5 -- preserve historical records.

    **Nothing reads them while the header is deleted.** That claim is what makes retention
    safe, so it is enumerated rather than asserted. Re-verify it with
    ``grep -rn 'query(BOMItem)\|select(BOMItem)\|join(BOMItem' app/``; every hit is one of:

    * reaches the lines through a ``BOM`` header that filters ``is_deleted == False``
      (``bom.py`` throughout, ``work_orders.py`` x2, ``setup.py`` x3, ``parts.py``,
      ``routing.py``, ``completion_inventory_service.py`` x2, and ``mrp_service.py`` via
      the ``BOM.items`` relationship) -- these are the safe majority; or
    * is one of the TWO deliberate exceptions: the "is this part referenced anywhere"
      probes behind ``hard_delete`` in ``parts.py`` (``DELETE /parts/{id}?hard_delete``)
      and ``materials.py`` (``DELETE /materials/{id}?hard_delete``). They query
      ``BOMItem.component_part_id`` with NO header join, and they must not filter one: a
      retained line still holds a real foreign key to the part, so hard-deleting it would
      fail or orphan the line.

    ``prediction_service._calculate_wo_demand`` used to be an unlisted THIRD direct reader
    (``query(BOMItem).filter(component_part_id == part_id)`` -- no join, no ``company_id``,
    no ``is_deleted``). It now reaches through the header like everything else. Anything
    added to this list that is NOT a hard-delete probe should be presumed a defect.

    So there are no per-line DELETE audit rows -- writing them for rows that still exist
    would be a false record. The count of retained lines goes on the header's DELETE row
    instead.
    """
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    if bom.status != "draft":
        raise HTTPException(status_code=400, detail="Only draft BOMs can be deleted")

    # log_delete serializes the model synchronously, so passing the live attached instance
    # captures the header's full pre-image. Logged BEFORE the terminal commit.
    audit.log_delete(
        "bom",
        bom.id,
        bom_identifier(_tenant_part_number(db, bom.part_id, company_id), bom.revision),
        old_values=bom,
        soft_delete=True,
        # ``previous_is_active`` is recorded because the clear below is unconditional and
        # ``restore_bom`` reinstates ``is_active = True`` unconditionally too. For every row
        # a supported verb can produce that round-trips exactly (this verb and ``restore_bom``
        # are the only two writers of the flag in the backend), but if a legacy or
        # hand-edited inactive row ever exists the chain says what the flag was. It is
        # EVIDENCE only -- ``restore_bom`` deliberately does not read it back to decide what
        # to reinstate; see its docstring for why.
        extra_data={"retained_line_count": len(bom.items), "previous_is_active": bool(bom.is_active)},
    )

    bom.soft_delete(current_user.id)
    # ``is_active`` is what the ten-odd "the active BOM for this part" lookups across the
    # backend key on; clearing it alongside the tombstone means a reader that predates the
    # ``is_deleted`` predicate still cannot resolve this row.
    bom.is_active = False

    db.commit()
    return {"message": "BOM deleted", "can_restore": True}


@router.post("/{bom_id}/restore", summary="Restore a soft-deleted BOM")
def restore_bom(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Restore a soft-deleted BOM.

    **Required roles**: Admin or Manager (the same tier that can delete one).

    This endpoint is not optional garnish on the soft-delete conversion, it is a
    PREREQUISITE for it. ``BOM.part_id`` is UNIQUE with no soft-delete carve-out, so a
    soft-deleted BOM permanently occupies its part's only BOM slot: without a restore verb,
    deleting a BOM would mean that part can never have a BOM again. Mirrors
    ``parts.restore_part``, which ``_reject_existing_bom`` and ``create_bom`` already point
    callers at by name.

    Restoring cannot collide with anything: the unique constraint guarantees no second BOM
    row took the slot while this one was deleted. The BOM comes back as a DRAFT -- only
    drafts can be deleted, so that is the state it left in, and no approval is resurrected.

    ``is_active`` comes back TRUE unconditionally, which is the exact inverse of
    ``delete_bom`` clearing it unconditionally. That is a faithful round trip for every row
    the API can produce: ``delete_bom`` and this verb are the ONLY two writers of
    ``BOM.is_active`` in the backend -- no create path sets it (the column default is TRUE),
    ``BOMUpdate`` does not carry the field, and neither importer touches it -- so a BOM that
    is deleted was active.

    **It deliberately does NOT read ``previous_is_active`` back off the DELETE audit row**,
    even though ``delete_bom`` records it. Two reasons, and the first is the general one:
    the audit chain is a RECORD, not a source of truth. Driving a business decision off it
    inverts that -- this verb's behaviour would then depend on the ``extra_data`` shape of a
    row nothing enforces a schema on, and on that row still being resolvable (aged segments
    are archived to cold storage, and a partition drop is a documented DBA option; see
    ``docs/AUDIT_LOG_RETENTION_RUNBOOK.md``), so the same BOM could restore differently
    depending on how old it is. Second, there is nothing to read: per the paragraph above,
    the value is TRUE for every row a supported verb can produce, so the branch would be
    dead code guarding an unreachable state. ``previous_is_active`` stays on the delete row
    as EVIDENCE for a legacy or hand-edited row, which is the right job for an audit row.

    If BOM deactivation ever becomes a real verb ("obsolete this BOM without deleting it"),
    the prior value belongs on the ``boms`` row itself -- a column, set by that verb and
    consulted here -- not in the chain. Adding that column now, for a state nothing can
    reach, would be a migration over live multi-tenant data buying nothing.
    """
    bom = db.query(BOM).filter(BOM.id == bom_id, BOM.company_id == company_id).first()
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    if not bom.is_deleted:
        raise HTTPException(status_code=400, detail="BOM is not deleted")

    bom.restore()
    bom.is_active = True

    # BEFORE the terminal commit. ``action="restore"`` is the same verb ``restore_part``
    # uses, and log_update's empty-diff self-suppression does not apply to it.
    audit.log_update(
        "bom",
        bom.id,
        bom_identifier(_tenant_part_number(db, bom.part_id, company_id), bom.revision),
        old_values={"is_deleted": True, "is_active": False},
        new_values={"is_deleted": False, "is_active": True},
        action="restore",
    )

    db.commit()
    return {"message": "BOM restored", "bom_id": bom.id}


# BOM Item operations
@router.post("/{bom_id}/items", response_model=BOMItemResponse)
def add_bom_item(
    bom_id: int,
    item_in: BOMItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Add an item to a BOM.

    Audited as ``bom_line`` (CREATE). BOM-line writes carried NO audit row of any kind
    before this -- an AS9100D records gap on the controlled document the shop builds parts
    from, and the stated reason the UoM-mismatch worklist shipped read-only.
    """
    try:
        bom = (
            db.query(BOM)
            .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
            .first()
        )
        if not bom:
            raise HTTPException(status_code=404, detail="BOM not found")

        # Parent 404 FIRST, then the status gate: a foreign ``bom_id`` must not be
        # distinguishable from a released one by the shape of the refusal.
        _assert_bom_lines_editable(bom)

        # Validate component exists IN THIS COMPANY (invariant #1) -- see the matching
        # scoping in ``create_bom``. A 404 rather than a 403 so a foreign id cannot be probed.
        component = db.query(Part).filter(Part.id == item_in.component_part_id, Part.company_id == company_id).first()
        if not component:
            raise HTTPException(status_code=404, detail="Component part not found")

        # Same rule for the OPTIONAL work center (invariant #1) -- BOM-line write path 4 of 4.
        _assert_work_center_owned(db, item_in.work_center_id, company_id)

        # Check for circular reference
        if item_in.component_part_id == bom.part_id:
            raise HTTPException(status_code=400, detail="BOM cannot contain itself")

        # Check for deeper circular references
        if would_create_circular_reference(db, bom.part_id, item_in.component_part_id, company_id):
            raise HTTPException(
                status_code=400, detail="Adding this component would create a circular reference in the BOM structure"
            )

        # Inherit customer_name from parent assembly if component doesn't have one.
        # Tenant-scoped for the same reason: this line WRITES onto ``component``, so an
        # unscoped pair here was a cross-tenant write, not just a read.
        parent_part = db.query(Part).filter(Part.id == bom.part_id, Part.company_id == company_id).first()
        if parent_part and parent_part.customer_name and not component.customer_name:
            component.customer_name = parent_part.customer_name

        # Get item data and ensure enum values are lowercase for PostgreSQL
        item_data = item_in.model_dump()

        # Convert item_type to lowercase string
        if 'item_type' in item_data and item_data['item_type']:
            val = item_data['item_type']
            if hasattr(val, 'value'):
                item_data['item_type'] = val.value.lower()
            elif isinstance(val, str):
                item_data['item_type'] = val.lower()

        # Convert line_type to lowercase string
        if 'line_type' in item_data and item_data['line_type']:
            val = item_data['line_type']
            if hasattr(val, 'value'):
                item_data['line_type'] = val.value.lower()
            elif isinstance(val, str):
                item_data['line_type'] = val.lower()

        # BOM-LINE WRITE PATH 4 of 4 (single add — the one the BOM page and the part BOM
        # tab actually use; NEITHER of them sends a unit at all). An unstated unit inherits
        # the component part's rather than the literal "each" the schema used to supply.
        item_data['unit_of_measure'] = _resolve_line_uom(item_data.get('unit_of_measure'), component)

        item = BOMItem(bom_id=bom_id, company_id=company_id, **item_data)
        db.add(item)
        db.flush()  # assigns item.id for the audit row below

        armed = armed_parts_affected_by_bom(db, bom, company_id=company_id)
        warning = backflush_armed_edit_warning(armed, item_number=item.item_number)

        # Logged BEFORE the terminal commit so the CREATE row commits atomically with the
        # line. AuditService.log() only FLUSHES, and the request session never commits on
        # teardown -- an audit call placed after db.commit() opens a fresh transaction that
        # get_db teardown rolls back, silently discarding the row. Same rule and same
        # comment as routing.py's add_operation.
        audit.log_create(
            "bom_line",
            item.id,
            bom_line_identifier(
                parent_part.part_number if parent_part else None, item.item_number, component.part_number
            ),
            new_values=item,
            extra_data=_armed_extra_data(armed),
        )

        db.commit()
        db.refresh(item)

        # Build response manually to avoid joinedload issues. Single-item path, so the
        # probing helper is the right one -- one component, one probe.
        component_info = get_component_part_info(component, db, company_id) if component else None

        return BOMItemResponse(
            id=item.id,
            bom_id=item.bom_id,
            component_part_id=item.component_part_id,
            item_number=item.item_number,
            quantity=item.quantity,
            item_type=item.item_type,
            line_type=item.line_type,
            unit_of_measure=item.unit_of_measure or "each",
            reference_designator=item.reference_designator,
            find_number=item.find_number,
            notes=item.notes,
            torque_spec=item.torque_spec,
            installation_notes=item.installation_notes,
            work_center_id=item.work_center_id,
            operation_sequence=item.operation_sequence or 10,
            scrap_factor=item.scrap_factor or 0.0,
            lead_time_offset=item.lead_time_offset or 0,
            is_optional=item.is_optional or False,
            is_alternate=item.is_alternate or False,
            alternate_group=item.alternate_group,
            component_part=component_info,
            backflush_armed_warning=warning,
            created_at=item.created_at,
            updated_at=item.updated_at,
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Error adding BOM item to BOM %s", bom_id)
        raise HTTPException(status_code=500, detail="Error adding BOM item")


@router.put("/items/{item_id}", response_model=BOMItemResponse)
def update_bom_item(
    item_id: int,
    item_in: BOMItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Update a BOM item.

    A request that does not mention ``unit_of_measure`` leaves it alone (``exclude_unset``)
    — this endpoint is NOT a backfill, and existing lines keep what they have. A request
    that CLEARS it (explicit ``null`` / blank) is treated as "no stated unit" and resolves
    to the component part's, the same rule the four create paths use, rather than writing a
    NULL nobody asked for.

    Audited as ``bom_line`` (UPDATE). ``log_update`` self-suppresses on an empty diff, so an
    idempotent PUT that restates the current values writes no row -- the same behaviour
    ``update_work_center`` relies on.
    """
    item = (
        db.query(BOMItem)
        .join(BOM)
        .filter(BOMItem.id == item_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")

    bom = (
        db.query(BOM)
        .filter(BOM.id == item.bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:  # pragma: no cover - the join above already proved it
        raise HTTPException(status_code=404, detail="BOM item not found")

    # The scoped join above IS the tenant proof, so the parent status gate comes next --
    # still before the first setattr.
    _assert_bom_lines_editable(bom)

    # Hoisted out of the UoM branch below: the identifier, the UoM resolution and the
    # cross-tenant guard all need the TENANT-SCOPED component, and one indexed read serves
    # all three. Never ``item.component_part`` -- that relationship carries no company_id.
    component = db.query(Part).filter(Part.id == item.component_part_id, Part.company_id == company_id).first()

    update_data = item_in.model_dump(exclude_unset=True)

    # Same ownership check the two create paths make: ``work_center_id`` is settable here
    # too, so without this the update path reopens the hole they close.
    _assert_work_center_owned(db, update_data.get("work_center_id"), company_id)

    if "unit_of_measure" in update_data:
        # Defense-in-depth (invariant #1): resolve a cleared UoM through the same
        # TENANT-SCOPED Part lookup the other three UoM-resolution sites use — never
        # the unscoped ``component_part`` relationship, which on a mis-parented line
        # would read (and store) a FOREIGN part's stocking unit.
        stated = update_data["unit_of_measure"]
        if component is None and not (stated is not None and str(stated).strip()):
            # Cleared value + unresolvable (foreign/missing) component: NO inheritance.
            # Keep the line's existing value rather than manufacturing one from nothing.
            update_data.pop("unit_of_measure")
        else:
            update_data["unit_of_measure"] = _resolve_line_uom(stated, component)

    # Full-row snapshot BEFORE the setattr loop, so the diff is column-true rather than
    # whatever the request happened to name (mirrors ``update_work_center``).
    old_values = {c.key: getattr(item, c.key) for c in item.__table__.columns}

    for field, value in update_data.items():
        setattr(item, field, value)

    armed = armed_parts_affected_by_bom(db, bom, company_id=company_id)
    warning = backflush_armed_edit_warning(armed, item_number=item.item_number)

    # BEFORE the terminal commit -- see add_bom_item for why that ordering is load-bearing.
    db.flush()
    audit.log_update(
        "bom_line",
        item.id,
        bom_line_identifier(
            _tenant_part_number(db, bom.part_id, company_id),
            item.item_number,
            component.part_number if component else None,
        ),
        old_values=old_values,
        new_values=item,
        extra_data=_armed_extra_data(armed),
    )

    db.commit()
    db.refresh(item)
    return build_bom_item_response(item, db, company_id=company_id, backflush_armed_warning=warning)


@router.delete("/items/{item_id}")
def delete_bom_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role([UserRole.ADMIN, UserRole.MANAGER])),
    company_id: int = Depends(get_current_company_id),
    audit: AuditService = Depends(get_audit_service),
):
    """Delete a BOM item. The delete is PHYSICAL, and audited.

    ``BOMItem`` has no ``SoftDeleteMixin`` (only ``TenantMixin``), so there is no tombstone
    to write and the row really goes. Adding the mixin is a schema change -- an Alembic
    revision plus an ``is_deleted == False`` predicate at all eleven ``query(BOMItem)``
    sites across eight modules, any one of which, if missed, silently changes what the
    backflush explodes. That is deliberately NOT bundled into a compliance/security fix;
    the record is instead made complete the way ``routing.py``'s ``delete_operation``
    (same shape: a child line row of a controlled document, no mixin) already does it --
    ``log_delete(..., soft_delete=False)`` with the full pre-image, logged while the
    instance is still attached and BEFORE the delete.
    """
    item = (
        db.query(BOMItem)
        .join(BOM)
        .filter(BOMItem.id == item_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="BOM item not found")

    bom = (
        db.query(BOM)
        .filter(BOM.id == item.bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:  # pragma: no cover - the join above already proved it
        raise HTTPException(status_code=404, detail="BOM item not found")

    # Parent status gate, before the delete is staged.
    _assert_bom_lines_editable(bom)

    armed = armed_parts_affected_by_bom(db, bom, company_id=company_id)
    warning = backflush_armed_edit_warning(armed, item_number=item.item_number)

    # log_delete serializes the model synchronously, so passing the live attached instance
    # captures its full old state. After db.delete() it is unusable.
    audit.log_delete(
        "bom_line",
        item.id,
        bom_line_identifier(
            _tenant_part_number(db, bom.part_id, company_id),
            item.item_number,
            _tenant_part_number(db, item.component_part_id, company_id),
        ),
        old_values=item,
        soft_delete=False,
        extra_data=_armed_extra_data(armed),
    )

    db.delete(item)
    db.commit()
    return {"message": "BOM item deleted", "backflush_armed_warning": warning}


# Multi-level BOM operations
def would_create_circular_reference(
    db: Session, parent_part_id: int, component_part_id: int, company_id: int, visited: Optional[Set[int]] = None
) -> bool:
    """Check if adding component would create a circular reference.

    Tenant-scoped (invariant #1): every BOM lookup in the walk carries ``company_id`` so the
    recursion can never traverse into — or leak the structure of — another company's BOMs.
    Soft-deleted BOMs are skipped (invariant 3): a structure the shop has deleted cannot
    close a loop, and refusing a legal line because of one would be a phantom refusal.
    """
    if visited is None:
        visited = set()

    if component_part_id in visited:
        return True

    if component_part_id == parent_part_id:
        return True

    visited.add(component_part_id)

    # Get the component's BOM (scoped to the active company)
    component_bom = (
        db.query(BOM)
        .filter(
            BOM.part_id == component_part_id,
            BOM.company_id == company_id,
            BOM.is_active == True,  # noqa: E712
            BOM.is_deleted == False,  # noqa: E712
        )
        .first()
    )

    if not component_bom:
        return False

    # Check each child
    for item in component_bom.items:
        if would_create_circular_reference(db, parent_part_id, item.component_part_id, company_id, visited.copy()):
            return True

    return False


def explode_bom_recursive(
    db: Session,
    bom_id: int,
    company_id: int,
    parent_qty: float = 1.0,
    level: int = 0,
    max_levels: int = 20,
    visited: Optional[Set[int]] = None,
) -> List[BOMItemWithChildren]:
    """Recursively explode a BOM to get all levels.

    Tenant-scoped (invariant #1): the top-level and every sub-BOM lookup filter on
    ``company_id`` so the recursive walk cannot cross a tenant boundary even through a
    corrupt/mis-parented row. Soft-deleted BOMs are not exploded (invariant 3), and every
    component is resolved through ``tenant_parts_by_id`` rather than the unscoped
    ``BOMItem.component_part`` relationship the explosion used to render.
    """
    if visited is None:
        visited = set()

    if level >= max_levels:
        return []

    bom = (
        db.query(BOM)
        .options(joinedload(BOM.items))
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )

    if not bom:
        return []

    # One scoped component read per BOM level, not one per line.
    components_by_id = tenant_parts_by_id(db, [i.component_part_id for i in bom.items], company_id)

    result = []
    for item in bom.items:
        if item.component_part_id in visited:
            continue  # Skip to prevent infinite loops

        # Handle NULL values defensively
        qty = item.quantity or 1.0
        scrap = item.scrap_factor if item.scrap_factor is not None else 0.0
        extended_qty = qty * parent_qty * (1 + scrap)

        # Check if component has its own BOM (scoped to the active company). This row is
        # what the recursion descends into AND, below, the component's ``has_bom`` flag --
        # the predicate is identical to ``parts_with_active_bom``'s, so the response builder
        # is handed the answer instead of re-running the same query per line.
        component_bom = (
            db.query(BOM)
            .filter(
                BOM.part_id == item.component_part_id,
                BOM.company_id == company_id,
                BOM.is_active == True,  # noqa: E712
                BOM.is_deleted == False,  # noqa: E712
            )
            .first()
        )

        children = []
        item_type = item.item_type or BOMItemType.MAKE
        if component_bom and item_type != BOMItemType.BUY:
            new_visited = visited.copy()
            new_visited.add(item.component_part_id)
            children = explode_bom_recursive(
                db, component_bom.id, company_id, extended_qty, level + 1, max_levels, new_visited
            )

        item_response = BOMItemWithChildren(
            id=item.id,
            bom_id=item.bom_id,
            component_part_id=item.component_part_id,
            item_number=item.item_number,
            quantity=qty,
            item_type=item_type,
            line_type=item.line_type if item.line_type else BOMLineType.COMPONENT,
            unit_of_measure=item.unit_of_measure or "each",
            reference_designator=item.reference_designator,
            find_number=item.find_number,
            notes=item.notes,
            torque_spec=item.torque_spec,
            installation_notes=item.installation_notes,
            work_center_id=item.work_center_id,
            operation_sequence=item.operation_sequence if item.operation_sequence is not None else 10,
            scrap_factor=scrap,
            lead_time_offset=item.lead_time_offset if item.lead_time_offset is not None else 0,
            is_optional=item.is_optional or False,
            is_alternate=item.is_alternate or False,
            alternate_group=item.alternate_group,
            component_part=(
                component_part_info(components_by_id[item.component_part_id], has_bom=component_bom is not None)
                if item.component_part_id in components_by_id
                else None
            ),
            created_at=item.created_at,
            updated_at=item.updated_at,
            children=children,
            level=level,
            extended_quantity=extended_qty,
        )
        result.append(item_response)

    return result


def get_max_level(items: List[BOMItemWithChildren], current_max: int = 0) -> int:
    """Get the maximum nesting level in exploded BOM"""
    for item in items:
        current_max = max(current_max, item.level)
        if item.children:
            current_max = get_max_level(item.children, current_max)
    return current_max


@router.get("/{bom_id}/explode", response_model=BOMExploded)
def explode_bom(
    bom_id: int,
    max_levels: int = Query(default=10, ge=1, le=20, description="Maximum levels to explode"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """Explode a BOM to show all levels (multi-level BOM)"""
    # Tenant-scoped (invariant #1): a foreign id gets a flat 404 — never confirm existence.
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    # The parent part TENANT-SCOPED, never the unscoped ``BOM.part`` relationship this used
    # to ``joinedload`` — see ``tenant_parts_by_id``.
    parent_part = tenant_parts_by_id(db, [bom.part_id], company_id).get(bom.part_id)

    items = explode_bom_recursive(db, bom_id, company_id, 1.0, 0, max_levels)
    total_levels = get_max_level(items) + 1 if items else 0

    return BOMExploded(
        bom_id=bom.id,
        part_id=bom.part_id,
        part_number=parent_part.part_number if parent_part else "",
        part_name=parent_part.name if parent_part else "",
        revision=bom.revision,
        total_levels=total_levels,
        items=items,
    )


def flatten_bom_items(items: List[BOMItemWithChildren], flat_list: List[BOMFlatItem], parent_qty: float = 1.0):
    """Flatten nested BOM items into a single list"""
    for item in items:
        flat_item = BOMFlatItem(
            level=item.level,
            item_number=item.item_number,
            find_number=item.find_number,
            part_id=item.component_part_id,
            part_number=item.component_part.part_number if item.component_part else "",
            part_name=item.component_part.name if item.component_part else "",
            # component_part here is ComponentPartInfo (part_type: str) — enum-or-string safe
            part_type=(
                item.component_part.part_type.value
                if item.component_part and hasattr(item.component_part.part_type, "value")
                else (item.component_part.part_type if item.component_part else "")
            ),
            item_type=item.item_type,
            line_type=item.line_type if item.line_type else BOMLineType.COMPONENT,
            quantity_per=item.quantity,
            extended_quantity=item.extended_quantity,
            unit_of_measure=item.unit_of_measure,
            scrap_factor=item.scrap_factor,
            lead_time_offset=item.lead_time_offset,
            is_optional=item.is_optional,
            is_alternate=item.is_alternate,
            has_children=len(item.children) > 0,
            torque_spec=item.torque_spec,
            installation_notes=item.installation_notes,
        )
        flat_list.append(flat_item)

        if item.children:
            flatten_bom_items(item.children, flat_list, item.extended_quantity)


@router.get("/{bom_id}/flatten", response_model=BOMFlattened)
def flatten_bom(
    bom_id: int,
    max_levels: int = Query(default=10, ge=1, le=20),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """Get a flattened view of a multi-level BOM (for reports/MRP)"""
    # Tenant-scoped (invariant #1): a foreign id gets a flat 404 — never confirm existence.
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    # The parent part TENANT-SCOPED — see ``tenant_parts_by_id``.
    parent_part = tenant_parts_by_id(db, [bom.part_id], company_id).get(bom.part_id)

    exploded = explode_bom_recursive(db, bom_id, company_id, 1.0, 0, max_levels)

    flat_items: List[BOMFlatItem] = []
    flatten_bom_items(exploded, flat_items)

    unique_parts = set(item.part_id for item in flat_items)

    return BOMFlattened(
        bom_id=bom.id,
        part_number=parent_part.part_number if parent_part else "",
        part_name=parent_part.name if parent_part else "",
        revision=bom.revision,
        total_items=len(flat_items),
        total_unique_parts=len(unique_parts),
        items=flat_items,
    )


@router.get("/{bom_id}/where-used")
def where_used(
    bom_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    company_id: int = Depends(get_current_company_id),
):
    """Find all parent assemblies that use this BOM's part"""
    # Tenant-scoped (invariant #1): a foreign id gets a flat 404 — never confirm existence.
    bom = (
        db.query(BOM)
        .filter(BOM.id == bom_id, BOM.company_id == company_id, BOM.is_deleted == False)  # noqa: E712
        .first()
    )
    if not bom:
        raise HTTPException(status_code=404, detail="BOM not found")

    # Find all BOM items that reference this part — scoped to the active company's BOMs
    usages = (
        db.query(BOMItem)
        .join(BOM, BOMItem.bom_id == BOM.id)
        .options(joinedload(BOMItem.bom))
        .filter(
            BOMItem.component_part_id == bom.part_id,
            BOM.company_id == company_id,
            # A deleted parent assembly is not a "where used" (invariant 3) -- its lines are
            # retained by ``delete_bom`` on purpose, so without this they keep reporting a
            # usage the shop has deleted.
            BOM.is_deleted == False,  # noqa: E712
        )
        .all()
    )

    # Every part this response names — this BOM's own part and each parent assembly's —
    # resolved TENANT-SCOPED in one read, never off ``BOM.part``. That relationship carries
    # no ``company_id`` predicate, and this endpoint prints the part number and name of
    # every assembly it finds; on a mis-parented row it printed a foreign one.
    parent_part_ids = {usage.bom.part_id for usage in usages if usage.bom}
    parent_part_ids.add(bom.part_id)
    parts_by_id = tenant_parts_by_id(db, parent_part_ids, company_id)

    result = []
    for usage in usages:
        parent_part = parts_by_id.get(usage.bom.part_id) if usage.bom else None
        if parent_part:
            result.append(
                {
                    "parent_part_id": usage.bom.part_id,
                    "parent_part_number": parent_part.part_number,
                    "parent_part_name": parent_part.name,
                    "bom_id": usage.bom_id,
                    "quantity_used": usage.quantity,
                    # item_type is a String(20) column — handle enum-or-string like get_bom does
                    "item_type": usage.item_type.value if hasattr(usage.item_type, "value") else usage.item_type,
                }
            )

    own_part = parts_by_id.get(bom.part_id)
    return {
        "part_id": bom.part_id,
        "part_number": own_part.part_number if own_part else None,
        "used_in": result,
    }
