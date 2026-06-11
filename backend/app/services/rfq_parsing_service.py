"""RFQ package ingestion and parsing for sheet-metal AI estimating."""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from app.models.rfq_quote import RfqPackageFile
from app.services.pdf_service import extract_text_from_pdf
from app.services.storage_service import ref_as_local_path

GAUGE_TO_INCHES: Dict[str, float] = {
    "24ga": 0.0239,
    "22ga": 0.0299,
    "20ga": 0.0359,
    "18ga": 0.0478,
    "16ga": 0.0598,
    "14ga": 0.0747,
    "12ga": 0.1046,
    "11ga": 0.1196,
    "10ga": 0.1345,
    "7ga": 0.1793,
}


MATERIAL_PATTERNS = [
    (re.compile(r"\b(304|316)\s*stainless|\bstainless\b", re.IGNORECASE), "Stainless Steel"),
    (re.compile(r"\b(5052|6061|aluminum|aluminium)\b", re.IGNORECASE), "Aluminum"),
    (re.compile(r"\b(copper|c110|electrolytic tough pitch)\b", re.IGNORECASE), "Copper"),
    (re.compile(r"\b(a36|1018|mild steel|carbon steel|crs|hrs|steel)\b", re.IGNORECASE), "Carbon Steel"),
]


FINISH_PATTERNS = [
    re.compile(r"powder\s*coat(?:ing)?", re.IGNORECASE),
    re.compile(r"\bpaint(?:ed|ing)?\b", re.IGNORECASE),
    re.compile(r"\banodiz(?:e|ed|ing)\b", re.IGNORECASE),
    re.compile(r"\bpassivat(?:e|ed|ion)\b", re.IGNORECASE),
    re.compile(r"\bzinc\s*plate|plating\b", re.IGNORECASE),
]


HARDWARE_HINTS = ("pem", "nut", "bolt", "screw", "standoff", "washer", "rivet", "insert", "stud")
PROCESS_HINTS = ("chemical film", "chem film", "tin plate", "plate", "plating", "anodize", "passivate", "process")
REFERENCE_HINTS = ("practice", "specification", "standard", "reference", "document")
MATERIAL_HINTS = ("aluminum", "aluminium", "steel", "stainless", "copper", "brass", "bronze")
MATCH_STOPWORDS = (
    "flat",
    "pattern",
    "rev",
    "revision",
    "sheet",
    "part",
    "dxf",
    "pdf",
    "step",
    "stp",
    "drawing",
    "dwg",
    "laser",
    "cut",
)


def _clean_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def _safe_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        return float(text)
    except Exception:
        return None


def _detect_material(text: str) -> Optional[str]:
    for pattern, material in MATERIAL_PATTERNS:
        if pattern.search(text):
            return material
    return None


def _detect_finish(text: str) -> Optional[str]:
    for pattern in FINISH_PATTERNS:
        match = pattern.search(text)
        if match:
            return match.group(0)
    return None


def _parse_thickness(text: str) -> Tuple[Optional[str], Optional[float], float]:
    if not text:
        return None, None, 0.0

    gauge_match = re.search(r"\b(\d{1,2})\s*(ga|gauge)\b", text, re.IGNORECASE)
    if gauge_match:
        gauge_key = f"{gauge_match.group(1)}ga"
        thickness = GAUGE_TO_INCHES.get(gauge_key)
        return gauge_key, thickness, 0.85 if thickness else 0.55

    inch_match = re.search(r"(\d*\.?\d+)\s*(?:in|\"|\binch(?:es)?)\b", text, re.IGNORECASE)
    if inch_match:
        raw = inch_match.group(1)
        value = _safe_float(raw)
        return raw, value, 0.80 if value else 0.4

    mm_match = re.search(r"(\d*\.?\d+)\s*mm\b", text, re.IGNORECASE)
    if mm_match:
        raw_mm = mm_match.group(1)
        mm_value = _safe_float(raw_mm)
        if mm_value is None:
            return None, None, 0.0
        return f"{raw_mm}mm", mm_value / 25.4, 0.70

    return None, None, 0.0


def _extract_pdf_part_hint(text: str) -> Optional[str]:
    patterns = [
        re.compile(r"(?:part\s*(?:no|number)|p\/n|dwg(?:\s*no)?)[\s:#-]*([A-Z0-9._-]{3,})", re.IGNORECASE),
    ]
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return None


def _extract_drawing_number(text: str) -> Optional[str]:
    patterns = [
        re.compile(r"(?:drawing|dwg)\s*(?:no|number|#)?[\s:#-]*([A-Z0-9._-]{3,})", re.IGNORECASE),
        re.compile(r"(?:print|drawing)\s*[\s:#-]+([A-Z0-9._-]{3,})", re.IGNORECASE),
    ]
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return None


def _extract_revision(text: str) -> Optional[str]:
    patterns = [
        re.compile(r"\brev(?:ision)?\b[\s:#-]*([A-Z0-9]{1,4})\b", re.IGNORECASE),
        re.compile(r"\b([A-Z0-9]{1,4})\s+rev(?:ision)?\b", re.IGNORECASE),
    ]
    for pattern in patterns:
        match = pattern.search(text)
        if match:
            return match.group(1).strip()
    return None


def _parse_feature_count(text: str, feature_terms: Tuple[str, ...]) -> Optional[int]:
    if not text:
        return None
    feature_pattern = "|".join(feature_terms)
    patterns = [
        re.compile(rf"\b(?:{feature_pattern})\s*(?:count|qty|quantity)?\s*[:#-]\s*(\d{{1,4}})\b", re.IGNORECASE),
        re.compile(
            rf"\b(\d{{1,4}})\s*x?\s*(?:[oØø⌀]?\s*\d*\.?\d+\s*(?:in|mm|\")?\s*)?(?:{feature_pattern})\b", re.IGNORECASE
        ),
    ]
    candidates: List[int] = []
    for pattern in patterns:
        for match in pattern.finditer(text):
            count = _parse_int(match.group(1))
            if count and 0 < count < 1000:
                candidates.append(count)
    return max(candidates) if candidates else None


def _hint_variants(*values: Any) -> List[str]:
    variants: List[str] = []

    def add_variant(value: str) -> None:
        clean = _clean_key(value)
        if clean and clean not in variants:
            variants.append(clean)

    for value in values:
        if value is None:
            continue
        raw = str(value).strip()
        if not raw:
            continue
        stem = Path(raw).stem
        add_variant(stem)
        without_rev = re.sub(r"(?:[\s._-]*rev(?:ision)?[\s._-]*[A-Z0-9]{1,4})$", "", stem, flags=re.IGNORECASE)
        add_variant(without_rev)
        tokens = re.findall(r"[a-z0-9]+", stem.lower())
        filtered = [token for token in tokens if token not in MATCH_STOPWORDS]
        add_variant("".join(filtered))
    return variants


def _append_unique(target: Dict[str, List[str]], key: str, values: List[str]) -> None:
    bucket = target.setdefault(key, [])
    for value in values:
        if value and value not in bucket:
            bucket.append(value)


def _parse_int(value: Any) -> Optional[int]:
    float_value = _safe_float(value)
    if float_value is None:
        return None
    try:
        return int(round(float_value))
    except Exception:
        return None


def _file_stem_part_number(file_name: str) -> Optional[str]:
    stem = Path(file_name).stem
    stem = re.sub(r"(?:[\s._-]*rev(?:ision)?[\s._-]*[A-Z0-9]+)$", "", stem, flags=re.IGNORECASE)
    match = re.search(r"\b([0-9]{3,}-[0-9A-Z]{3,}-[0-9A-Z]{3,})\b", stem, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    clean = stem.strip()
    return clean or None


def _revision_from_filename(file_name: str) -> Optional[str]:
    match = re.search(r"(?:^|[\s._-])rev(?:ision)?[\s._-]*([A-Z0-9]{1,4})(?:$|[\s._-])", file_name, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None


def _extract_title_from_text(text: str) -> Optional[str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    prep_match = re.search(r"\bPREP\s*\n\s*([A-Z][A-Z0-9 ,/&()'.-]{3,})\s*\n", text, re.IGNORECASE)
    if prep_match:
        candidate = prep_match.group(1).strip()
        if candidate and not re.search(r"\b(CHK|APVD|DWG|CAGEC|SCALE|AXIS|TOLERANCE)\b", candidate, re.IGNORECASE):
            return candidate
    for idx, line in enumerate(lines):
        if line.upper() == "TITLE" and idx > 0:
            candidate = lines[idx - 1].strip()
            if candidate and not re.search(r"\b(PREP|APVD|CHK|DWG|CAGEC|SCALE)\b", candidate, re.IGNORECASE):
                return candidate
    return None


def _extract_calculated_weight_lbs(text: str) -> Optional[float]:
    match = re.search(r"CALCULATED\s+WEIGHT\s*:\s*\(?\s*(\d*\.?\d+)\s*LBS?\s*\)?", text, re.IGNORECASE)
    if not match:
        return None
    return _safe_float(match.group(1))


def _find_dimension_pair(text: str) -> Optional[Tuple[float, float]]:
    if not text:
        return None
    # Common forms: 12 x 8, 12.5 X 8.2 in, 320 x 220 mm.
    match = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:in|mm|\"|inches)?\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:in|mm|\"|inches)?",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None
    first = _safe_float(match.group(1))
    second = _safe_float(match.group(2))
    if first is None or second is None:
        return None
    # If millimeters are explicit, convert to inches.
    if "mm" in text.lower():
        return first / 25.4, second / 25.4
    return first, second


def _line_type_for_assembly_item(qty_token: str, description: str, part_number: Optional[str]) -> Tuple[str, str]:
    text = f"{part_number or ''} {description or ''}".lower()
    if any(keyword in text for keyword in HARDWARE_HINTS):
        return "hardware", "buy"
    if any(keyword in text for keyword in ("ink", "epoxy", "adhesive", "paint", "sealant")):
        return "consumable", "buy"
    if any(keyword in text for keyword in PROCESS_HINTS):
        if any(keyword in text for keyword in ("fabrication practice", "assembly practice", "practices")):
            return "reference", "reference"
        return "process", "buy"
    if str(qty_token or "").strip().upper() == "REF":
        return "reference", "reference"
    if any(keyword in text for keyword in MATERIAL_HINTS):
        return "reference", "buy"
    if any(keyword in text for keyword in REFERENCE_HINTS):
        return "reference", "reference"
    return "purchased", "buy"


def _parse_qty_token(value: str) -> Tuple[float, bool]:
    token = str(value or "").strip().upper()
    if token in {"AR", "A/R", "ASREQD", "ASREQUIRED"}:
        return 1.0, True
    if token == "REF":
        return 0.0, False
    parsed = _safe_float(token)
    return (parsed if parsed and parsed > 0 else 1.0), False


def _extract_parts_list_rows(text: str) -> List[Dict[str, Any]]:
    if "PARTS LIST" not in text.upper():
        return []
    after = re.split(r"\bPARTS\s+LIST\b", text, flags=re.IGNORECASE, maxsplit=1)[-1]
    before = re.split(
        r"\n\s*(?:ASSEMBLY|REFERENCE VIEW|STATE\s+\d+|PART\s+1\b)", after, flags=re.IGNORECASE, maxsplit=1
    )[0]
    rows: List[Dict[str, Any]] = []
    for raw_line in before.splitlines():
        line = re.sub(r"\s+", " ", raw_line.strip())
        if not line:
            continue
        match = re.match(
            r"^(?P<qty>REF|AR|A/R|\d+(?:\.\d+)?)\s+"
            r"(?P<item>\d{1,4})\s+"
            r"(?P<part>[A-Z0-9][A-Z0-9._/-]{2,})\s+"
            r"(?P<desc>.+)$",
            line,
            re.IGNORECASE,
        )
        if not match:
            continue
        qty_token = match.group("qty").upper()
        qty, qty_as_required = _parse_qty_token(qty_token)
        item_number = match.group("item")
        part_number = match.group("part").strip()
        description = match.group("desc").strip()
        note_ref = None
        note_match = re.search(r"\s+(\d{1,3})$", description)
        if note_match:
            note_ref = note_match.group(1)
            description = description[: note_match.start()].strip()
        line_type, item_type = _line_type_for_assembly_item(qty_token, description, part_number)
        rows.append(
            {
                "line_number": int(item_number) * 10,
                "item_number": item_number,
                "find_number": item_number,
                "part_number": part_number,
                "part_name": description,
                "description": description,
                "qty": qty,
                "quantity_per_assembly": qty,
                "qty_as_required": qty_as_required,
                "unit_of_measure": "each",
                "line_type": line_type,
                "item_type": item_type,
                "note_ref": note_ref,
                "source": "parts-list",
            }
        )
    return rows


def _find_parts_list_item(rows: List[Dict[str, Any]], item_number: str) -> Optional[Dict[str, Any]]:
    for row in rows:
        if str(row.get("item_number")) == str(item_number):
            return row
    return None


def _part_numbers_from_phrase(phrase: str) -> List[int]:
    return [int(value) for value in re.findall(r"\b\d{1,3}\b", phrase or "")]


def _clean_finish_name(value: str) -> str:
    text = re.sub(r"\bPROCESS,\s*", "", value or "", flags=re.IGNORECASE).strip(" .")
    return text.title() if text.isupper() else text


def _extract_manufactured_detail_specs(
    text: str, assembly_number: str, rows: List[Dict[str, Any]], file_name: str
) -> List[Dict[str, Any]]:
    discovered = {int(value) for value in re.findall(r"\bPART\s+(\d{1,3})\b", text, re.IGNORECASE)}
    if not discovered:
        return []

    qty_by_part = {part_no: 1.0 for part_no in discovered}
    for match in re.finditer(r"((?:PART\s+\d+\s*)+)\s+QUANTITY\s*:\s*(\d+(?:\.\d+)?)", text, re.IGNORECASE):
        qty = _safe_float(match.group(2)) or 1.0
        for part_no in _part_numbers_from_phrase(match.group(1)):
            qty_by_part[part_no] = qty

    material_by_part: Dict[int, str] = {}
    thickness_by_part: Dict[int, Tuple[str, Optional[float], float]] = {}
    notes_by_part: Dict[int, List[str]] = {part_no: [] for part_no in discovered}

    for match in re.finditer(
        r"\bPARTS?\s+([0-9,\sANDand]+):\s*ITEM\s+(\d+),\s*([.\d]+)\s*STOCK",
        text,
        re.IGNORECASE,
    ):
        item = _find_parts_list_item(rows, match.group(2))
        material = _detect_material(item.get("description", "") if item else "") or (item or {}).get("description")
        raw_thickness = match.group(3)
        thickness = _parse_thickness(f"{raw_thickness} in")
        for part_no in _part_numbers_from_phrase(match.group(1)):
            material_by_part[part_no] = material or "Aluminum"
            thickness_by_part[part_no] = thickness
            notes_by_part.setdefault(part_no, []).append(
                f"Material from note: item {match.group(2)}, {raw_thickness} stock"
            )

    for match in re.finditer(r"\bPART\s+(\d+):\s*(?!ITEM\b)(.+?)\s+([.\d]+)\s*STOCK", text, re.IGNORECASE | re.DOTALL):
        part_no = int(match.group(1))
        if part_no in material_by_part and part_no in thickness_by_part:
            continue
        material_text = re.sub(r"\s+", " ", match.group(2)).strip(" ,")
        material = _detect_material(material_text) or material_text
        raw_thickness = match.group(3)
        material_by_part[part_no] = material
        thickness_by_part[part_no] = _parse_thickness(f"{raw_thickness} in")
        notes_by_part.setdefault(part_no, []).append(f"Material from note: {material}, {raw_thickness} stock")

    finish_by_part: Dict[int, str] = {}
    finish_section_match = re.search(
        r"\bFINISH\s*:\s*(.+?)(?:\n\s*\d+\s+IDENTIFICATION|\n\s*\d+\s+RIVET|\n\s*\d+\s+UNLESS|\Z)",
        text,
        re.IGNORECASE | re.DOTALL,
    )
    finish_section = re.sub(r"\s+", " ", finish_section_match.group(1).strip()) if finish_section_match else ""
    for match in re.finditer(r"\bPARTS?\s+([0-9,\sANDand]+):\s*ITEM\s+(\d+)", finish_section, re.IGNORECASE):
        item = _find_parts_list_item(rows, match.group(2))
        finish = _clean_finish_name(item.get("description", "") if item else f"Item {match.group(2)}")
        for part_no in _part_numbers_from_phrase(match.group(1)):
            finish_by_part[part_no] = finish
            notes_by_part.setdefault(part_no, []).append(f"Finish from note: item {match.group(2)}")
    for match in re.finditer(
        r"\bPART\s+(\d+):\s*([^.;]+(?:PLATE|PLATING|ANODIZE|FILM)[^.;]*)", finish_section, re.IGNORECASE
    ):
        part_no = int(match.group(1))
        finish = _clean_finish_name(match.group(2))
        finish_by_part[part_no] = finish
        notes_by_part.setdefault(part_no, []).append(f"Finish from note: {finish}")

    details: List[Dict[str, Any]] = []
    for part_no in sorted(discovered):
        thickness_raw, thickness_in, thickness_conf = thickness_by_part.get(part_no, (None, None, 0.0))
        material = material_by_part.get(part_no)
        finish = finish_by_part.get(part_no)
        detail_id = f"{assembly_number}-PART-{part_no}"
        sources: Dict[str, List[str]] = {
            "drawing_pdf": [f"{file_name}:part-{part_no}"],
            "cross_reference": [f"PDF {file_name} (assembly-detail)"],
        }
        if material:
            sources["material"] = [f"{file_name}:note-material-part-{part_no}"]
        if thickness_raw:
            sources["thickness"] = [f"{file_name}:note-thickness-part-{part_no}"]
        if finish:
            sources["finish"] = [f"{file_name}:note-finish-part-{part_no}"]
        details.append(
            {
                "part_number": detail_id,
                "part_name": f"PART {part_no}",
                "description": f"{assembly_number} detail PART {part_no}",
                "find_number": f"PART {part_no}",
                "item_number": f"P{part_no}",
                "qty": qty_by_part.get(part_no, 1.0),
                "quantity_per_assembly": qty_by_part.get(part_no, 1.0),
                "material": material,
                "thickness": thickness_raw,
                "thickness_in": thickness_in,
                "finish": finish,
                "line_type": "manufactured",
                "item_type": "make",
                "unit_of_measure": "each",
                "notes": "; ".join(notes_by_part.get(part_no, [])),
                "confidence": {
                    "material": 0.78 if material else 0.0,
                    "thickness": thickness_conf,
                    "finish": 0.72 if finish else 0.0,
                    "geometry": 0.0,
                },
                "sources": sources,
            }
        )
    return details


def _extract_assembly_payload(text: str, file_name: str) -> Optional[Dict[str, Any]]:
    rows = _extract_parts_list_rows(text)
    is_assembly = bool(rows) or bool(re.search(r"\bPARTS\s+LIST\b|\bNEXT\s+ASSY\b|\bASSEMBLY\b", text, re.IGNORECASE))
    if not is_assembly:
        return None

    file_part_number = _file_stem_part_number(file_name)
    drawing_number = file_part_number or _extract_drawing_number(text)
    assembly_number = drawing_number or file_part_number or Path(file_name).stem
    revision = _revision_from_filename(file_name)
    if not revision:
        rev_history = re.search(r"REVISION HISTORY.*?\n\s*([A-Z0-9]{1,4})\s+", text, re.IGNORECASE | re.DOTALL)
        revision = rev_history.group(1).upper() if rev_history else _extract_revision(text)
    title = _extract_title_from_text(text) or assembly_number
    manufactured_parts = _extract_manufactured_detail_specs(text, assembly_number, rows, file_name)
    return {
        "part_number": assembly_number,
        "part_name": title,
        "drawing_number": drawing_number,
        "revision": revision,
        "calculated_weight_lbs": _extract_calculated_weight_lbs(text),
        "bom_items": rows,
        "manufactured_parts": manufactured_parts,
    }


def parse_bom_xlsx(file_path: str, file_name: str) -> Dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True)
    part_rows: List[Dict[str, Any]] = []
    hardware_rows: List[Dict[str, Any]] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        header_row_index = None
        header_map: Dict[str, int] = {}
        for idx, row in enumerate(rows[:30]):
            normalized = [str(cell).strip().lower() if cell is not None else "" for cell in row]
            clean_tokens = [_clean_key(token) for token in normalized if token]
            has_qty = any("qty" in token or "quantity" in token for token in clean_tokens)
            has_part = any("part" in token or token in ("pn", "pn", "item", "itemno") for token in clean_tokens)
            if has_qty and has_part:
                header_row_index = idx
                for col_idx, token in enumerate(normalized):
                    if token:
                        header_map[_clean_key(token)] = col_idx
                break

        if header_row_index is None:
            continue

        def _get_value(data_row: Tuple[Any, ...], *candidates: str) -> Any:
            candidate_keys = [_clean_key(candidate) for candidate in candidates]
            for key, col_idx in header_map.items():
                if col_idx >= len(data_row):
                    continue
                if any(candidate_key in key for candidate_key in candidate_keys):
                    return data_row[col_idx]
            # Fallback: direct token startswith for noisy headers like "qty ea".
            for key, col_idx in header_map.items():
                if col_idx >= len(data_row):
                    continue
                if any(key.startswith(candidate_key) for candidate_key in candidate_keys):
                    return data_row[col_idx]
            return None

        def _get_notes_blob(data_row: Tuple[Any, ...]) -> str:
            notes_cells: List[str] = []
            for key, col_idx in header_map.items():
                if col_idx >= len(data_row):
                    continue
                if any(token in key for token in ("note", "remark", "comment", "description", "desc")):
                    value = data_row[col_idx]
                    if value is not None and str(value).strip():
                        notes_cells.append(str(value).strip())
            return " ".join(notes_cells)

        for row_idx, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
            part_number = _get_value(row, "part number", "part no", "part", "pn", "p/n", "item", "item no")
            description = _get_value(row, "description", "desc", "name", "part description")
            qty = _safe_float(_get_value(row, "qty", "quantity", "order qty", "required qty")) or 0
            if not part_number and not description:
                continue
            if qty <= 0:
                qty = 1

            material = _get_value(row, "material", "matl", "alloy")
            thickness = _get_value(row, "thickness", "gauge", "ga")
            finish = _get_value(row, "finish", "coating", "paint", "plating")
            item_type = str(_get_value(row, "type", "item type", "category") or "").lower()
            notes_blob = _get_notes_blob(row)

            flat_area = _safe_float(_get_value(row, "flat area", "area in2", "area"))
            cut_length = _safe_float(_get_value(row, "cut length", "perimeter", "total cut"))
            hole_count = _parse_int(_get_value(row, "hole count", "holes"))
            bend_count = _parse_int(_get_value(row, "bend count", "bends"))
            length = _safe_float(_get_value(row, "flat length", "length"))
            width = _safe_float(_get_value(row, "flat width", "width"))
            if flat_area is None and length and width:
                flat_area = length * width
            if cut_length is None and length and width:
                cut_length = 2.0 * (length + width)

            if (flat_area is None or cut_length is None) and notes_blob:
                dim_pair = _find_dimension_pair(notes_blob)
                if dim_pair:
                    dim_l, dim_w = dim_pair
                    if flat_area is None:
                        flat_area = dim_l * dim_w
                    if cut_length is None:
                        cut_length = 2.0 * (dim_l + dim_w)

            source_ref = f"{file_name}!{sheet.title}:row{row_idx}"
            row_data = {
                "part_number": str(part_number).strip() if part_number else None,
                "part_name": str(description).strip() if description else str(part_number).strip(),
                "qty": qty,
                "material": str(material).strip() if material else None,
                "thickness": str(thickness).strip() if thickness else None,
                "finish": str(finish).strip() if finish else None,
                "flat_area": flat_area,
                "cut_length": cut_length,
                "hole_count": hole_count,
                "bend_count": bend_count,
                "notes": (str(description).strip() if description else "") + (f" | {notes_blob}" if notes_blob else ""),
                "source": source_ref,
            }

            combined = " ".join([str(part_number or ""), str(description or ""), item_type, notes_blob]).lower()
            part_number_value = str(part_number or "").upper()
            is_hardware = (
                "hardware" in item_type
                or part_number_value.startswith(("HW", "BOLT", "NUT", "SCREW", "RVT", "PEM"))
                or any(keyword in combined for keyword in HARDWARE_HINTS)
            )
            if is_hardware:
                hardware_rows.append(row_data)
            else:
                part_rows.append(row_data)

    return {"parts": part_rows, "hardware": hardware_rows}


def parse_pdf_drawing(file_path: str, file_name: str) -> Dict[str, Any]:
    result = extract_text_from_pdf(file_path)
    text = (result.text or "").strip()
    assembly_payload = _extract_assembly_payload(text, file_name)
    material = _detect_material(text)
    thickness_raw, thickness_in, thickness_conf = _parse_thickness(text)
    finish = _detect_finish(text)
    drawing_number = _extract_drawing_number(text)
    if not drawing_number or drawing_number.upper() in {"NO", "NO.", "NUMBER"}:
        drawing_number = (assembly_payload or {}).get("drawing_number") or _file_stem_part_number(file_name)
    revision = _revision_from_filename(file_name) or _extract_revision(text)
    if assembly_payload and assembly_payload.get("revision"):
        revision = assembly_payload["revision"]
    hole_count = _parse_feature_count(text, (r"holes?", r"hole\s*qty", r"hole\s*quantity"))
    bend_count = _parse_feature_count(text, (r"bends?", r"bend\s*lines?"))
    tolerances_flag = bool(re.search(r"(?:\+\/-|±|tolerance)", text, re.IGNORECASE))
    weld_required = bool(re.search(r"(?:\bweld\b|fillet|gmaw|mig|tig)", text, re.IGNORECASE))
    assembly_required = bool(re.search(r"(?:assembly|assy)", text, re.IGNORECASE))
    part_hint = _extract_pdf_part_hint(text) or drawing_number or Path(file_name).stem
    dim_pair = _find_dimension_pair(text)
    inferred_flat_area = None
    inferred_cut_length = None
    geometry_confidence = 0.0
    if dim_pair:
        dim_l, dim_w = dim_pair
        inferred_flat_area = dim_l * dim_w
        inferred_cut_length = 2.0 * (dim_l + dim_w)
        geometry_confidence = 0.45

    sources: Dict[str, List[str]] = {}
    if material:
        sources["material"] = [f"{file_name}:text"]
    if thickness_raw:
        sources["thickness"] = [f"{file_name}:text"]
    if finish:
        sources["finish"] = [f"{file_name}:text"]
    if inferred_flat_area:
        sources["geometry"] = [f"{file_name}:text-dimensions"]
    drawing_details: List[str] = []
    if drawing_number:
        drawing_details.append(f"drawing_number={drawing_number}")
    if revision:
        drawing_details.append(f"revision={revision}")
    if hole_count is not None:
        drawing_details.append(f"holes={hole_count}")
    if bend_count is not None:
        drawing_details.append(f"bends={bend_count}")
    if drawing_details:
        sources["drawing_detail"] = [f"{file_name}:text:{';'.join(drawing_details)}"]

    return {
        "file_name": file_name,
        "source_type": "pdf",
        "document_kind": "assembly" if assembly_payload else "part",
        "part_hint": part_hint,
        "drawing_number": drawing_number,
        "revision": revision,
        "assembly": assembly_payload,
        "bom_items": (assembly_payload or {}).get("bom_items", []),
        "manufactured_parts": (assembly_payload or {}).get("manufactured_parts", []),
        "material": material,
        "thickness": thickness_raw,
        "thickness_in": thickness_in,
        "flat_area": inferred_flat_area,
        "cut_length": inferred_cut_length,
        "hole_count": hole_count,
        "bend_count": bend_count,
        "finish": finish,
        "weld_required": weld_required,
        "assembly_required": assembly_required,
        "tolerances_flag": tolerances_flag,
        "confidence": {
            "material": 0.80 if material else 0.0,
            "thickness": thickness_conf,
            "finish": 0.75 if finish else 0.0,
            "geometry": geometry_confidence,
        },
        "sources": sources,
        "text_length": len(text),
    }


def _poly_area(points: List[Tuple[float, float]]) -> float:
    if len(points) < 3:
        return 0.0
    area = 0.0
    for i in range(len(points)):
        x1, y1 = points[i]
        x2, y2 = points[(i + 1) % len(points)]
        area += x1 * y2 - x2 * y1
    return abs(area) / 2.0


def parse_dxf_geometry(file_path: str, file_name: str) -> Dict[str, Any]:
    import ezdxf

    doc = ezdxf.readfile(file_path)
    msp = doc.modelspace()
    entities = list(msp)

    min_x = float("inf")
    min_y = float("inf")
    max_x = float("-inf")
    max_y = float("-inf")
    cut_length = 0.0
    bend_count = 0
    hole_count = 0
    closed_areas: List[float] = []
    layers = sorted({str(entity.dxf.layer or "0") for entity in entities if hasattr(entity, "dxf")})
    entity_counts: Dict[str, int] = {}

    def _update_bounds(x: float, y: float) -> None:
        nonlocal min_x, min_y, max_x, max_y
        min_x = min(min_x, x)
        min_y = min(min_y, y)
        max_x = max(max_x, x)
        max_y = max(max_y, y)

    for entity in entities:
        entity_type = entity.dxftype()
        entity_counts[entity_type] = entity_counts.get(entity_type, 0) + 1
        layer = (entity.dxf.layer or "").lower()
        is_bend = any(token in layer for token in ("bend", "fold", "brake"))

        if entity_type == "LINE":
            start = entity.dxf.start
            end = entity.dxf.end
            _update_bounds(start.x, start.y)
            _update_bounds(end.x, end.y)
            length = math.sqrt((end.x - start.x) ** 2 + (end.y - start.y) ** 2)
            if is_bend:
                bend_count += 1
            else:
                cut_length += length
        elif entity_type == "CIRCLE":
            center = entity.dxf.center
            radius = entity.dxf.radius
            _update_bounds(center.x - radius, center.y - radius)
            _update_bounds(center.x + radius, center.y + radius)
            cut_length += 2 * math.pi * radius
            area = math.pi * radius * radius
            closed_areas.append(area)
            if radius * 2 <= 2.0:
                hole_count += 1
        elif entity_type == "ARC":
            center = entity.dxf.center
            radius = entity.dxf.radius
            start_angle = entity.dxf.start_angle
            end_angle = entity.dxf.end_angle
            _update_bounds(center.x - radius, center.y - radius)
            _update_bounds(center.x + radius, center.y + radius)
            span = end_angle - start_angle
            if span < 0:
                span += 360
            cut_length += 2 * math.pi * radius * (abs(span) / 360.0)
        elif entity_type in ("LWPOLYLINE", "POLYLINE"):
            points: List[Tuple[float, float]] = []
            if entity_type == "LWPOLYLINE":
                points = [(point[0], point[1]) for point in entity.get_points()]
                is_closed = bool(entity.closed)
            else:
                points = [(vertex.dxf.location.x, vertex.dxf.location.y) for vertex in entity.vertices]
                is_closed = bool(entity.is_closed)
            if not points:
                continue
            for x, y in points:
                _update_bounds(x, y)
            length = 0.0
            for idx in range(len(points) - 1):
                length += math.dist(points[idx], points[idx + 1])
            if is_closed and len(points) > 2:
                length += math.dist(points[-1], points[0])
                closed_areas.append(_poly_area(points))
            if is_bend:
                bend_count += 1
            else:
                cut_length += length

    flat_area = None
    if closed_areas:
        flat_area = max(closed_areas)
    elif min_x != float("inf"):
        flat_area = max(max_x - min_x, 0.0) * max(max_y - min_y, 0.0)

    confidence = 0.90 if closed_areas else 0.65 if flat_area else 0.0
    return {
        "file_name": file_name,
        "source_type": "dxf",
        "part_hint": Path(file_name).stem,
        "flat_area": flat_area,
        "cut_length": cut_length if cut_length > 0 else None,
        "hole_count": hole_count,
        "bend_count": bend_count,
        "bbox": {
            "min_x": None if min_x == float("inf") else min_x,
            "max_x": None if max_x == float("-inf") else max_x,
            "min_y": None if min_y == float("inf") else min_y,
            "max_y": None if max_y == float("-inf") else max_y,
        },
        "layers": layers,
        "entity_counts": entity_counts,
        "confidence": {"geometry": confidence},
        "sources": {"geometry": [f"{file_name}:modelspace"]},
    }


def parse_step_fallback(file_path: str, file_name: str) -> Dict[str, Any]:
    """
    STEP fallback when full geometry parser is unavailable.
    Attempts rough bounding box extraction from CARTESIAN_POINT coordinates.
    """
    content = Path(file_path).read_text(errors="ignore")
    product_names = []
    for match in re.finditer(r"PRODUCT\s*\(\s*'([^']+)'\s*,", content, re.IGNORECASE):
        name = match.group(1).strip()
        if name and name not in product_names:
            product_names.append(name)
    components = [
        {
            "component_name": name,
            "part_hint": name,
            "quantity": product_names.count(name) or 1,
            "source": f"{file_name}:PRODUCT",
        }
        for name in product_names[:100]
    ]
    points = re.findall(
        r"CARTESIAN_POINT\s*\(\s*'[^']*'\s*,\s*\(\s*([-+0-9.Ee]+)\s*,\s*([-+0-9.Ee]+)\s*,\s*([-+0-9.Ee]+)\s*\)\s*\)",
        content,
    )
    if not points:
        return {
            "file_name": file_name,
            "source_type": "step",
            "part_hint": Path(file_name).stem,
            "flat_area": None,
            "cut_length": None,
            "hole_count": None,
            "bend_count": None,
            "bbox": None,
            "components": components,
            "confidence": {"geometry": 0.0},
            "low_confidence": True,
            "sources": {"geometry": [f"{file_name}:fallback-none"]},
            "warning": "STEP parsing unavailable. Provide flat pattern DXF for high-confidence estimate.",
        }

    coords = [(float(x), float(y), float(z)) for x, y, z in points]
    xs = [p[0] for p in coords]
    ys = [p[1] for p in coords]
    zs = [p[2] for p in coords]
    length = max(xs) - min(xs)
    width = max(ys) - min(ys)
    thickness_guess = max(min(max(zs) - min(zs), max(length, width)), 0.0)
    flat_area = max(length * width, 0.0)

    return {
        "file_name": file_name,
        "source_type": "step",
        "part_hint": Path(file_name).stem,
        "flat_area": flat_area if flat_area > 0 else None,
        "cut_length": 2.0 * (length + width) if length > 0 and width > 0 else None,
        "hole_count": None,
        "bend_count": None,
        "bbox": {"length": length, "width": width, "thickness_guess": thickness_guess},
        "components": components,
        "confidence": {"geometry": 0.35},
        "low_confidence": True,
        "sources": {"geometry": [f"{file_name}:cartesian_points"]},
        "warning": "STEP parsed with bounding-box fallback. Confirm thickness and flat pattern DXF before release.",
    }


def build_normalized_part_specs(
    bom_parts: List[Dict[str, Any]],
    pdf_specs: List[Dict[str, Any]],
    dxf_specs: List[Dict[str, Any]],
    step_specs: List[Dict[str, Any]],
) -> Dict[str, Any]:
    parts: Dict[str, Dict[str, Any]] = {}
    assumptions: List[Dict[str, Any]] = []
    missing: List[Dict[str, Any]] = []
    global_sources: Dict[str, List[str]] = {}

    def ensure_part(key: str, default_name: str) -> Dict[str, Any]:
        if key not in parts:
            parts[key] = {
                "part_id": key,
                "part_name": default_name,
                "qty": 1,
                "material": None,
                "thickness": None,
                "thickness_in": None,
                "flat_area": None,
                "cut_length": None,
                "hole_count": None,
                "bend_count": None,
                "finish": None,
                "weld_required": False,
                "assembly_required": False,
                "tolerances_flag": False,
                "drawing_number": None,
                "revision": None,
                "line_type": "manufactured",
                "item_type": "make",
                "bom_level": 0,
                "parent_part_number": None,
                "item_number": None,
                "find_number": None,
                "quantity_per_assembly": 1,
                "unit_of_measure": "each",
                "calculated_weight_lbs": None,
                "notes": "",
                "confidence": {"material": 0.0, "thickness": 0.0, "geometry": 0.0, "finish": 0.0},
                "sources": {},
            }
        return parts[key]

    def score_part_match(payload: Dict[str, Any], part_key: str, part_value: Dict[str, Any]) -> Tuple[int, str]:
        target_variants = set(
            _hint_variants(
                part_key,
                part_value.get("part_id"),
                part_value.get("part_name"),
                part_value.get("drawing_number"),
            )
        )
        if not target_variants:
            return 0, ""

        hint_variants = set(_hint_variants(payload.get("part_hint"), payload.get("drawing_number")))
        file_variants = set(_hint_variants(payload.get("file_name")))
        if hint_variants & target_variants:
            return 100, "part-or-drawing-number"
        if file_variants & target_variants:
            return 90, "file-name"

        for payload_variant in hint_variants | file_variants:
            for target_variant in target_variants:
                if len(payload_variant) < 3 or len(target_variant) < 3:
                    continue
                if payload_variant in target_variant or target_variant in payload_variant:
                    return 75, "reduced-file-hint"
        return 0, ""

    def attach_by_hint(payload: Dict[str, Any], source_type: str) -> Tuple[Dict[str, Any], str]:
        best_part: Optional[Dict[str, Any]] = None
        best_score = 0
        best_reason = ""
        for part_key, part_value in parts.items():
            score, reason = score_part_match(payload, part_key, part_value)
            if score > best_score:
                best_part = part_value
                best_score = score
                best_reason = reason

        if best_part is not None and best_score >= 75:
            return best_part, best_reason

        if len(parts) == 1:
            only_part = next(iter(parts.values()))
            if only_part.get("sources", {}).get("bom"):
                return only_part, "single-bom-part"

        key_seed = (
            payload.get("part_hint")
            or payload.get("drawing_number")
            or payload.get("file_name")
            or f"part-{len(parts)+1}"
        )
        key_variants = _hint_variants(key_seed)
        key = key_variants[0] if key_variants else _clean_key(str(key_seed)) or f"part-{len(parts)+1}"
        part = ensure_part(key, str(payload.get("part_hint") or Path(str(payload.get("file_name") or key)).stem))
        if payload.get("part_hint"):
            part["part_id"] = str(payload["part_hint"])
        return part, f"created-from-{source_type}"

    def append_cross_reference(
        part: Dict[str, Any], source_type: str, payload: Dict[str, Any], matched_by: str
    ) -> None:
        file_name = str(payload.get("file_name") or payload.get("part_hint") or "unknown-file")
        source_key = {
            "bom": "bom",
            "pdf": "drawing_pdf",
            "dxf": "flat_pattern_dxf",
            "step": "step_model",
        }.get(source_type, source_type)
        ref = f"{file_name} ({matched_by})"
        _append_unique(part["sources"], source_key, [ref])
        if source_type in ("pdf", "dxf", "step"):
            _append_unique(part["sources"], "cross_reference", [f"{source_type.upper()} {ref}"])

    def append_once_assumption(item: Dict[str, Any]) -> None:
        if item not in assumptions:
            assumptions.append(item)

    def apply_common_line_fields(
        part: Dict[str, Any],
        *,
        line_type: Optional[str] = None,
        item_type: Optional[str] = None,
        bom_level: Optional[int] = None,
        parent_part_number: Optional[str] = None,
        item_number: Optional[str] = None,
        find_number: Optional[str] = None,
        quantity_per_assembly: Optional[float] = None,
        unit_of_measure: Optional[str] = None,
    ) -> None:
        if line_type:
            part["line_type"] = line_type
        if item_type:
            part["item_type"] = item_type
        if bom_level is not None:
            part["bom_level"] = bom_level
        if parent_part_number:
            part["parent_part_number"] = parent_part_number
        if item_number:
            part["item_number"] = str(item_number)
        if find_number:
            part["find_number"] = str(find_number)
        if quantity_per_assembly is not None:
            part["quantity_per_assembly"] = quantity_per_assembly
        if unit_of_measure:
            part["unit_of_measure"] = unit_of_measure

    def apply_assembly_child_payload(
        child: Dict[str, Any], parent_part_number: str, source_file: str
    ) -> Dict[str, Any]:
        key_seed = child.get("part_number") or child.get("part_name") or f"{parent_part_number}-item-{len(parts)+1}"
        key = _clean_key(str(key_seed)) or f"part-{len(parts)+1}"
        part = ensure_part(key, child.get("part_name") or str(key_seed))
        part["part_id"] = child.get("part_number") or part["part_id"]
        part["part_name"] = child.get("part_name") or part["part_name"]
        part["qty"] = child.get("qty") if child.get("qty") is not None else part["qty"]
        part["notes"] = child.get("notes") or part["notes"]
        apply_common_line_fields(
            part,
            line_type=child.get("line_type") or "manufactured",
            item_type=child.get("item_type") or "make",
            bom_level=1,
            parent_part_number=parent_part_number,
            item_number=child.get("item_number"),
            find_number=child.get("find_number"),
            quantity_per_assembly=child.get("quantity_per_assembly"),
            unit_of_measure=child.get("unit_of_measure"),
        )
        append_cross_reference(
            part, "pdf", {"file_name": source_file, "part_hint": child.get("part_number")}, "assembly-bom"
        )
        for source_key, refs in (child.get("sources") or {}).items():
            _append_unique(part["sources"], source_key, refs)
        if child.get("material"):
            part["material"] = child["material"]
            part["confidence"]["material"] = max(
                part["confidence"]["material"], child.get("confidence", {}).get("material", 0.72)
            )
            _append_unique(
                part["sources"], "material", child.get("sources", {}).get("material", [f"{source_file}:assembly-note"])
            )
        if child.get("thickness"):
            part["thickness"] = child["thickness"]
            part["thickness_in"] = child.get("thickness_in")
            part["confidence"]["thickness"] = max(
                part["confidence"]["thickness"], child.get("confidence", {}).get("thickness", 0.7)
            )
            _append_unique(
                part["sources"],
                "thickness",
                child.get("sources", {}).get("thickness", [f"{source_file}:assembly-note"]),
            )
        if child.get("finish"):
            part["finish"] = child["finish"]
            part["confidence"]["finish"] = max(
                part["confidence"]["finish"], child.get("confidence", {}).get("finish", 0.7)
            )
            _append_unique(
                part["sources"], "finish", child.get("sources", {}).get("finish", [f"{source_file}:assembly-note"])
            )
        return part

    for row in bom_parts:
        key_seed = row.get("part_number") or row.get("part_name") or f"row-{len(parts)+1}"
        key = _clean_key(key_seed) or f"part-{len(parts)+1}"
        part = ensure_part(key, row.get("part_name") or key_seed)
        part["part_id"] = row.get("part_number") or key
        part["part_name"] = row.get("part_name") or part["part_name"]
        part["qty"] = row.get("qty") or part["qty"]
        part["notes"] = row.get("notes") or part["notes"]
        append_cross_reference(part, "bom", {"file_name": row.get("source") or "bom"}, "bom-row")
        if row.get("material"):
            part["material"] = row["material"]
            part["confidence"]["material"] = max(part["confidence"]["material"], 0.9)
            _append_unique(part["sources"], "material", [row["source"]])
        if row.get("thickness"):
            raw, inches, conf = _parse_thickness(str(row["thickness"]))
            part["thickness"] = raw or str(row["thickness"])
            part["thickness_in"] = inches
            part["confidence"]["thickness"] = max(part["confidence"]["thickness"], conf)
            _append_unique(part["sources"], "thickness", [row["source"]])
        if row.get("finish"):
            part["finish"] = row["finish"]
            part["confidence"]["finish"] = max(part["confidence"]["finish"], 0.85)
            _append_unique(part["sources"], "finish", [row["source"]])
        if row.get("flat_area") is not None:
            part["flat_area"] = row["flat_area"]
            part["confidence"]["geometry"] = max(part["confidence"]["geometry"], 0.65)
            _append_unique(part["sources"], "geometry", [row["source"]])
            append_once_assumption(
                {
                    "part_id": part["part_id"],
                    "field": "geometry",
                    "assumption": "Geometry derived from BOM dimensions/columns.",
                    "confidence": 0.65,
                }
            )
        if row.get("cut_length") is not None:
            part["cut_length"] = row["cut_length"]
            part["confidence"]["geometry"] = max(part["confidence"]["geometry"], 0.65)
            _append_unique(part["sources"], "geometry", [row["source"]])
        if row.get("hole_count") is not None:
            part["hole_count"] = row["hole_count"]
            _append_unique(part["sources"], "geometry", [row["source"]])
        if row.get("bend_count") is not None:
            part["bend_count"] = row["bend_count"]
            _append_unique(part["sources"], "geometry", [row["source"]])

    for payload in pdf_specs:
        assembly_payload = payload.get("assembly") or {}
        if payload.get("document_kind") == "assembly" and assembly_payload:
            parent_number = (
                assembly_payload.get("part_number") or payload.get("drawing_number") or payload.get("part_hint")
            )
            parent_key = _clean_key(str(parent_number or payload.get("file_name") or f"assembly-{len(parts)+1}"))
            parent = ensure_part(
                parent_key or f"assembly-{len(parts)+1}", assembly_payload.get("part_name") or str(parent_number)
            )
            parent["part_id"] = parent_number or parent["part_id"]
            parent["part_name"] = assembly_payload.get("part_name") or parent["part_name"]
            parent["assembly_required"] = True
            parent["drawing_number"] = assembly_payload.get("drawing_number") or payload.get("drawing_number")
            parent["revision"] = assembly_payload.get("revision") or payload.get("revision")
            parent["calculated_weight_lbs"] = assembly_payload.get("calculated_weight_lbs")
            parent["notes"] = f"Assembly drawing {payload.get('file_name')}"
            parent["confidence"]["material"] = max(parent["confidence"]["material"], 0.0)
            apply_common_line_fields(
                parent,
                line_type="assembly",
                item_type="make",
                bom_level=0,
                quantity_per_assembly=1,
                unit_of_measure="each",
            )
            append_cross_reference(parent, "pdf", payload, "assembly-parent")
            _append_unique(parent["sources"], "drawing_detail", payload.get("sources", {}).get("drawing_detail", []))

            for child in payload.get("manufactured_parts") or []:
                apply_assembly_child_payload(child, str(parent["part_id"]), payload.get("file_name") or "assembly-pdf")

            for row in payload.get("bom_items") or []:
                line_type = row.get("line_type") or "purchased"
                child = {
                    "part_number": row.get("part_number") or f"{parent['part_id']}-ITEM-{row.get('item_number')}",
                    "part_name": row.get("part_name") or row.get("description") or f"Item {row.get('item_number')}",
                    "qty": row.get("qty"),
                    "quantity_per_assembly": row.get("quantity_per_assembly"),
                    "line_type": line_type,
                    "item_type": row.get("item_type") or ("buy" if line_type != "reference" else "reference"),
                    "unit_of_measure": row.get("unit_of_measure") or "each",
                    "item_number": row.get("item_number"),
                    "find_number": row.get("find_number"),
                    "notes": "As required quantity." if row.get("qty_as_required") else row.get("note_ref"),
                    "sources": {"bom": [f"{payload.get('file_name')}:parts-list-item-{row.get('item_number')}"]},
                    "confidence": {"material": 0.0, "thickness": 0.0, "geometry": 0.0, "finish": 0.0},
                }
                part = apply_assembly_child_payload(
                    child, str(parent["part_id"]), payload.get("file_name") or "assembly-pdf"
                )
                if line_type == "reference":
                    part["qty"] = 0
            continue

        part, matched_by = attach_by_hint(payload, "pdf")
        append_cross_reference(part, "pdf", payload, matched_by)
        if payload.get("drawing_number") and not part.get("drawing_number"):
            part["drawing_number"] = payload["drawing_number"]
        if payload.get("revision") and not part.get("revision"):
            part["revision"] = payload["revision"]
        _append_unique(part["sources"], "drawing_detail", payload.get("sources", {}).get("drawing_detail", []))
        if payload.get("material") and not part.get("material"):
            part["material"] = payload["material"]
            part["confidence"]["material"] = max(
                part["confidence"]["material"], payload.get("confidence", {}).get("material", 0.0)
            )
            _append_unique(part["sources"], "material", payload.get("sources", {}).get("material", []))
        if payload.get("thickness") and not part.get("thickness"):
            part["thickness"] = payload["thickness"]
            part["thickness_in"] = payload.get("thickness_in")
            part["confidence"]["thickness"] = max(
                part["confidence"]["thickness"], payload.get("confidence", {}).get("thickness", 0.0)
            )
            _append_unique(part["sources"], "thickness", payload.get("sources", {}).get("thickness", []))
        if payload.get("finish") and not part.get("finish"):
            part["finish"] = payload["finish"]
            part["confidence"]["finish"] = max(
                part["confidence"]["finish"], payload.get("confidence", {}).get("finish", 0.0)
            )
            _append_unique(part["sources"], "finish", payload.get("sources", {}).get("finish", []))
        if payload.get("flat_area") is not None and part.get("flat_area") is None:
            part["flat_area"] = payload["flat_area"]
            part["confidence"]["geometry"] = max(
                part["confidence"]["geometry"],
                payload.get("confidence", {}).get("geometry", 0.0),
            )
            _append_unique(part["sources"], "geometry", payload.get("sources", {}).get("geometry", []))
            append_once_assumption(
                {
                    "part_id": part["part_id"],
                    "field": "geometry",
                    "assumption": "Geometry inferred from PDF dimension text.",
                    "confidence": payload.get("confidence", {}).get("geometry", 0.0),
                }
            )
        if payload.get("cut_length") is not None and part.get("cut_length") is None:
            part["cut_length"] = payload["cut_length"]
            _append_unique(part["sources"], "geometry", payload.get("sources", {}).get("geometry", []))
        if payload.get("hole_count") is not None and part.get("hole_count") is None:
            part["hole_count"] = payload["hole_count"]
            _append_unique(part["sources"], "features", payload.get("sources", {}).get("drawing_detail", []))
        if payload.get("bend_count") is not None and part.get("bend_count") is None:
            part["bend_count"] = payload["bend_count"]
            _append_unique(part["sources"], "features", payload.get("sources", {}).get("drawing_detail", []))
        part["weld_required"] = part["weld_required"] or payload.get("weld_required", False)
        part["assembly_required"] = part["assembly_required"] or payload.get("assembly_required", False)
        part["tolerances_flag"] = part["tolerances_flag"] or payload.get("tolerances_flag", False)
        if payload.get("drawing_number") or payload.get("revision"):
            note_parts = []
            if payload.get("drawing_number"):
                note_parts.append(f"Drawing {payload['drawing_number']}")
            if payload.get("revision"):
                note_parts.append(f"Rev {payload['revision']}")
            note = " ".join(note_parts)
            if note and note not in (part.get("notes") or ""):
                part["notes"] = f"{part.get('notes') or ''} | {note}".strip(" |")

    for payload in dxf_specs + step_specs:
        source_type = str(payload.get("source_type") or "geometry")
        part, matched_by = attach_by_hint(payload, source_type)
        append_cross_reference(part, source_type, payload, matched_by)
        if source_type == "step" and payload.get("components"):
            for component in payload.get("components") or []:
                component_payload = {
                    "part_hint": component.get("part_hint") or component.get("component_name"),
                    "file_name": payload.get("file_name"),
                    "source_type": "step",
                }
                component_part, component_match = attach_by_hint(component_payload, "step")
                if component_part is part and component_part.get("line_type") == "assembly":
                    continue
                append_cross_reference(component_part, "step", component_payload, f"step-component-{component_match}")
                _append_unique(
                    component_part["sources"],
                    "step_component",
                    [f"{payload.get('file_name')}:PRODUCT:{component.get('component_name')}"],
                )
        if payload.get("flat_area") is not None:
            part["flat_area"] = payload["flat_area"]
        if payload.get("cut_length") is not None:
            part["cut_length"] = payload["cut_length"]
        if payload.get("hole_count") is not None:
            part["hole_count"] = payload["hole_count"]
        if payload.get("bend_count") is not None:
            part["bend_count"] = payload["bend_count"]
        part["confidence"]["geometry"] = max(
            part["confidence"]["geometry"], payload.get("confidence", {}).get("geometry", 0.0)
        )
        _append_unique(part["sources"], "geometry", payload.get("sources", {}).get("geometry", []))
        if source_type == "dxf" and part.get("sources", {}).get("drawing_pdf"):
            append_once_assumption(
                {
                    "part_id": part["part_id"],
                    "field": "cross_reference",
                    "assumption": "Cross-referenced PDF drawing details with DXF flat-pattern geometry.",
                    "confidence": min(0.95, max(0.75, payload.get("confidence", {}).get("geometry", 0.0))),
                }
            )
        warning = payload.get("warning")
        if warning:
            append_once_assumption(
                {
                    "part_id": part["part_id"],
                    "field": "geometry",
                    "assumption": warning,
                    "confidence": payload.get("confidence", {}).get("geometry", 0.0),
                }
            )

    # Controlled inference: if all resolved materials are identical, apply to unresolved parts.
    known_materials = {part["material"] for part in parts.values() if part.get("material")}
    if len(known_materials) == 1:
        inferred_material = next(iter(known_materials))
        for part in parts.values():
            if part.get("material"):
                continue
            part["material"] = inferred_material
            part["confidence"]["material"] = 0.6
            part["sources"].setdefault("material", []).append("inferred:single-material-across-package")
            assumptions.append(
                {
                    "part_id": part["part_id"],
                    "field": "material",
                    "assumption": f"Inferred {inferred_material} from package-wide consistency.",
                    "confidence": 0.6,
                }
            )

    for part in parts.values():
        line_type = str(part.get("line_type") or "manufactured")
        required = ("material", "thickness", "flat_area", "cut_length") if line_type == "manufactured" else ()
        for field in required:
            missing_condition = part.get(field) in (None, "", 0)
            if missing_condition:
                missing.append(
                    {
                        "part_id": part["part_id"],
                        "field": field,
                        "message": f"Missing required {field} for reliable sheet-metal quote.",
                    }
                )
        for field_key, refs in part["sources"].items():
            _append_unique(global_sources, field_key, refs)

    part_list = sorted(parts.values(), key=lambda item: str(item.get("part_id")))
    return {
        "parts": part_list,
        "assumptions": assumptions,
        "missing_specs": missing,
        "source_attribution": global_sources,
    }


def parse_rfq_package_files(files: List[RfqPackageFile]) -> Dict[str, Any]:
    bom_parts: List[Dict[str, Any]] = []
    hardware_items: List[Dict[str, Any]] = []
    pdf_specs: List[Dict[str, Any]] = []
    dxf_specs: List[Dict[str, Any]] = []
    step_specs: List[Dict[str, Any]] = []
    warnings: List[str] = []
    file_results: Dict[int, Dict[str, Any]] = {}

    for file_record in files:
        ext = (file_record.file_ext or "").lower()
        name = file_record.file_name
        try:
            if ext not in (".xlsx", ".xls", ".pdf", ".dxf", ".step", ".stp"):
                file_results[file_record.id] = {
                    "parse_status": "skipped",
                    "summary": {"reason": f"Unsupported extension {ext}"},
                }
                continue
            # Stored refs may live on object storage (s3://...); the parsing libs
            # (openpyxl, pdf2image, ezdxf) need a real local file, so materialize
            # the ref first. Local refs pass through unchanged.
            with ref_as_local_path(file_record.file_path) as local_file:
                local_path = str(local_file)
                if ext in (".xlsx", ".xls"):
                    bom = parse_bom_xlsx(local_path, name)
                    bom_parts.extend(bom["parts"])
                    hardware_items.extend(bom["hardware"])
                    file_results[file_record.id] = {
                        "parse_status": "parsed",
                        "summary": {
                            "parts_found": len(bom["parts"]),
                            "hardware_found": len(bom["hardware"]),
                        },
                    }
                elif ext == ".pdf":
                    pdf = parse_pdf_drawing(local_path, name)
                    pdf_specs.append(pdf)
                    if pdf.get("text_length", 0) < 120:
                        warnings.append(
                            f"{name}: low extracted text volume. If this is a scanned drawing, upload DXF or machine-readable PDF for better accuracy."
                        )
                    file_results[file_record.id] = {
                        "parse_status": "parsed",
                        "summary": {
                            "text_length": pdf["text_length"],
                            "document_kind": pdf.get("document_kind"),
                            "part_hint": pdf["part_hint"],
                            "drawing_number": pdf.get("drawing_number"),
                            "revision": pdf.get("revision"),
                            "assembly_items": len(pdf.get("bom_items") or []),
                            "manufactured_details": len(pdf.get("manufactured_parts") or []),
                            "material": pdf.get("material"),
                            "thickness": pdf.get("thickness"),
                            "finish": pdf.get("finish"),
                            "holes": pdf.get("hole_count"),
                            "bends": pdf.get("bend_count"),
                        },
                    }
                elif ext == ".dxf":
                    dxf = parse_dxf_geometry(local_path, name)
                    dxf_specs.append(dxf)
                    if not dxf.get("flat_area") or not dxf.get("cut_length"):
                        warnings.append(
                            f"{name}: geometry extracted with limited confidence. Verify units and closed outer profile."
                        )
                    file_results[file_record.id] = {
                        "parse_status": "parsed",
                        "summary": {
                            "part_hint": dxf.get("part_hint"),
                            "flat_area": dxf["flat_area"],
                            "cut_length": dxf["cut_length"],
                            "holes": dxf.get("hole_count"),
                            "bends": dxf.get("bend_count"),
                            "layers": dxf.get("layers", [])[:8],
                        },
                    }
                else:
                    step = parse_step_fallback(local_path, name)
                    step_specs.append(step)
                    file_results[file_record.id] = {
                        "parse_status": "parsed_with_fallback",
                        "summary": {
                            "flat_area": step.get("flat_area"),
                            "components": len(step.get("components") or []),
                            "warning": step.get("warning"),
                        },
                    }
        except Exception as exc:
            warnings.append(f"{name}: parse failed ({exc})")
            file_results[file_record.id] = {
                "parse_status": "error",
                "parse_error": str(exc),
            }

    normalized = build_normalized_part_specs(bom_parts, pdf_specs, dxf_specs, step_specs)
    for part in normalized["parts"]:
        part_sources = part.get("sources") or {}
        part_label = part.get("part_id") or part.get("part_name") or "part"
        has_pdf = bool(part_sources.get("drawing_pdf"))
        has_dxf = bool(part_sources.get("flat_pattern_dxf"))
        has_bom = bool(part_sources.get("bom"))
        if has_pdf and not has_dxf:
            warnings.append(
                f"{part_label}: drawing PDF parsed without a matching DXF flat pattern. Geometry may be inferred from drawing dimensions."
            )
        if has_dxf and not (has_pdf or has_bom):
            warnings.append(
                f"{part_label}: DXF geometry parsed without matching PDF/BOM specifications. Confirm material, thickness, finish, and revision."
            )
    return {
        "parts": normalized["parts"],
        "hardware_items": hardware_items,
        "assumptions": normalized["assumptions"],
        "missing_specs": normalized["missing_specs"],
        "source_attribution": normalized["source_attribution"],
        "warnings": warnings,
        "file_results": file_results,
    }
