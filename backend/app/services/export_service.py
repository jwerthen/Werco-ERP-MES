"""Data export service for CSV and Excel generation."""
import csv
import io
from datetime import datetime, date
from typing import Any, Dict, List, Optional
from decimal import Decimal


def format_value(value: Any) -> Any:
    """Format a value for export."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, 'value'):
        return value.value
    return value


def generate_csv(data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> str:
    """Generate CSV string from list of dictionaries."""
    if not data:
        return ""
    
    if columns:
        headers = columns
    else:
        headers = list(data[0].keys())
    
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    
    for row in data:
        writer.writerow([format_value(row.get(col)) for col in headers])
    
    return output.getvalue()


def generate_excel(
    data: List[Dict[str, Any]], 
    columns: Optional[List[str]] = None,
    sheet_name: str = "Export"
) -> bytes:
    """Generate Excel file from list of dictionaries."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    if not data:
        wb_output = io.BytesIO()
        wb.save(wb_output)
        return wb_output.getvalue()
    
    if columns:
        headers = columns
    else:
        headers = list(data[0].keys())
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col in enumerate(headers, 1):
            value = format_value(row_data.get(col))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal="right")
    
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        for row_idx in range(2, len(data) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    ws.freeze_panes = "A2"
    
    wb_output = io.BytesIO()
    wb.save(wb_output)
    return wb_output.getvalue()
