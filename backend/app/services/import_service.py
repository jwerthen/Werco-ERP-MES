"""Shared Excel/CSV import kit (A0.2 — Excel migration kit).

One parsing front door for every tabular import endpoint:

* Accepts **.xlsx** (openpyxl) and **.csv** (utf-8 / utf-8-sig) uploads and
  normalizes both into the same row shape the existing per-entity validators
  already consume: ``dict[normalized_header -> stripped string]``. The service
  deliberately does NOT re-implement any entity business rules — endpoints keep
  their own validators; this module only handles file mechanics.
* Header normalization: first non-empty row is the header; values are
  stripped/lowered/snake_cased (``"Part Number"`` -> ``part_number``).
* Defensive cell coercion: Excel dates -> ISO strings, integral floats ->
  ``"5"`` not ``"5.0"``, booleans -> ``"true"``/``"false"``, everything else ->
  stripped text, so numbers-typed-as-text and text-typed-as-numbers both land
  in the shape the row validators expect.
* Blank rows are dropped; rows whose first cell starts with ``#`` are treated
  as guidance/example rows (the server-generated templates mark their guidance
  row this way) and skipped.
* Sane caps on file size and row count so a stray 300 MB workbook can't take
  the API down.

It also builds the downloadable per-entity XLSX templates (styled header row,
a ``#``-marked plain-English guidance row, and example rows on a separate
``Examples`` sheet) that replace the CSV templates the frontend used to
hardcode client-side.
"""

from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime, time
from typing import Any, Dict, Iterator, List, Optional, Set, Tuple

# Caps chosen for go-live master-data loads: big enough for a full part master
# (thousands of rows), small enough that parsing stays interactive.
MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024  # 10 MB
MAX_IMPORT_ROWS = 10_000

# Rows whose FIRST cell starts with this marker are guidance/example rows
# (written by the server-generated templates) and are skipped on import.
SKIP_ROW_MARKER = "#"

CSV_EXTENSIONS = {".csv"}
XLSX_EXTENSIONS = {".xlsx"}
SUPPORTED_IMPORT_EXTENSIONS = CSV_EXTENSIONS | XLSX_EXTENSIONS

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_ACCEPTED_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%m/%d/%Y",
    "%m/%d/%y",
    "%m-%d-%Y",
    "%d-%b-%Y",
)


class ImportFileError(ValueError):
    """File-level import problem (bad type/encoding/headers/size). Maps to HTTP 400."""


def normalize_import_header(value: Any) -> str:
    """Normalize a header cell: strip, lower, snake_case, ascii-ish.

    Matches (and supersedes) the per-router ``_normalize_csv_header`` helpers so
    XLSX and CSV uploads resolve to the same column keys.
    """
    normalized = coerce_cell(value).strip().lower()
    normalized = normalized.replace("-", "_").replace(" ", "_").replace("/", "_")
    return re.sub(r"[^a-z0-9_]", "", normalized)


def coerce_cell(value: Any) -> str:
    """Coerce an arbitrary cell value (XLSX-typed or CSV text) to clean text.

    The per-entity validators all parse from strings (``_parse_float`` etc.), so
    the defensive coercions live here once:

    * ``None`` -> ``""``
    * bool -> ``"true"`` / ``"false"`` (matches the ``_parse_bool`` vocabularies)
    * integral floats -> ``"5"`` (Excel stores most numbers as floats; ``"5.0"``
      would break ``int()`` parses for quantities/lead times)
    * datetime/date/time -> ISO text (midnight datetimes collapse to the date,
      which is how Excel represents date-only cells)
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:.10f}".rstrip("0").rstrip(".")
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    return str(value).strip()


def parse_date_field(value: str, field_name: str) -> Optional[date]:
    """Parse a (already coerced) cell into a date; '' -> None; raises ValueError."""
    text = (value or "").strip()
    if not text:
        return None
    # Coerced Excel datetimes may carry a time component ("2026-06-01 07:30:00").
    text = text.split(" ")[0].split("T")[0]
    for fmt in _ACCEPTED_DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name} must be a date (YYYY-MM-DD)")


@dataclass
class ParsedTable:
    """Normalized tabular upload: headers + (file_row_number, row_dict) pairs."""

    headers: List[str]
    rows: List[Tuple[int, Dict[str, str]]] = field(default_factory=list)

    def iter_rows(self) -> Iterator[Tuple[int, Dict[str, str]]]:
        return iter(self.rows)

    def __len__(self) -> int:  # pragma: no cover - convenience
        return len(self.rows)


def _file_extension(filename: Optional[str]) -> str:
    if not filename or "." not in filename:
        return ""
    return "." + filename.rsplit(".", 1)[-1].lower()


def _rows_from_csv(content: bytes) -> Iterator[List[Any]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ImportFileError("CSV must be UTF-8 encoded") from exc
    yield from csv.reader(io.StringIO(decoded))


def _rows_from_xlsx(content: bytes) -> Iterator[Tuple[Any, ...]]:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except ImportFileError:  # pragma: no cover - defensive
        raise
    except Exception as exc:
        raise ImportFileError("Could not read the .xlsx file. Re-save it as a standard Excel workbook.") from exc
    try:
        # Import reads the FIRST sheet only; the template's "Examples" sheet is
        # therefore ignored by design.
        worksheet = workbook.worksheets[0]
        yield from worksheet.iter_rows(values_only=True)
    finally:
        workbook.close()


def parse_import_file(
    filename: Optional[str],
    content: bytes,
    *,
    required_columns: Optional[Set[str]] = None,
    max_rows: int = MAX_IMPORT_ROWS,
    max_bytes: int = MAX_IMPORT_FILE_BYTES,
) -> ParsedTable:
    """Parse an uploaded CSV/XLSX into normalized row dicts.

    Raises :class:`ImportFileError` for any file-level problem; row-level
    validation stays in the calling endpoint so its error contract is unchanged.
    """
    ext = _file_extension(filename)
    if ext not in SUPPORTED_IMPORT_EXTENSIONS:
        raise ImportFileError("Please upload a CSV or Excel (.xlsx) file")
    if not content:
        raise ImportFileError("Uploaded file is empty")
    if len(content) > max_bytes:
        raise ImportFileError(f"File is too large (max {max_bytes // (1024 * 1024)} MB)")

    raw_rows = _rows_from_csv(content) if ext in CSV_EXTENSIONS else _rows_from_xlsx(content)

    headers: List[str] = []
    table_rows: List[Tuple[int, Dict[str, str]]] = []

    for row_number, raw_row in enumerate(raw_rows, start=1):
        cells = [coerce_cell(cell) for cell in raw_row]
        if not any(cells):
            continue  # tolerate blank rows anywhere
        if not headers:
            headers = [normalize_import_header(cell) for cell in cells]
            continue
        if cells[0].startswith(SKIP_ROW_MARKER):
            continue  # template guidance/example row
        if len(table_rows) >= max_rows:
            raise ImportFileError(f"Too many rows (max {max_rows}). Split the file and import in batches.")
        row: Dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue  # trailing/unnamed columns are ignored
            if header in row and not (idx < len(cells) and cells[idx]):
                continue  # first occurrence of a duplicated header wins
            row[header] = cells[idx] if idx < len(cells) else ""
        table_rows.append((row_number, row))

    if not headers:
        raise ImportFileError("File must include a header row")

    if required_columns:
        missing = sorted(set(required_columns) - {h for h in headers if h})
        if missing:
            raise ImportFileError(f"Missing required columns: {', '.join(missing)}")

    return ParsedTable(headers=[h for h in headers if h], rows=table_rows)


# ---------------------------------------------------------------------------
# Server-generated XLSX templates
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class TemplateColumn:
    name: str
    note: str
    examples: Tuple[str, ...] = ()


@dataclass(frozen=True)
class ImportTemplateSpec:
    entity: str
    title: str
    description: str
    columns: Tuple[TemplateColumn, ...]


def _col(name: str, note: str, *examples: str) -> TemplateColumn:
    return TemplateColumn(name=name, note=note, examples=tuple(examples))


IMPORT_TEMPLATES: Dict[str, ImportTemplateSpec] = {
    spec.entity: spec
    for spec in (
        ImportTemplateSpec(
            entity="users",
            title="Users",
            description="Employee accounts. POST /api/v1/users/import-csv",
            columns=(
                _col("employee_id", "REQUIRED. Unique badge/employee number.", "EMP-1001", "EMP-1002", "EMP-1003"),
                _col("first_name", "REQUIRED.", "Maria", "Dale", "Quinn"),
                _col("last_name", "REQUIRED.", "Lopez", "Hardy", "Tran"),
                _col("email", "Optional. Generated from employee_id when blank.", "maria@werco.com", "", ""),
                _col("password", "Optional for operators (auto-generated). Required otherwise.", "", "Str0ng!Pass", ""),
                _col(
                    "role",
                    "Optional. operator (default), supervisor, manager, admin, viewer.",
                    "operator",
                    "supervisor",
                    "operator",
                ),
                _col("department", "Optional free text.", "Machining", "Welding", ""),
            ),
        ),
        ImportTemplateSpec(
            entity="parts",
            title="Engineering parts",
            description="Manufactured/assembly part master. POST /api/v1/parts/import-csv",
            columns=(
                _col("part_number", "REQUIRED. Unique; uppercased on import.", "1042-100", "1042-200", "ASSY-9001"),
                _col("name", "REQUIRED.", "Bracket, Left", "Bracket, Right", "Frame Assembly"),
                _col("part_type", "REQUIRED. manufactured or assembly.", "manufactured", "manufactured", "assembly"),
                _col("revision", "Optional. Defaults to A.", "B", "A", "C"),
                _col("description", "Optional.", "6061-T6 machined bracket", "", ""),
                _col("unit_of_measure", "Optional. Defaults to each.", "each", "each", "each"),
                _col("standard_cost", "Optional number.", "12.50", "14.25", "310"),
                _col("lead_time_days", "Optional whole number.", "10", "10", "30"),
                _col("is_critical", "Optional. true/false.", "false", "false", "true"),
                _col("requires_inspection", "Optional. true/false (default true).", "true", "true", "true"),
                _col("customer_name", "Optional.", "Acme Aero", "", ""),
                _col("customer_part_number", "Optional.", "ACME-771", "", ""),
                _col("drawing_number", "Optional.", "DWG-1042", "", ""),
            ),
        ),
        ImportTemplateSpec(
            entity="materials",
            title="Materials & supplies",
            description="Purchased/raw material/hardware/consumable master. POST /api/v1/materials/import-csv",
            columns=(
                _col("part_number", "REQUIRED. Unique; uppercased on import.", "RM-AL-0250", "HW-10-32", "CON-TAPE"),
                _col("name", "REQUIRED.", '0.250" 6061 plate', "10-32 locknut", "Masking tape"),
                _col(
                    "part_type",
                    "REQUIRED. purchased, raw_material, hardware, or consumable.",
                    "raw_material",
                    "hardware",
                    "consumable",
                ),
                _col("description", "Optional.", "4x8 sheet", "", ""),
                _col("unit_of_measure", "Optional. Defaults to each.", "pounds", "each", "each"),
                _col("standard_cost", "Optional number.", "3.10", "0.08", "2.50"),
                _col("lead_time_days", "Optional whole number.", "14", "5", "3"),
                _col("reorder_point", "Optional number.", "500", "1000", "12"),
                _col("reorder_quantity", "Optional number.", "1000", "5000", "24"),
            ),
        ),
        ImportTemplateSpec(
            entity="customers",
            title="Customers",
            description="Customer master. POST /api/v1/customers/import-csv",
            columns=(
                _col("name", "REQUIRED. Unique customer name.", "Acme Aero", "Borealis Defense", "Cardinal Pumps"),
                _col("code", "Optional. Generated when blank.", "ACM001", "", ""),
                _col("contact_name", "Optional.", "Pat Chu", "", ""),
                _col("email", "Optional.", "pat@acmeaero.com", "", ""),
                _col("phone", "Optional.", "555-201-7788", "", ""),
                _col("address_line1", "Optional.", "100 Flight Way", "", ""),
                _col("city", "Optional.", "Wichita", "", ""),
                _col("state", "Optional.", "KS", "", ""),
                _col("zip_code", "Optional.", "67202", "", ""),
                _col("payment_terms", "Optional. Defaults to Net 30.", "Net 30", "Net 45", ""),
                _col("requires_coc", "Optional. true/false (default true).", "true", "true", "false"),
                _col("requires_fai", "Optional. true/false (default false).", "false", "true", "false"),
            ),
        ),
        ImportTemplateSpec(
            entity="vendors",
            title="Vendors",
            description="Vendor/supplier master. POST /api/v1/purchasing/vendors/import-csv",
            columns=(
                _col("name", "REQUIRED.", "Apex Metals", "Bolt Bin Supply", "Crest Finishing"),
                _col("code", "Optional. Generated when blank.", "APX001", "", ""),
                _col("contact_name", "Optional.", "Sam Reyes", "", ""),
                _col("email", "Optional.", "sales@apexmetals.com", "", ""),
                _col("phone", "Optional.", "555-310-4400", "", ""),
                _col("payment_terms", "Optional.", "Net 30", "", ""),
                _col("lead_time_days", "Optional whole number (default 14).", "10", "5", "21"),
                _col("is_approved", "Optional. true/false.", "true", "true", "false"),
                _col("is_as9100_certified", "Optional. true/false.", "true", "false", "false"),
                _col("is_iso9001_certified", "Optional. true/false.", "true", "false", "true"),
            ),
        ),
        ImportTemplateSpec(
            entity="work-centers",
            title="Work centers",
            description="Work center master. POST /api/v1/work-centers/import-csv",
            columns=(
                _col("code", "REQUIRED. Unique; uppercased on import.", "MILL-01", "LATHE-01", "WELD-01"),
                _col("name", "REQUIRED.", "Haas VF-2", "Mazak QT-250", "Weld Bay 1"),
                _col(
                    "work_center_type",
                    "REQUIRED. Must match a configured type (e.g. machining, fabrication, welding, inspection).",
                    "machining",
                    "machining",
                    "welding",
                ),
                _col("description", "Optional.", "3-axis vertical mill", "", ""),
                _col("hourly_rate", "Optional number.", "95", "90", "80"),
                _col("capacity_hours_per_day", "Optional number (default 8).", "8", "8", "8"),
                _col("efficiency_factor", "Optional number (default 1.0).", "0.9", "0.85", "1.0"),
                _col("building", "Optional.", "Main", "Main", "North"),
                _col("area", "Optional.", "CNC", "CNC", "Fab"),
            ),
        ),
        ImportTemplateSpec(
            entity="work-orders",
            title="Open work orders",
            description=(
                "Open (in-flight) work orders for go-live. POST /api/v1/work-orders/import. "
                "Operations come from the part's released routing."
            ),
            columns=(
                _col("wo_number", "Optional. Generated when blank. Must be unique.", "WO-LEGACY-0188", "", ""),
                _col(
                    "part_number",
                    "REQUIRED. Part must already exist with a released routing.",
                    "1042-100",
                    "1042-200",
                    "ASSY-9001",
                ),
                _col("quantity", "REQUIRED. Quantity ordered (> 0).", "25", "100", "4"),
                _col("due_date", "Optional date (YYYY-MM-DD).", "2026-07-15", "2026-06-30", ""),
                _col("customer", "Optional. Existing customer code or name.", "ACM001", "Borealis Defense", ""),
                _col("customer_po", "Optional.", "PO-77812", "", ""),
                _col("priority", "Optional 1-10 (1 = highest, default 5).", "3", "5", ""),
                _col(
                    "completed_through_seq",
                    "Optional. Last routing operation sequence ALREADY finished on paper before migration; "
                    "those operations are marked complete and the next one becomes ready.",
                    "20",
                    "",
                    "",
                ),
            ),
        ),
        ImportTemplateSpec(
            entity="purchase-orders",
            title="Open purchase orders",
            description=(
                "Open (issued, not yet received) purchase orders for go-live. "
                "POST /api/v1/purchasing/purchase-orders/import. "
                "Rows sharing a po_number become lines of one PO."
            ),
            columns=(
                _col(
                    "po_number",
                    "Optional. Rows with the same po_number form one PO; generated when blank.",
                    "PO-LEGACY-2201",
                    "PO-LEGACY-2201",
                    "",
                ),
                _col("vendor_code", "REQUIRED. Vendor must already exist.", "APX001", "APX001", "BLT001"),
                _col(
                    "part_number", "REQUIRED. Part/material must already exist.", "RM-AL-0250", "HW-10-32", "CON-TAPE"
                ),
                _col("quantity", "REQUIRED. Quantity ordered (> 0).", "500", "2000", "24"),
                _col("unit_price", "REQUIRED. Price per unit (>= 0).", "3.10", "0.08", "2.50"),
                _col(
                    "promised_date",
                    "Optional date (YYYY-MM-DD) the vendor promised delivery.",
                    "2026-06-20",
                    "2026-06-20",
                    "",
                ),
            ),
        ),
        ImportTemplateSpec(
            entity="bom",
            title="Bill of materials",
            description="BOM line items. POST /api/v1/bom/import/preview (review the mapping, then commit).",
            columns=(
                _col(
                    "part_number", "Component part number (created if missing on commit).", "1042-100", "HW-10-32", ""
                ),
                _col("description", "Component description.", "Bracket, Left", "10-32 locknut", "Loctite 242"),
                _col("quantity", "Quantity per assembly.", "2", "8", "0.1"),
                _col("unit_of_measure", "Optional. each, pounds, feet, ...", "each", "each", "each"),
                _col("item_type", "Optional. make, buy, or phantom.", "make", "buy", "buy"),
                _col(
                    "line_type",
                    "Optional. component, hardware, consumable, reference.",
                    "component",
                    "hardware",
                    "consumable",
                ),
            ),
        ),
    )
}


def list_import_templates() -> List[Dict[str, Any]]:
    """Summaries for the template index endpoint."""
    return [
        {
            "entity": spec.entity,
            "title": spec.title,
            "description": spec.description,
            "columns": [column.name for column in spec.columns],
            "download_path": f"/api/v1/import/templates/{spec.entity}",
        }
        for spec in IMPORT_TEMPLATES.values()
    ]


def build_import_template_workbook(entity: str) -> bytes:
    """Build the downloadable XLSX template for ``entity``.

    Layout contract (must stay in sync with :func:`parse_import_file`):

    * Sheet 1 ("Import"): styled header row, then ONE guidance row whose cells
      all start with ``#`` — the importer skips ``#``-prefixed rows.
    * Sheet 2 ("Examples"): the same header plus 2-3 realistic example rows.
      The importer only reads the first sheet, so examples can never be
      imported by accident.
    """
    spec = IMPORT_TEMPLATES.get(entity)
    if spec is None:
        raise KeyError(entity)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Werco brand: navy header, instrument-panel hairline borders (see export_service).
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1B4D9C", end_color="1B4D9C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )
    guidance_font = Font(italic=True, color="6B7280")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Import"

    for col_idx, column in enumerate(spec.columns, 1):
        cell = sheet.cell(row=1, column=col_idx, value=column.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        guidance = sheet.cell(row=2, column=col_idx, value=f"# {column.note}")
        guidance.font = guidance_font
        width = max(len(column.name) + 4, min(len(column.note) + 4, 46))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width
    sheet.freeze_panes = "A3"

    examples = workbook.create_sheet("Examples")
    example_count = max((len(column.examples) for column in spec.columns), default=0)
    for col_idx, column in enumerate(spec.columns, 1):
        cell = examples.cell(row=1, column=col_idx, value=column.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        examples.column_dimensions[get_column_letter(col_idx)].width = max(len(column.name) + 4, 14)
        for example_idx in range(example_count):
            value = column.examples[example_idx] if example_idx < len(column.examples) else ""
            examples.cell(row=2 + example_idx, column=col_idx, value=value)
    examples.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()
