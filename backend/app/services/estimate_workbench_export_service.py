"""Estimate Workbench exports — customer PDF + internal audit (Phase 6).

Customer PDF: sell price + line descriptions only (no rates, hours, confidence).
Internal audit: full fab/buyout/machined breakdown, hours, confidence, notes,
rate snapshot, verification — Excel workbook and/or JSON.
"""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional

from app.models.estimate_workbench import (
    QuoteAssembly,
    QuoteBuyoutLineItem,
    QuoteFabLineItem,
    QuoteMachinedLineItem,
)
from app.models.rfq_quote import QuoteEstimate, RfqPackage
from app.services.estimate_workbench_service import build_verification_report
from app.services.quote_pdf_service import build_customer_quote_pdf


class ExportBlockedError(Exception):
    def __init__(self, message: str, blockers: Optional[List[Dict[str, Any]]] = None):
        self.message = message
        self.blockers = blockers or []
        super().__init__(message)


def _alive(rows: Optional[List[Any]]) -> List[Any]:
    return [r for r in (rows or []) if not getattr(r, "is_deleted", False)]


def _money(n: Any) -> float:
    try:
        return round(float(n or 0), 4)
    except (TypeError, ValueError):
        return 0.0


def collect_workbench_export_payload(
    estimate: QuoteEstimate,
    *,
    package: Optional[RfqPackage] = None,
) -> Dict[str, Any]:
    """Full internal audit payload (JSON-serializable)."""
    verification = build_verification_report(estimate)
    breakdown = dict(estimate.internal_breakdown or {})

    assemblies_out: List[Dict[str, Any]] = []
    for asm in _alive(estimate.assemblies):
        assert isinstance(asm, QuoteAssembly)
        fab_rows = []
        for fl in _alive(asm.fab_line_items):
            assert isinstance(fl, QuoteFabLineItem)
            fab_rows.append(
                {
                    "id": fl.id,
                    "part_number": fl.part_number,
                    "detail_name": fl.detail_name,
                    "material": fl.material,
                    "qty": fl.qty,
                    "thickness_in": fl.thickness_in,
                    "width_in": fl.width_in,
                    "length_in": fl.length_in,
                    "cut_length_in": fl.cut_length_in,
                    "pierce_count": fl.pierce_count,
                    "bend_count": fl.bend_count,
                    "weld_length_in": fl.weld_length_in,
                    "weld_minutes_ea": fl.weld_minutes_ea,
                    "include_material": fl.include_material,
                    "include_laser": fl.include_laser,
                    "include_brake": fl.include_brake,
                    "include_weld": fl.include_weld,
                    "weight_ea_lb": fl.weight_ea_lb,
                    "material_cost": _money(fl.material_cost),
                    "laser_cost": _money(fl.laser_cost),
                    "laser_hours": _money(fl.laser_hours),
                    "brake_cost": _money(fl.brake_cost),
                    "brake_hours": _money(fl.brake_hours),
                    "weld_cost": _money(fl.weld_cost),
                    "weld_hours": _money(fl.weld_hours),
                    "line_total": _money(fl.line_total),
                    "confidence": fl.confidence,
                    "verification_note": fl.verification_note,
                    "calc_warnings": fl.calc_warnings,
                    "calc_errors": fl.calc_errors,
                }
            )
        buy_rows = []
        for bl in _alive(asm.buyout_line_items):
            assert isinstance(bl, QuoteBuyoutLineItem)
            buy_rows.append(
                {
                    "id": bl.id,
                    "part_number": bl.part_number,
                    "description": bl.description,
                    "category": bl.category,
                    "vendor": bl.vendor,
                    "qty": bl.qty,
                    "unit_cost": _money(bl.unit_cost),
                    "extended_cost": _money(bl.extended_cost),
                    "price_source": bl.price_source,
                    "confidence": bl.confidence,
                    "verification_note": bl.verification_note,
                }
            )
        assemblies_out.append(
            {
                "id": asm.id,
                "name": asm.name,
                "assembly_labor_hrs": asm.assembly_labor_hrs,
                "electrical_labor_hrs": asm.electrical_labor_hrs,
                "notes": asm.notes,
                "fab_lines": fab_rows,
                "buyout_lines": buy_rows,
            }
        )

    machined_out = []
    for mp in _alive(estimate.machined_line_items):
        assert isinstance(mp, QuoteMachinedLineItem)
        machined_out.append(
            {
                "id": mp.id,
                "part_number": mp.part_number,
                "description": mp.description,
                "material": mp.material,
                "qty": mp.qty,
                "stock_dia_in": mp.stock_dia_in,
                "stock_length_in": mp.stock_length_in,
                "turning_minutes": mp.turning_minutes,
                "milling_minutes": mp.milling_minutes,
                "weight_ea_lb": mp.weight_ea_lb,
                "material_cost": _money(mp.material_cost),
                "turning_cost": _money(mp.turning_cost),
                "turning_hours": _money(mp.turning_hours),
                "milling_cost": _money(mp.milling_cost),
                "milling_hours": _money(mp.milling_hours),
                "line_total": _money(mp.line_total),
                "confidence": mp.confidence,
                "verification_note": mp.verification_note,
            }
        )

    return {
        "exported_at": datetime.utcnow().isoformat() + "Z",
        "estimate_id": estimate.id,
        "rfq_package_id": estimate.rfq_package_id,
        "rfq_number": package.rfq_number if package else None,
        "customer_name": package.customer_name if package else None,
        "quote_id": estimate.quote_id,
        "version": estimate.version,
        "currency": estimate.currency,
        "totals": {
            "material": _money(estimate.material_total),
            "hardware_consumables": _money(estimate.hardware_consumables_total),
            "shop_labor_oh": _money(estimate.shop_labor_oh_total),
            "margin": _money(estimate.margin_total),
            "grand_total": _money(estimate.grand_total),
        },
        "bid_summary": {
            "laser_hours": breakdown.get("laser_hours"),
            "brake_hours": breakdown.get("brake_hours"),
            "weld_hours": breakdown.get("weld_hours"),
            "cogs": breakdown.get("cogs"),
            "sell_price": breakdown.get("sell_price") or estimate.grand_total,
            "target_margin": breakdown.get("target_margin"),
        },
        "rate_snapshot": breakdown.get("rate_snapshot"),
        "verification": verification,
        "assumptions": estimate.assumptions or [],
        "source_attribution": estimate.source_attribution or {},
        "assemblies": assemblies_out,
        "machined_parts": machined_out,
    }


def build_workbench_audit_json_bytes(
    estimate: QuoteEstimate,
    *,
    package: Optional[RfqPackage] = None,
) -> bytes:
    payload = collect_workbench_export_payload(estimate, package=package)
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


def build_workbench_audit_xlsx(
    estimate: QuoteEstimate,
    *,
    package: Optional[RfqPackage] = None,
) -> bytes:
    """Multi-sheet Excel: Summary, Fab, Buyout, Machined, Verification."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    payload = collect_workbench_export_payload(estimate, package=package)
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(bold=True, color="FFFFFF")

    summary_rows = [
        ("Estimate ID", payload["estimate_id"]),
        ("RFQ", payload.get("rfq_number") or payload["rfq_package_id"]),
        ("Customer", payload.get("customer_name") or ""),
        ("Quote ID", payload.get("quote_id") or ""),
        ("Version", payload.get("version")),
        ("Grand Total (sell)", payload["totals"]["grand_total"]),
        ("Material Total", payload["totals"]["material"]),
        ("Hardware / Buyout", payload["totals"]["hardware_consumables"]),
        ("Shop Labor + OH", payload["totals"]["shop_labor_oh"]),
        ("Margin", payload["totals"]["margin"]),
        ("Laser Hours", payload["bid_summary"].get("laser_hours")),
        ("Brake Hours", payload["bid_summary"].get("brake_hours")),
        ("Weld Hours", payload["bid_summary"].get("weld_hours")),
        ("COGS", payload["bid_summary"].get("cogs")),
        ("Target Margin", payload["bid_summary"].get("target_margin")),
        ("Exported At", payload["exported_at"]),
        ("Verification Status", payload["verification"].get("status")),
        ("Can Finalize", payload["verification"].get("can_finalize")),
        ("Review Count", payload["verification"].get("review_count")),
    ]
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in summary_rows:
        ws.append(list(row))

    rate_snap = payload.get("rate_snapshot") or {}
    if rate_snap:
        ws.append([])
        ws.append(["Rate Snapshot (frozen on finalize)", ""])
        for k, v in rate_snap.items():
            ws.append([k, v])

    # --- Fab ---
    fab_ws = wb.create_sheet("Fab Lines")
    fab_headers = [
        "assembly",
        "part_number",
        "detail_name",
        "material",
        "qty",
        "thickness_in",
        "width_in",
        "length_in",
        "cut_length_in",
        "pierce_count",
        "bend_count",
        "weld_length_in",
        "weight_ea_lb",
        "material_cost",
        "laser_cost",
        "laser_hours",
        "brake_cost",
        "brake_hours",
        "weld_cost",
        "weld_hours",
        "line_total",
        "confidence",
        "verification_note",
        "ops_material",
        "ops_laser",
        "ops_brake",
        "ops_weld",
    ]
    fab_ws.append(fab_headers)
    for cell in fab_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for asm in payload["assemblies"]:
        for fl in asm["fab_lines"]:
            fab_ws.append(
                [
                    asm["name"],
                    fl.get("part_number"),
                    fl.get("detail_name"),
                    fl.get("material"),
                    fl.get("qty"),
                    fl.get("thickness_in"),
                    fl.get("width_in"),
                    fl.get("length_in"),
                    fl.get("cut_length_in"),
                    fl.get("pierce_count"),
                    fl.get("bend_count"),
                    fl.get("weld_length_in"),
                    fl.get("weight_ea_lb"),
                    fl.get("material_cost"),
                    fl.get("laser_cost"),
                    fl.get("laser_hours"),
                    fl.get("brake_cost"),
                    fl.get("brake_hours"),
                    fl.get("weld_cost"),
                    fl.get("weld_hours"),
                    fl.get("line_total"),
                    fl.get("confidence"),
                    fl.get("verification_note"),
                    fl.get("include_material"),
                    fl.get("include_laser"),
                    fl.get("include_brake"),
                    fl.get("include_weld"),
                ]
            )

    # --- Buyout ---
    buy_ws = wb.create_sheet("Buyouts")
    buy_headers = [
        "assembly",
        "part_number",
        "description",
        "category",
        "vendor",
        "qty",
        "unit_cost",
        "extended_cost",
        "price_source",
        "confidence",
        "verification_note",
    ]
    buy_ws.append(buy_headers)
    for cell in buy_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for asm in payload["assemblies"]:
        for bl in asm["buyout_lines"]:
            buy_ws.append(
                [
                    asm["name"],
                    bl.get("part_number"),
                    bl.get("description"),
                    bl.get("category"),
                    bl.get("vendor"),
                    bl.get("qty"),
                    bl.get("unit_cost"),
                    bl.get("extended_cost"),
                    bl.get("price_source"),
                    bl.get("confidence"),
                    bl.get("verification_note"),
                ]
            )

    # --- Machined ---
    mach_ws = wb.create_sheet("Machined")
    mach_headers = [
        "part_number",
        "description",
        "material",
        "qty",
        "stock_dia_in",
        "stock_length_in",
        "turning_minutes",
        "milling_minutes",
        "weight_ea_lb",
        "material_cost",
        "turning_cost",
        "turning_hours",
        "milling_cost",
        "milling_hours",
        "line_total",
        "confidence",
        "verification_note",
    ]
    mach_ws.append(mach_headers)
    for cell in mach_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for mp in payload["machined_parts"]:
        mach_ws.append([mp.get(h) for h in mach_headers])

    # --- Verification ---
    ver_ws = wb.create_sheet("Verification")
    ver_ws.append(["category", "label", "confidence", "reason", "line_total", "anchor"])
    for cell in ver_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for action in payload["verification"].get("priority_actions") or []:
        ver_ws.append(
            [
                action.get("category"),
                action.get("label"),
                action.get("confidence"),
                action.get("reason"),
                action.get("line_total"),
                action.get("anchor"),
            ]
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_workbench_customer_line_summaries(estimate: QuoteEstimate) -> List[Dict[str, Any]]:
    """Customer-facing lines: description + qty + sell allocation (no rates/hours)."""
    lines: List[Dict[str, Any]] = []
    for asm in _alive(estimate.assemblies):
        for fl in _alive(asm.fab_line_items):
            label = fl.detail_name or "Fab detail"
            if fl.part_number:
                label = f"{fl.part_number} — {label}"
            lines.append(
                {
                    "part_display": f"{asm.name}: {label}",
                    "qty": fl.qty or 1,
                    "material": fl.material or None,
                    "thickness": f'{fl.thickness_in}"' if fl.thickness_in is not None else None,
                    "finish": None,
                    "part_total": _money(fl.line_total),
                }
            )
        for bl in _alive(asm.buyout_line_items):
            label = bl.description or "Buyout"
            if bl.part_number:
                label = f"{bl.part_number} — {label}"
            lines.append(
                {
                    "part_display": f"{asm.name}: {label}",
                    "qty": bl.qty or 1,
                    "material": None,
                    "thickness": None,
                    "finish": bl.category,
                    "part_total": _money(bl.extended_cost),
                }
            )
    for mp in _alive(estimate.machined_line_items):
        label = mp.description or "Machined part"
        if mp.part_number:
            label = f"{mp.part_number} — {label}"
        lines.append(
            {
                "part_display": label,
                "qty": mp.qty or 1,
                "material": mp.material,
                "thickness": None,
                "finish": None,
                "part_total": _money(mp.line_total),
            }
        )
    return lines


def build_workbench_customer_pdf(
    estimate: QuoteEstimate,
    *,
    package: Optional[RfqPackage] = None,
    quote_number: Optional[str] = None,
    require_clear_verification: bool = True,
) -> bytes:
    """Customer PDF from workbench. Blocks while Review items remain (unless finalized)."""
    report = build_verification_report(estimate)
    if require_clear_verification and not report["can_finalize"] and not estimate.quote_id:
        raise ExportBlockedError(
            "Cannot export customer PDF while Review items remain — finalize gate applies",
            blockers=report.get("blockers") or [],
        )

    breakdown = estimate.internal_breakdown or {}
    sell = float(breakdown.get("sell_price") or estimate.grand_total or 0)
    qn = quote_number or (f"EW-{estimate.id}" if not estimate.quote_id else f"QTE-EW-{estimate.id}")

    assumptions_raw = estimate.assumptions or []
    assumptions: List[Dict[str, Any]] = []
    for item in assumptions_raw:
        if isinstance(item, dict):
            assumptions.append(
                {
                    "field": item.get("field") or item.get("source") or "note",
                    "assumption": item.get("assumption") or item.get("note") or str(item),
                }
            )
        else:
            assumptions.append({"field": "note", "assumption": str(item)})

    return build_customer_quote_pdf(
        quote_number=qn,
        revision="A",
        customer_name=(package.customer_name if package else None) or "Customer",
        customer_contact=None,
        customer_email=None,
        rfq_reference=(package.rfq_number if package else None)
        or (package.rfq_reference if package else None),
        quote_date=datetime.utcnow().strftime("%Y-%m-%d"),
        valid_until=None,
        lead_time_label=None,
        total_amount=sell,
        line_summaries=build_workbench_customer_line_summaries(estimate),
        assumptions=assumptions,
        exclusions=[
            "Quote excludes taxes, freight, and duties unless stated otherwise.",
            "Subject to drawing/specification review at order entry.",
            "Internal machine rates, labor hours, and verification notes are omitted from this customer document.",
        ],
    )
