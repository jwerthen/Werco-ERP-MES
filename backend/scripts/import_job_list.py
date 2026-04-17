"""One-shot importer for the legacy Excel job list.

Reads an .xlsx with the columns:
    CUSTOMER | QTY | PART NUMBER | REV | SERIAL # | DESCRIPTION |
    WERCO JOB # | P.O. # | P.O. LINE ITEM # | DATE OF PO |
    REQ. DELIVERY DATE | INVOICE DATE | ACTUAL SHIP DATE

Filters out shipped rows (rows with ACTUAL SHIP DATE set), then creates:
  - Customer records (one per unique CUSTOMER value, scoped to --company-id)
  - Part records (one per unique PART NUMBER + REV, placeholder if blank)
  - WorkOrder records (one per row; WO numbers ending a suffix if the
    same Werco Job # appears on multiple rows, e.g. W1772-1, W1772-2)

Usage:
    # dry-run — prints the plan, no DB writes
    python -m scripts.import_job_list --xlsx /path/to/file.xlsx --company-id 1

    # commit (single transaction, rolls back on any error)
    python -m scripts.import_job_list --xlsx /path/to/file.xlsx --company-id 1 --commit
"""
import argparse
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, date

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from sqlalchemy.exc import IntegrityError

from app.db.database import SessionLocal
from app.models.company import Company
from app.models.customer import Customer
from app.models.part import Part, PartType, UnitOfMeasure
from app.models.work_order import WorkOrder, WorkOrderStatus


EXPECTED_HEADERS = [
    "CUSTOMER", "QTY", "PART NUMBER", "REV", "SERIAL #", "DESCRIPTION",
    "WERCO JOB #", "P.O. #", "P.O. LINE ITEM #", "DATE OF PO",
    "REQ. DELIVERY DATE", "INVOICE DATE", "ACTUAL SHIP DATE",
]


def _s(v):
    """Normalize cell value to trimmed string or None."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v).strip() or None


def _to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _slug_code(name: str) -> str:
    """Slug for Customer.code — upper-case alnum+dash, max 50 chars."""
    s = re.sub(r"[^A-Z0-9]+", "-", (name or "").upper()).strip("-")
    return s[:50]


def _placeholder_part_number(customer: str, desc: str) -> str:
    base = re.sub(r"[^A-Z0-9]+", "-", f"{customer}-{desc or 'ITEM'}".upper()).strip("-")
    return f"PLACEHOLDER-{base[:80]}"


def _read_rows(xlsx_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [_s(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]
    # validate required headers are present in the first row
    for h in EXPECTED_HEADERS:
        if h not in headers:
            raise SystemExit(f"Missing expected column in xlsx: {h!r}. Found: {headers}")

    idx = {h: headers.index(h) for h in EXPECTED_HEADERS}

    rows = []
    for r in range(2, ws.max_row + 1):
        cells = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in cells):
            continue
        rows.append({
            "_row": r,
            "customer": _s(cells[idx["CUSTOMER"]]),
            "qty": cells[idx["QTY"]],
            "part_number": _s(cells[idx["PART NUMBER"]]),
            "rev": _s(cells[idx["REV"]]),
            "serial": _s(cells[idx["SERIAL #"]]),
            "description": _s(cells[idx["DESCRIPTION"]]),
            "wo_number": _s(cells[idx["WERCO JOB #"]]),
            "po_number": _s(cells[idx["P.O. #"]]),
            "po_line_item": _s(cells[idx["P.O. LINE ITEM #"]]),
            "po_date": _to_date(cells[idx["DATE OF PO"]]),
            "due_date": _to_date(cells[idx["REQ. DELIVERY DATE"]]),
            "invoice_date": _to_date(cells[idx["INVOICE DATE"]]),
            "ship_date": _to_date(cells[idx["ACTUAL SHIP DATE"]]),
        })
    return rows


def _filter_open(rows):
    """Keep rows that have NOT shipped."""
    return [r for r in rows if r["ship_date"] is None]


def _apply_wo_suffixes(rows):
    """When the same WERCO JOB # appears on >1 row, suffix -1, -2, ... in
    spreadsheet order so the resulting WO numbers are unique per company."""
    counts = defaultdict(int)
    for r in rows:
        counts[r["wo_number"]] += 1

    seen = defaultdict(int)
    for r in rows:
        wo = r["wo_number"]
        if counts[wo] > 1:
            seen[wo] += 1
            r["final_wo_number"] = f"{wo}-{seen[wo]}"
        else:
            r["final_wo_number"] = wo
    return rows


def build_plan(rows, db, company_id: int):
    """Return (customers_to_create, parts_to_create, wos_to_create, warnings)."""
    warnings = []

    # Existing lookups
    existing_customers = {
        c.name: c for c in db.query(Customer).filter(Customer.company_id == company_id).all()
    }
    existing_parts = {
        (p.part_number, (p.revision or "")): p
        for p in db.query(Part).filter(Part.company_id == company_id).all()
    }
    existing_wo_numbers = {
        wo for (wo,) in db.query(WorkOrder.work_order_number).filter(
            WorkOrder.company_id == company_id
        ).all()
    }

    customers_to_create = {}  # name -> dict
    parts_to_create = {}       # (pn, rev) -> dict
    wos_to_create = []

    for r in rows:
        cust = r["customer"]
        if not cust:
            warnings.append(f"Row {r['_row']}: blank CUSTOMER, skipping")
            continue
        if cust not in existing_customers and cust not in customers_to_create:
            customers_to_create[cust] = {"name": cust, "code": _slug_code(cust)}

        pn = r["part_number"]
        rev = r["rev"] or ""
        if not pn:
            # Placeholder part: synthesize number from customer + description
            pn = _placeholder_part_number(cust, r["description"])
            r["part_number"] = pn
            r["rev"] = rev
            warnings.append(
                f"Row {r['_row']} (WO {r['wo_number']}): blank part #, "
                f"creating placeholder {pn!r}"
            )

        key = (pn, rev)
        if key not in existing_parts and key not in parts_to_create:
            parts_to_create[key] = {
                "part_number": pn,
                "revision": rev or "",
                "name": r["description"] or pn,
                "description": r["description"],
                "part_type": PartType.MANUFACTURED,
                "customer_name": cust,
            }

        final_wo = r["final_wo_number"]
        if final_wo in existing_wo_numbers:
            warnings.append(
                f"Row {r['_row']}: WO {final_wo} already exists in company "
                f"{company_id} — row will be SKIPPED on commit"
            )
            r["_skip"] = True
            continue

        try:
            qty = float(r["qty"]) if r["qty"] is not None else 0.0
        except (TypeError, ValueError):
            warnings.append(f"Row {r['_row']}: non-numeric QTY {r['qty']!r}, using 0")
            qty = 0.0

        wos_to_create.append({
            "work_order_number": final_wo,
            "original_wo": r["wo_number"],
            "part_key": key,
            "customer_name": cust,
            "customer_po": r["po_number"],
            "po_line_item": r["po_line_item"],
            "po_date": r["po_date"],
            "quantity_ordered": qty,
            "due_date": r["due_date"],
            "serial": r["serial"],
            "description": r["description"],
            "row": r["_row"],
        })

    return customers_to_create, parts_to_create, wos_to_create, warnings


def print_plan(customers, parts, wos, warnings, open_count, total_count):
    print("=" * 72)
    print("IMPORT PLAN")
    print("=" * 72)
    print(f"Rows in file          : {total_count}")
    print(f"Shipped (filtered out): {total_count - open_count}")
    print(f"Open rows to import   : {open_count}")
    print()
    print(f"New Customers to create : {len(customers)}")
    for name, c in sorted(customers.items()):
        print(f"   - {name}  (code={c['code']})")
    print()
    print(f"New Parts to create     : {len(parts)}")
    for (pn, rev), p in sorted(parts.items()):
        print(f"   - {pn} rev={rev or '-'}  name={p['name']!r}")
    print()
    print(f"Work Orders to create   : {len(wos)}")
    # Preview first 10
    for w in wos[:10]:
        pn, rev = w["part_key"]
        print(f"   - {w['work_order_number']:<12}  "
              f"{w['customer_name']:<20}  "
              f"qty={w['quantity_ordered']:<6}  "
              f"part={pn} rev={rev or '-'}  "
              f"PO={w['customer_po']}/{w['po_line_item']}  "
              f"due={w['due_date']}")
    if len(wos) > 10:
        print(f"   ... ({len(wos) - 10} more)")
    print()
    print(f"Warnings                : {len(warnings)}")
    for w in warnings:
        print(f"   ! {w}")
    print("=" * 72)


def commit_plan(db, company_id, created_by, customers, parts, wos):
    customer_map = {
        c.name: c for c in db.query(Customer).filter(Customer.company_id == company_id).all()
    }
    for name, data in customers.items():
        c = Customer(company_id=company_id, name=data["name"], code=data["code"])
        db.add(c)
        customer_map[name] = c
    db.flush()

    part_map = {
        (p.part_number, (p.revision or "")): p
        for p in db.query(Part).filter(Part.company_id == company_id).all()
    }
    for key, data in parts.items():
        p = Part(
            company_id=company_id,
            part_number=data["part_number"],
            revision=data["revision"],
            name=data["name"][:255],
            description=data["description"],
            part_type=data["part_type"],
            unit_of_measure=UnitOfMeasure.EACH,
            customer_name=data["customer_name"],
            created_by=created_by,
        )
        db.add(p)
        part_map[key] = p
    db.flush()

    for w in wos:
        part = part_map[w["part_key"]]
        wo = WorkOrder(
            company_id=company_id,
            work_order_number=w["work_order_number"],
            part_id=part.id,
            quantity_ordered=w["quantity_ordered"],
            status=WorkOrderStatus.RELEASED,
            priority=5,
            due_date=w["due_date"],
            customer_name=w["customer_name"],
            customer_po=w["customer_po"],
            po_line_item=w["po_line_item"],
            po_date=w["po_date"],
            serial_numbers=w["serial"],
            notes=w["description"],
            created_by=created_by,
            released_by=created_by,
            released_at=datetime.utcnow(),
        )
        db.add(wo)
    db.flush()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to the xlsx job list")
    parser.add_argument("--company-id", type=int, required=True,
                        help="Tenant company_id to import under")
    parser.add_argument("--created-by", type=int, default=None,
                        help="User id to stamp on created_by / released_by")
    parser.add_argument("--commit", action="store_true",
                        help="Actually write to the DB (default is dry-run)")
    args = parser.parse_args()

    db = SessionLocal()
    try:
        company = db.query(Company).filter(Company.id == args.company_id).first()
        if not company:
            raise SystemExit(f"No company with id={args.company_id}")
        print(f"Target company: {company.name} (id={company.id}, slug={company.slug})")
        print(f"Source xlsx   : {args.xlsx}")
        print()

        all_rows = _read_rows(args.xlsx)
        open_rows = _filter_open(all_rows)
        open_rows = _apply_wo_suffixes(open_rows)

        customers, parts, wos, warnings = build_plan(open_rows, db, args.company_id)
        print_plan(customers, parts, wos, warnings,
                   open_count=len(open_rows), total_count=len(all_rows))

        if not args.commit:
            print("\nDRY-RUN — no changes written. Re-run with --commit to apply.")
            return

        print("\nCOMMITTING...")
        commit_plan(db, args.company_id, args.created_by, customers, parts, wos)
        db.commit()
        print(
            f"Done. Created {len(customers)} customers, {len(parts)} parts, "
            f"{len(wos)} work orders."
        )
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
