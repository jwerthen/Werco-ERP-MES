"""Unit tests for the shared Excel/CSV import kit (A0.2).

Covers the parsing edge cases the migration depends on: header normalization,
duplicate-header rejection, blank-row tolerance, ``"# "`` guidance-row skipping
(while ``#``-prefixed part numbers like ``#10-32X1/2`` stay importable),
defensive cell coercion (Excel floats/dates/bools, numbers-as-text), size/row
caps, the bounded-scan limits (column cap, blank-run cutoff with its loud
look-ahead — data past the gap refuses the file rather than silently
truncating — and the scanned-row backstop), and the invariant that every
server-generated template round-trips through the parser.
"""

import io
import time
from datetime import date, datetime

import pytest
from openpyxl import Workbook, load_workbook

from app.services.import_service import (
    IMPORT_TEMPLATES,
    MAX_CONSECUTIVE_BLANK_ROWS,
    MAX_IMPORT_COLUMNS,
    ImportFileError,
    build_import_template_workbook,
    coerce_cell,
    list_import_templates,
    normalize_import_header,
    parse_date_field,
    parse_import_file,
)


def _xlsx_bytes(rows, sheet_rows_extra=None):
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    if sheet_rows_extra:
        extra = workbook.create_sheet("Second")
        for row in sheet_rows_extra:
            extra.append(row)
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


@pytest.mark.unit
class TestHeaderNormalization:
    def test_strip_lower_snake_case(self):
        assert normalize_import_header(" Part Number ") == "part_number"
        assert normalize_import_header("PART-NUMBER") == "part_number"
        assert normalize_import_header("Lead Time (Days)") == "lead_time_days"
        assert normalize_import_header("customer/po") == "customer_po"

    def test_non_string_headers(self):
        assert normalize_import_header(42) == "42"
        assert normalize_import_header(None) == ""


@pytest.mark.unit
class TestCellCoercion:
    def test_integral_float_becomes_int_string(self):
        assert coerce_cell(5.0) == "5"
        assert coerce_cell(10.0) == "10"

    def test_non_integral_float_keeps_decimals(self):
        assert coerce_cell(2.5) == "2.5"

    def test_bool_and_none(self):
        assert coerce_cell(True) == "true"
        assert coerce_cell(False) == "false"
        assert coerce_cell(None) == ""

    def test_dates(self):
        assert coerce_cell(datetime(2026, 7, 15)) == "2026-07-15"  # date-only Excel cell
        assert coerce_cell(datetime(2026, 7, 15, 7, 30)) == "2026-07-15 07:30:00"
        assert coerce_cell(date(2026, 7, 15)) == "2026-07-15"

    def test_text_is_stripped(self):
        assert coerce_cell("  42 ") == "42"


@pytest.mark.unit
class TestParseDateField:
    def test_accepted_formats(self):
        assert parse_date_field("2026-07-15", "due_date") == date(2026, 7, 15)
        assert parse_date_field("07/15/2026", "due_date") == date(2026, 7, 15)
        assert parse_date_field("2026-07-15 07:30:00", "due_date") == date(2026, 7, 15)
        assert parse_date_field("", "due_date") is None

    def test_invalid_raises(self):
        with pytest.raises(ValueError, match="due_date"):
            parse_date_field("July fifteen", "due_date")


@pytest.mark.unit
class TestParseImportFileXlsx:
    def test_happy_path_with_blank_and_guidance_rows(self):
        content = _xlsx_bytes(
            [
                [None, None, None],  # leading blank row tolerated
                ["Part Number", "Name", "Lead Time Days"],
                ["# REQUIRED. Unique part number.", "# REQUIRED.", "# Optional."],
                ["P-100", "Bracket", 10.0],
                [None, None, None],  # blank row mid-file tolerated
                ["P-200", "Plate", " 12 "],
            ]
        )
        table = parse_import_file("upload.xlsx", content, required_columns={"part_number", "name"})
        assert table.headers == ["part_number", "name", "lead_time_days"]
        rows = list(table.iter_rows())
        assert len(rows) == 2
        first_row_number, first = rows[0]
        assert first == {"part_number": "P-100", "name": "Bracket", "lead_time_days": "10"}
        assert first_row_number == 4  # real file row number for error reporting
        _, second = rows[1]
        assert second["lead_time_days"] == "12"

    def test_only_first_sheet_is_read(self):
        content = _xlsx_bytes(
            [["part_number", "name"], ["P-1", "One"]],
            sheet_rows_extra=[["part_number", "name"], ["P-EXAMPLE", "Should not import"]],
        )
        table = parse_import_file("upload.xlsx", content)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["P-1"]

    def test_trailing_unnamed_columns_ignored(self):
        content = _xlsx_bytes([["part_number", "name", None], ["P-1", "One", "stray"]])
        table = parse_import_file("upload.xlsx", content)
        _, row = next(table.iter_rows())
        assert row == {"part_number": "P-1", "name": "One"}

    def test_typed_cells_coerced(self):
        content = _xlsx_bytes(
            [
                ["part_number", "quantity", "due_date", "is_critical"],
                ["P-1", 25.0, datetime(2026, 7, 15), True],
            ]
        )
        _, row = next(parse_import_file("u.xlsx", content).iter_rows())
        assert row == {"part_number": "P-1", "quantity": "25", "due_date": "2026-07-15", "is_critical": "true"}

    def test_corrupt_xlsx_rejected(self):
        with pytest.raises(ImportFileError, match="xlsx"):
            parse_import_file("upload.xlsx", b"this is not a zip archive")


@pytest.mark.unit
class TestGuidanceRowMarker:
    """Only the exact ``"# "`` template marker skips a row — ``#``-prefixed
    part numbers (e.g. ``#10-32X1/2``) are real data and must import."""

    def test_hash_prefixed_part_number_is_imported_not_skipped(self):
        content = _xlsx_bytes(
            [
                ["part_number", "quantity"],
                ["# REQUIRED. Part must exist.", "# REQUIRED."],  # template guidance row
                ["#10-32X1/2", 500.0],  # real hardware part number
                ["#8-32X3/8", 250.0],
            ]
        )
        table = parse_import_file("upload.xlsx", content)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["#10-32X1/2", "#8-32X3/8"]

    def test_hash_prefixed_part_number_in_csv(self):
        content = b"part_number,quantity\n# guidance row,skipped\n#10-32X1/2,500\n"
        table = parse_import_file("upload.csv", content)
        rows = list(table.iter_rows())
        assert len(rows) == 1
        assert rows[0][1] == {"part_number": "#10-32X1/2", "quantity": "500"}

    def test_bare_hash_without_space_is_data(self):
        content = b"part_number,name\n#SPECIAL,Hash part\n"
        table = parse_import_file("upload.csv", content)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["#SPECIAL"]


@pytest.mark.unit
class TestDuplicateHeaders:
    """Two columns that collide after normalization must fail loudly at parse
    time — silently merging them is data loss in a migration tool."""

    def test_distinct_headers_colliding_after_normalization_rejected(self):
        content = b"Part Number,part-number,name\nP-1,P-2,Bracket\n"
        with pytest.raises(ImportFileError) as excinfo:
            parse_import_file("upload.csv", content)
        message = str(excinfo.value)
        assert "Part Number" in message and "part-number" in message and "part_number" in message

    def test_exact_duplicate_header_rejected(self):
        content = _xlsx_bytes([["name", "name"], ["One", "Two"]])
        with pytest.raises(ImportFileError, match="Duplicate column"):
            parse_import_file("upload.xlsx", content)

    def test_multiple_unnamed_trailing_columns_still_tolerated(self):
        # Empty header cells are ignored, not treated as colliding duplicates.
        content = _xlsx_bytes([["part_number", None, None], ["P-1", "x", "y"]])
        table = parse_import_file("upload.xlsx", content)
        _, row = next(table.iter_rows())
        assert row == {"part_number": "P-1"}


@pytest.mark.unit
class TestParseImportFileCsv:
    def test_utf8_sig_and_blank_rows(self):
        content = ("﻿" + "Part Number,Name\n\nP-100,Bracket\n# guidance,skipped\n").encode("utf-8")
        table = parse_import_file("upload.csv", content)
        rows = list(table.iter_rows())
        assert len(rows) == 1
        assert rows[0][1] == {"part_number": "P-100", "name": "Bracket"}

    def test_non_utf8_rejected(self):
        with pytest.raises(ImportFileError, match="UTF-8"):
            parse_import_file("upload.csv", "pärt".encode("latin-1"))


@pytest.mark.unit
class TestParseImportFileGuards:
    def test_unsupported_extension(self):
        with pytest.raises(ImportFileError, match="CSV or Excel"):
            parse_import_file("upload.xls", b"x")
        with pytest.raises(ImportFileError, match="CSV or Excel"):
            parse_import_file(None, b"x")

    def test_empty_file(self):
        with pytest.raises(ImportFileError, match="empty"):
            parse_import_file("upload.csv", b"")

    def test_size_cap(self):
        with pytest.raises(ImportFileError, match="too large"):
            parse_import_file("upload.csv", b"a,b\n1,2\n", max_bytes=4)

    def test_row_cap(self):
        content = b"a,b\n" + b"1,2\n" * 5
        with pytest.raises(ImportFileError, match="Too many rows"):
            parse_import_file("upload.csv", content, max_rows=3)

    def test_missing_required_columns(self):
        with pytest.raises(ImportFileError, match="part_number"):
            parse_import_file("upload.csv", b"name\nBracket\n", required_columns={"part_number", "name"})

    def test_header_only_blank_file(self):
        with pytest.raises(ImportFileError, match="header"):
            parse_import_file("upload.csv", b"\n\n\n")


@pytest.mark.unit
class TestBoundedScan:
    """The parser must never iterate a workbook's full *declared* grid: one stray
    formatted/whitespace cell at XFD1048576 used to make a 5 KB upload scan
    16,384 x 1,048,576 cells (~5 minutes of CPU on the event loop in prod).

    The bounds must also never *silently* drop data: the blank-run cutoff is
    followed by a bounded look-ahead that refuses the file loudly when real
    data exists past the gap."""

    def test_bloated_used_range_parses_fast_and_succeeds(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["part_number", "name"])
        sheet.append(["P-1", "One"])
        sheet.append(["P-2", "Two"])
        # The production failure mode: a single stray whitespace cell in the very
        # last cell of the grid bloats the declared used range to the maximum.
        sheet.cell(row=1_048_576, column=16_384, value=" ")
        out = io.BytesIO()
        workbook.save(out)

        started = time.monotonic()
        # max_scanned_rows tightened so the cutoff check is deterministic: with
        # the blank-run cutoff working, the main loop scans ~1k rows (the 50k
        # look-ahead runs outside this budget); if the cutoff ever regresses to
        # unbounded blank scanning, the backstop raises ImportFileError here
        # regardless of runner speed.
        table = parse_import_file(
            "upload.xlsx",
            out.getvalue(),
            required_columns={"part_number", "name"},
            max_scanned_rows=60_000,
        )
        elapsed = time.monotonic() - started

        assert [row["part_number"] for _, row in table.iter_rows()] == ["P-1", "P-2"]
        # Wall-clock is only a loose backstop — coverage-traced CI runners take
        # ~19s for the bounded scan, a full-grid regression takes many minutes.
        # Column-cap regressions are caught deterministically by
        # test_sheet_wider_than_column_cap_still_parses.
        assert elapsed < 90, f"bloated-dimension parse took {elapsed:.1f}s — grid scan regression"

    def test_sheet_wider_than_column_cap_still_parses(self):
        extra_headers = [f"extra_{i}" for i in range(MAX_IMPORT_COLUMNS)]
        content = _xlsx_bytes(
            [
                ["part_number", "name"] + extra_headers,
                ["P-1", "One"] + ["x"] * MAX_IMPORT_COLUMNS,
            ]
        )
        table = parse_import_file("upload.xlsx", content, required_columns={"part_number", "name"})
        _, row = next(table.iter_rows())
        assert row["part_number"] == "P-1"
        assert row["name"] == "One"
        # Columns beyond the cap are ignored, not an error.
        assert len(table.headers) == MAX_IMPORT_COLUMNS
        assert f"extra_{MAX_IMPORT_COLUMNS - 3}" in table.headers  # last column inside the cap
        assert f"extra_{MAX_IMPORT_COLUMNS - 2}" not in table.headers  # first column past the cap

    def test_csv_columns_sliced_to_cap(self):
        header = ",".join(["part_number"] + [f"c{i}" for i in range(MAX_IMPORT_COLUMNS + 10)])
        data = ",".join(["P-1"] + ["x"] * (MAX_IMPORT_COLUMNS + 10))
        table = parse_import_file("upload.csv", f"{header}\n{data}\n".encode())
        assert len(table.headers) == MAX_IMPORT_COLUMNS
        _, row = next(table.iter_rows())
        assert row["part_number"] == "P-1"

    def test_data_after_blank_run_cutoff_refuses_file_loudly(self):
        """A run of blank rows longer than the cutoff is treated as end of data —
        but if real data sits past the gap (user cleared a block mid-sheet and
        kept rows below), the parser must refuse the whole file, not silently
        truncate it: silent truncation is data loss in a migration tool."""
        content = b"part_number\nP-1\n" + b"\n" * (MAX_CONSECUTIVE_BLANK_ROWS + 1) + b"P-AFTER-GAP\n"
        with pytest.raises(ImportFileError) as excinfo:
            parse_import_file("upload.csv", content)
        message = str(excinfo.value)
        assert "gap of more than" in message
        assert f"{MAX_CONSECUTIVE_BLANK_ROWS:,} blank rows" in message
        # Header=1, P-1=2, blanks 3..1003 (cutoff), so the post-gap data is row 1004.
        assert "row 1004" in message

    def test_data_beyond_lookahead_window_is_end_of_data(self):
        """Past the bounded look-ahead the gap really is end of data (innocent
        used-range bloat must stay a fast, clean parse). The window is shrunk
        via the kwarg so the fixture stays small, mirroring max_scanned_rows."""
        gap = MAX_CONSECUTIVE_BLANK_ROWS + 1 + 20  # cutoff fires, then 20 more blanks
        content = b"part_number\nP-1\n" + b"\n" * gap + b"P-FAR-AWAY\n"
        table = parse_import_file("upload.csv", content, blank_run_lookahead_rows=10)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["P-1"]

    def test_lookahead_does_not_trip_scanned_row_backstop(self):
        """The post-gap look-ahead runs outside the max_scanned_rows budget: a
        file that is nothing but trailing blanks after the cutoff must stay a
        clean end-of-data, not become the 'used range is enormous' error just
        because the look-ahead crosses that count."""
        content = b"part_number\nP-1\n" + b"\n" * 2000  # cutoff at row 1003 < 1500; look-ahead crosses 1500
        table = parse_import_file("upload.csv", content, max_scanned_rows=1500)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["P-1"]

    def test_header_not_found_after_cutoff_mentions_cutoff(self):
        """When the blank-run cutoff fired before any header was seen, the plain
        'must include a header row' message would be misleading — say why."""
        content = b"\n" * (MAX_CONSECUTIVE_BLANK_ROWS + 5)
        with pytest.raises(ImportFileError) as excinfo:
            parse_import_file("upload.csv", content)
        message = str(excinfo.value)
        assert "header row" in message
        assert f"{MAX_CONSECUTIVE_BLANK_ROWS:,} consecutive blank rows" in message

    def test_blank_runs_at_the_cutoff_still_tolerated(self):
        content = b"part_number\n" + b"\n" * MAX_CONSECUTIVE_BLANK_ROWS + b"P-1\n"
        table = parse_import_file("upload.csv", content)
        assert [row["part_number"] for _, row in table.iter_rows()] == ["P-1"]

    def test_scanned_row_backstop_raises_actionable_error(self):
        # Alternating data/blank rows defeat the consecutive-blank cutoff; the
        # hard scanned-row cap still bounds the scan (kwarg mirrors max_rows so
        # the test doesn't need 100k real rows).
        content = b"part_number\n" + b"P-1\n\n" * 10
        with pytest.raises(ImportFileError, match="used range is enormous"):
            parse_import_file("upload.csv", content, max_scanned_rows=10)


@pytest.mark.unit
class TestImportTemplates:
    def test_listing_covers_all_entities(self):
        listed = {item["entity"] for item in list_import_templates()}
        assert listed == set(IMPORT_TEMPLATES)
        assert "work-orders" in listed and "purchase-orders" in listed

    @pytest.mark.parametrize("entity", sorted(IMPORT_TEMPLATES))
    def test_every_template_round_trips_through_the_parser(self, entity):
        """The template's own header + guidance rows must parse to zero data rows."""
        content = build_import_template_workbook(entity)
        table = parse_import_file(f"{entity}.xlsx", content)
        assert table.headers == [column.name for column in IMPORT_TEMPLATES[entity].columns]
        assert len(list(table.iter_rows())) == 0  # guidance row skipped, examples on separate sheet

    def test_template_structure(self):
        workbook = load_workbook(io.BytesIO(build_import_template_workbook("work-orders")))
        assert workbook.sheetnames == ["Import", "Examples"]
        sheet = workbook["Import"]
        assert sheet.cell(row=1, column=1).value == "wo_number"
        assert str(sheet.cell(row=2, column=1).value).startswith("#")
        examples = workbook["Examples"]
        assert examples.cell(row=1, column=2).value == "part_number"
        assert examples.cell(row=2, column=2).value  # at least one example row

    def test_unknown_entity_raises(self):
        with pytest.raises(KeyError):
            build_import_template_workbook("not-an-entity")
