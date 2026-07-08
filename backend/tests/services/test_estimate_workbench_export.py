"""Phase 6 — workbench customer PDF + internal audit export."""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.models.estimate_workbench import (
    ConfidenceLevel,
    QuoteAssembly,
    QuoteFabLineItem,
)
from app.models.rfq_quote import QuoteEstimate, RfqPackage
from app.services.estimate_workbench_export_service import (
    ExportBlockedError,
    build_workbench_audit_json_bytes,
    build_workbench_audit_xlsx,
    build_workbench_customer_pdf,
)
from app.services.estimate_workbench_service import seed_cut_bend_defaults


def _seed_estimate(
    db_session,
    *,
    confidence: str = ConfidenceLevel.CONFIRMED.value,
    rfq_suffix: str = "1",
) -> QuoteEstimate:
    seed_cut_bend_defaults(db_session, company_id=1, force=False)
    pkg = RfqPackage(
        rfq_number=f"RFQ-EXP-{rfq_suffix}",
        customer_name="Export Customer",
        status="uploaded",
        company_id=1,
    )
    db_session.add(pkg)
    db_session.flush()
    est = QuoteEstimate(
        rfq_package_id=pkg.id,
        version=1,
        currency="USD",
        company_id=1,
        grand_total=1250.0,
        material_total=200.0,
        hardware_consumables_total=50.0,
        shop_labor_oh_total=700.0,
        margin_total=300.0,
        internal_breakdown={
            "laser_hours": 0.5,
            "brake_hours": 0.2,
            "weld_hours": 0.0,
            "cogs": 950.0,
            "sell_price": 1250.0,
            "target_margin": 0.3,
            "rate_snapshot": {"laser_rate": 185, "brake_rate": 95},
        },
        assumptions=[{"source": "test", "note": "Unit test estimate"}],
    )
    db_session.add(est)
    db_session.flush()
    asm = QuoteAssembly(
        quote_estimate_id=est.id,
        name="Assembly 1",
        sort_order=0,
        company_id=1,
    )
    db_session.add(asm)
    db_session.flush()
    db_session.add(
        QuoteFabLineItem(
            assembly_id=asm.id,
            sort_order=0,
            detail_name="Panel",
            part_number="DET-1",
            material="A36 Mild Steel",
            qty=2,
            thickness_in=0.075,
            width_in=10,
            length_in=12,
            cut_length_in=44,
            bend_count=4,
            pierce_count=2,
            material_cost=20.0,
            laser_cost=40.0,
            laser_hours=0.25,
            brake_cost=30.0,
            brake_hours=0.1,
            weld_cost=0.0,
            weld_hours=0.0,
            line_total=90.0,
            confidence=confidence,
            verification_note="Checked" if confidence != ConfidenceLevel.REVIEW.value else "Needs look",
            company_id=1,
        )
    )
    db_session.commit()
    db_session.refresh(est)
    return est


def test_audit_xlsx_and_json_contain_fab_and_snapshot(db_session):
    est = _seed_estimate(db_session)
    # Reload with relationships
    from app.services.estimate_workbench_service import get_estimate_tree

    tree = get_estimate_tree(db_session, est.id, 1)
    pkg = db_session.query(RfqPackage).filter(RfqPackage.id == est.rfq_package_id).one()

    xlsx = build_workbench_audit_xlsx(tree, package=pkg)
    wb = load_workbook(BytesIO(xlsx))
    assert set(wb.sheetnames) >= {"Summary", "Fab Lines", "Buyouts", "Machined", "Verification"}
    fab = wb["Fab Lines"]
    headers = [c.value for c in fab[1]]
    assert "confidence" in headers
    assert "laser_hours" in headers
    assert fab.max_row >= 2

    raw = build_workbench_audit_json_bytes(tree, package=pkg)
    import json

    payload = json.loads(raw.decode("utf-8"))
    assert payload["estimate_id"] == est.id
    assert payload["rate_snapshot"]["laser_rate"] == 185
    assert payload["assemblies"][0]["fab_lines"][0]["detail_name"] == "Panel"


def test_customer_pdf_blocked_on_review(db_session):
    est = _seed_estimate(db_session, confidence=ConfidenceLevel.REVIEW.value)
    from app.services.estimate_workbench_service import get_estimate_tree

    tree = get_estimate_tree(db_session, est.id, 1)
    pkg = db_session.query(RfqPackage).filter(RfqPackage.id == est.rfq_package_id).one()
    try:
        build_workbench_customer_pdf(tree, package=pkg, require_clear_verification=True)
        assert False, "expected ExportBlockedError"
    except ExportBlockedError as exc:
        assert "Review" in exc.message


def test_customer_pdf_ok_when_confirmed(db_session):
    est = _seed_estimate(db_session, confidence=ConfidenceLevel.CONFIRMED.value)
    from app.services.estimate_workbench_service import get_estimate_tree

    tree = get_estimate_tree(db_session, est.id, 1)
    pkg = db_session.query(RfqPackage).filter(RfqPackage.id == est.rfq_package_id).one()
    pdf = build_workbench_customer_pdf(tree, package=pkg, require_clear_verification=True)
    assert pdf[:4] == b"%PDF"
    assert b"laser_rate" not in pdf
    assert b"verification_note" not in pdf


def test_export_api_endpoints(client, admin_headers, db_session):
    est = _seed_estimate(db_session, confidence=ConfidenceLevel.CONFIRMED.value, rfq_suffix="api-ok")

    xlsx = client.get(
        f"/api/v1/estimate-workbench/{est.id}/export/audit.xlsx",
        headers=admin_headers,
    )
    assert xlsx.status_code == 200, xlsx.text
    assert "spreadsheetml" in xlsx.headers.get("content-type", "")

    js = client.get(
        f"/api/v1/estimate-workbench/{est.id}/export/audit.json",
        headers=admin_headers,
    )
    assert js.status_code == 200
    assert js.json()["estimate_id"] == est.id

    pdf = client.get(
        f"/api/v1/estimate-workbench/{est.id}/export/customer.pdf",
        headers=admin_headers,
    )
    assert pdf.status_code == 200, pdf.text
    assert pdf.content[:4] == b"%PDF"

    # Review blocks customer PDF
    review_est = _seed_estimate(
        db_session, confidence=ConfidenceLevel.REVIEW.value, rfq_suffix="api-rev"
    )
    blocked = client.get(
        f"/api/v1/estimate-workbench/{review_est.id}/export/customer.pdf",
        headers=admin_headers,
    )
    assert blocked.status_code == 422
