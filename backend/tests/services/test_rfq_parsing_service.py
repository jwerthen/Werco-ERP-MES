import time
from pathlib import Path
from types import SimpleNamespace

import ezdxf
import pytest
from openpyxl import Workbook

from app.services import rfq_parsing_service as rfq_parser
from app.services.import_service import MAX_CONSECUTIVE_BLANK_ROWS, ImportFileError
from app.services.rfq_parsing_service import (
    build_normalized_part_specs,
    parse_bom_xlsx,
    parse_dxf_geometry,
    parse_rfq_package_files,
)


def test_parse_bom_xlsx_extracts_parts_and_hardware(tmp_path: Path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(
        [
            "Part Number",
            "Description",
            "Qty",
            "Material",
            "Thickness",
            "Finish",
            "Type",
            "Flat Length",
            "Flat Width",
        ]
    )
    sheet.append(["P-100", "Main Bracket", 4, "A36 Steel", "0.125", "Powder Coat", "Part", 12, 8])
    sheet.append(["HW-200", "PEM Nut 1/4-20", 8, "", "", "", "Hardware"])

    file_path = tmp_path / "sample_bom.xlsx"
    workbook.save(file_path)

    result = parse_bom_xlsx(str(file_path), "sample_bom.xlsx")
    assert len(result["parts"]) == 1
    assert len(result["hardware"]) == 1
    assert result["parts"][0]["part_number"] == "P-100"
    assert result["parts"][0]["qty"] == 4
    assert result["parts"][0]["material"] == "A36 Steel"
    assert result["parts"][0]["flat_area"] == 96
    assert result["parts"][0]["cut_length"] == 40
    assert "sample_bom.xlsx!BOM:row2" in result["parts"][0]["source"]


def test_parse_dxf_geometry_extracts_area_perimeter_and_features(tmp_path: Path):
    doc = ezdxf.new()
    msp = doc.modelspace()
    msp.add_lwpolyline([(0, 0), (4, 0), (4, 2), (0, 2)], close=True)
    msp.add_circle((1, 1), 0.25)
    msp.add_line((0.5, 0.5), (3.5, 0.5), dxfattribs={"layer": "BEND"})

    file_path = tmp_path / "part_flat.dxf"
    doc.saveas(file_path)

    result = parse_dxf_geometry(str(file_path), "part_flat.dxf")
    assert result["flat_area"] is not None
    assert abs(result["flat_area"] - 8.0) < 0.25
    assert result["cut_length"] is not None
    assert result["cut_length"] > 12.0  # perimeter + hole cut
    assert result["hole_count"] == 1
    assert result["bend_count"] >= 1


def test_parse_pdf_drawing_extracts_drawing_details(monkeypatch):
    monkeypatch.setattr(
        rfq_parser,
        "extract_text_from_pdf",
        lambda *_args, **_kwargs: SimpleNamespace(
            text=(
                "PART NO P-900 DWG NO D-900 REV B MATERIAL 5052 ALUMINUM "
                "THICKNESS .125 in POWDER COAT 4X Ø.250 HOLES 3 BEND LINES "
                "+/- .005 WELD"
            )
        ),
    )

    result = rfq_parser.parse_pdf_drawing("/tmp/P-900.pdf", "P-900.pdf")

    assert result["part_hint"] == "P-900"
    assert result["drawing_number"] == "D-900"
    assert result["revision"] == "B"
    assert result["material"] == "Aluminum"
    assert result["thickness_in"] == 0.125
    assert result["finish"].lower() == "powder coat"
    assert result["hole_count"] == 4
    assert result["bend_count"] == 3
    assert result["weld_required"] is True
    assert result["tolerances_flag"] is True


def test_build_normalized_part_specs_cross_references_pdf_and_dxf():
    pdf_specs = [
        {
            "file_name": "P-500 Rev B.pdf",
            "source_type": "pdf",
            "part_hint": "P-500",
            "drawing_number": "D-500",
            "revision": "B",
            "material": "Aluminum",
            "thickness": "0.125",
            "thickness_in": 0.125,
            "flat_area": None,
            "cut_length": None,
            "hole_count": 6,
            "bend_count": 4,
            "finish": "Powder Coat",
            "weld_required": True,
            "assembly_required": False,
            "tolerances_flag": True,
            "confidence": {"material": 0.8, "thickness": 0.8, "finish": 0.75, "geometry": 0.0},
            "sources": {
                "material": ["P-500 Rev B.pdf:text"],
                "thickness": ["P-500 Rev B.pdf:text"],
                "finish": ["P-500 Rev B.pdf:text"],
                "drawing_detail": ["P-500 Rev B.pdf:text:drawing_number=D-500;revision=B;holes=6;bends=4"],
            },
        }
    ]
    dxf_specs = [
        {
            "file_name": "P-500_flat.dxf",
            "source_type": "dxf",
            "part_hint": "P-500_flat",
            "flat_area": 18.0,
            "cut_length": 36.0,
            "hole_count": 8,
            "bend_count": 4,
            "confidence": {"geometry": 0.9},
            "sources": {"geometry": ["P-500_flat.dxf:modelspace"]},
        }
    ]

    result = build_normalized_part_specs([], pdf_specs, dxf_specs, [])

    assert len(result["parts"]) == 1
    part = result["parts"][0]
    assert part["part_id"] == "P-500"
    assert part["material"] == "Aluminum"
    assert part["thickness_in"] == 0.125
    assert part["flat_area"] == 18.0
    assert part["cut_length"] == 36.0
    assert part["hole_count"] == 8
    assert part["bend_count"] == 4
    assert part["drawing_number"] == "D-500"
    assert part["revision"] == "B"
    assert "drawing_pdf" in part["sources"]
    assert "flat_pattern_dxf" in part["sources"]
    assert any(item["field"] == "cross_reference" for item in result["assumptions"])


def test_parse_pdf_drawing_extracts_assembly_bom(monkeypatch):
    text = """
PREP
RETAINER, CAPACITOR
TITLE
REVISION HISTORY
A CO-001 2026-01-01 JW
NOTES:
1. MATERIAL: PART 1: ITEM 1, .090 STOCK.
PARTS 2 AND 3: ITEM 1, .125 STOCK.
PART 4: COPPER, ELECTROLYTIC TOUGH PITCH TEMPER
SOFT-ANNEALED .0647 STOCK, IN ACCORDANCE WITH QQ-A-250/8.
2. FINISH: PARTS 1, 2 AND 3: ITEM 2.
PART 4: TIN PLATE, .00015 MIN.
9. CALCULATED WEIGHT: (0.73 LBS).
QTY ITEM
NO
PART OR IDENTIFYING NUMBER NOMENCLATURE OR DESCRIPTION DOCUMENT NO CAGEC NOTES REF DESIGNATOR
PARTS LIST
6 7 MS20426AD6-8 RIVET (CSK. .187 DIA X .500) 5
4 6 MS20426AD4-6 RIVET (CSK. .125 DIA X .375) 5
REF 2 580-0036-002 PROCESS, CHEMICAL FILM - YELLOW 2
AR 1 820-5052-010 ALUMINUM, 5052-H32 1
PART 1 PART 4
QUANTITY: 2
PART 2
PART 3
"""
    monkeypatch.setattr(rfq_parser, "extract_text_from_pdf", lambda *_args, **_kwargs: SimpleNamespace(text=text))

    pdf = rfq_parser.parse_pdf_drawing("/tmp/818-3928-638_RevA.pdf", "818-3928-638_RevA.pdf")
    result = build_normalized_part_specs([], [pdf], [], [])

    assert pdf["document_kind"] == "assembly"
    assert pdf["assembly"]["part_number"] == "818-3928-638"
    assert pdf["assembly"]["part_name"] == "RETAINER, CAPACITOR"
    assert len(pdf["assembly"]["bom_items"]) == 4
    parts = result["parts"]
    assert any(part["line_type"] == "assembly" and part["part_id"] == "818-3928-638" for part in parts)
    manufactured = [part for part in parts if part["line_type"] == "manufactured"]
    assert len(manufactured) == 4
    part4 = next(part for part in manufactured if part["part_id"].endswith("PART-4"))
    assert part4["material"] == "Copper"
    assert part4["thickness_in"] == 0.0647
    assert part4["quantity_per_assembly"] == 2.0
    assert len([part for part in parts if part["line_type"] == "hardware"]) == 2


def _save_simple_bom(path: Path, *data_rows: list) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["Part Number", "Description", "Qty"])
    for row in data_rows:
        sheet.append(row)
    workbook.save(path)


def test_parse_bom_xlsx_bloated_used_range_parses_fast(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """The production failure mode for this bug class: a single stray
    whitespace cell at XFD1048576 bloats the declared used range to
    16,384 x 1,048,576 — the old ``list(sheet.iter_rows())`` materialized the
    whole grid (minutes of CPU and potentially GBs of memory for a KB-sized
    upload)."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    sheet.append(["Part Number", "Description", "Qty"])
    sheet.append(["P-1", "Bracket", 2])
    sheet.append(["P-2", "Spacer", 4])
    sheet.cell(row=1_048_576, column=16_384, value=" ")
    file_path = tmp_path / "bloated.xlsx"
    workbook.save(file_path)

    # Deterministic guard (wall clock alone is too loose on coverage-traced CI
    # runners): with the per-sheet blank-run cutoff working, the scan stops
    # after ~1k consecutive blank rows; if the cutoff ever regresses to
    # unbounded blank scanning, this tightened backstop raises ImportFileError
    # regardless of runner speed.
    monkeypatch.setattr("app.services.rfq_parsing_service.MAX_SCANNED_ROWS", 60_000)

    started = time.monotonic()
    result = parse_bom_xlsx(str(file_path), "bloated.xlsx")
    elapsed = time.monotonic() - started

    assert [part["part_number"] for part in result["parts"]] == ["P-1", "P-2"]
    assert [part["qty"] for part in result["parts"]] == [2, 4]
    assert result["hardware"] == []
    # Wall clock is only a loose backstop — coverage-traced CI runners are
    # ~60x slower than local; a full-grid regression takes many minutes.
    assert elapsed < 90, f"bloated-dimension BOM parse took {elapsed:.1f}s — grid scan regression"


def test_parse_bom_xlsx_blank_run_cutoff_is_per_sheet(tmp_path: Path):
    """A gap longer than MAX_CONSECUTIVE_BLANK_ROWS ends only THAT sheet's
    scan (treated as used-range bloat); later sheets still parse with their
    own headers."""
    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["Part Number", "Description", "Qty"])
    first.append(["P-1", "Bracket", 2])
    first.cell(row=MAX_CONSECUTIVE_BLANK_ROWS + 100, column=1, value="P-DROPPED")
    second = workbook.create_sheet("Second")
    second.append(["Part Number", "Description", "Qty"])
    second.append(["P-2", "Spacer", 4])
    file_path = tmp_path / "gap.xlsx"
    workbook.save(file_path)

    result = parse_bom_xlsx(str(file_path), "gap.xlsx")

    part_numbers = [part["part_number"] for part in result["parts"]]
    assert "P-1" in part_numbers
    assert "P-2" in part_numbers
    assert "P-DROPPED" not in part_numbers
    p2 = next(part for part in result["parts"] if part["part_number"] == "P-2")
    assert p2["source"] == "gap.xlsx!Second:row2"


def test_parse_bom_xlsx_header_found_late_with_leading_junk_and_blank_gaps(tmp_path: Path):
    """A header anywhere in a sheet's first 30 rows is still honored, and
    blank rows — before the header and between the header and the data — are
    tolerated exactly as before bounding: collected blank rows stay in the row
    list, so the rows[:30] header window and spreadsheet row numbering are
    unchanged."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BOM"
    # Rows 1-24: scattered title junk; the rest of the rows are genuinely blank.
    for row_num in range(1, 25):
        if row_num % 3 == 0:
            sheet.cell(row=row_num, column=1, value=f"cover note {row_num}")
    sheet.cell(row=25, column=1, value="Part Number")
    sheet.cell(row=25, column=2, value="Description")
    sheet.cell(row=25, column=3, value="Qty")
    # Rows 26-27 stay blank between the header and the data row.
    sheet.cell(row=28, column=1, value="P-1")
    sheet.cell(row=28, column=2, value="Bracket")
    sheet.cell(row=28, column=3, value=2)
    file_path = tmp_path / "late_header.xlsx"
    workbook.save(file_path)

    result = parse_bom_xlsx(str(file_path), "late_header.xlsx")

    assert len(result["parts"]) == 1
    part = result["parts"][0]
    assert part["part_number"] == "P-1"
    assert part["qty"] == 2
    assert part["source"] == "late_header.xlsx!BOM:row28"


def test_parse_bom_xlsx_header_beyond_first_30_rows_is_still_skipped(tmp_path: Path):
    """Pins the existing header-window semantics: a header past a sheet's
    first 30 rows is not detected. Bounding the scan must neither widen nor
    narrow that window."""
    workbook = Workbook()
    sheet = workbook.active
    for row_num in range(1, 31):
        sheet.cell(row=row_num, column=1, value=f"cover note {row_num}")
    sheet.cell(row=31, column=1, value="Part Number")
    sheet.cell(row=31, column=2, value="Qty")
    sheet.cell(row=32, column=1, value="P-1")
    sheet.cell(row=32, column=2, value=2)
    file_path = tmp_path / "buried_header.xlsx"
    workbook.save(file_path)

    result = parse_bom_xlsx(str(file_path), "buried_header.xlsx")

    assert result == {"parts": [], "hardware": []}


def test_parse_bom_xlsx_scanned_row_cap_refuses_the_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """The workbook-wide scanned-row backstop refuses pathological files
    loudly with an actionable message instead of spinning."""
    monkeypatch.setattr("app.services.rfq_parsing_service.MAX_SCANNED_ROWS", 5)
    file_path = tmp_path / "huge.xlsx"
    _save_simple_bom(file_path, *[[f"P-{i}", "Bracket", 1] for i in range(10)])

    with pytest.raises(ImportFileError, match="used range is enormous"):
        parse_bom_xlsx(str(file_path), "huge.xlsx")


def test_parse_bom_xlsx_collected_row_cap_refuses_the_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """The collected-row cap (MAX_IMPORT_ROWS) is a backstop on non-blank rows
    per workbook — RFQ BOMs are far smaller."""
    monkeypatch.setattr("app.services.rfq_parsing_service.MAX_IMPORT_ROWS", 3)
    file_path = tmp_path / "many_rows.xlsx"
    _save_simple_bom(file_path, *[[f"P-{i}", "Bracket", 1] for i in range(4)])

    with pytest.raises(ImportFileError, match="Too many rows"):
        parse_bom_xlsx(str(file_path), "many_rows.xlsx")


def test_parse_rfq_package_files_records_import_file_error_per_file(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    """An ImportFileError raised by parse_bom_xlsx must land as THAT file's
    parse_error (parse_status "error") and must not abort the package parse;
    an openpyxl-unreadable .xls keeps failing into the same per-file path
    (pre-existing behavior, pinned here)."""
    monkeypatch.setattr("app.services.rfq_parsing_service.MAX_SCANNED_ROWS", 5)

    bloated_path = tmp_path / "huge.xlsx"
    _save_simple_bom(bloated_path, *[[f"P-{i}", "Bracket", 1] for i in range(10)])
    good_path = tmp_path / "good.xlsx"
    _save_simple_bom(good_path, ["P-2", "Spacer", 4])
    xls_path = tmp_path / "legacy.xls"
    xls_path.write_bytes(b"not really an xls workbook")

    files = [
        SimpleNamespace(id=1, file_ext=".xlsx", file_name="huge.xlsx", file_path=str(bloated_path)),
        SimpleNamespace(id=2, file_ext=".xlsx", file_name="good.xlsx", file_path=str(good_path)),
        SimpleNamespace(id=3, file_ext=".xls", file_name="legacy.xls", file_path=str(xls_path)),
    ]

    result = parse_rfq_package_files(files)

    assert result["file_results"][1]["parse_status"] == "error"
    assert "used range is enormous" in result["file_results"][1]["parse_error"]
    assert any("huge.xlsx: parse failed" in warning for warning in result["warnings"])
    # The good file still parses — the package parse was not aborted.
    assert result["file_results"][2]["parse_status"] == "parsed"
    assert result["file_results"][2]["summary"] == {"parts_found": 1, "hardware_found": 0}
    assert result["file_results"][3]["parse_status"] == "error"
    assert result["file_results"][3]["parse_error"]
