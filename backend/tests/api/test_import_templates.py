"""API tests for the server-generated XLSX import templates (A0.2)."""

import io

import pytest
from fastapi import status
from fastapi.testclient import TestClient
from openpyxl import load_workbook

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.mark.api
@pytest.mark.requires_db
class TestImportTemplates:
    def test_list_templates(self, client: TestClient, auth_headers: dict):
        response = client.get("/api/v1/import/templates", headers=auth_headers)
        assert response.status_code == status.HTTP_200_OK
        entities = {item["entity"] for item in response.json()["templates"]}
        assert {"users", "parts", "materials", "customers", "vendors", "work-centers"} <= entities
        assert {"work-orders", "purchase-orders"} <= entities

    def test_download_template_is_valid_xlsx(self, client: TestClient, auth_headers: dict):
        response = client.get("/api/v1/import/templates/work-orders", headers=auth_headers)
        assert response.status_code == status.HTTP_200_OK
        assert response.headers["content-type"].startswith(XLSX_MEDIA_TYPE)
        assert "werco-import-template-work-orders.xlsx" in response.headers["content-disposition"]
        workbook = load_workbook(io.BytesIO(response.content))
        assert workbook.sheetnames == ["Import", "Examples"]
        header = [cell.value for cell in workbook["Import"][1]]
        assert "part_number" in header and "completed_through_seq" in header

    def test_unknown_entity_404(self, client: TestClient, auth_headers: dict):
        response = client.get("/api/v1/import/templates/widgets", headers=auth_headers)
        assert response.status_code == status.HTTP_404_NOT_FOUND

    def test_requires_auth(self, client: TestClient):
        response = client.get("/api/v1/import/templates/parts")
        assert response.status_code == status.HTTP_401_UNAUTHORIZED
