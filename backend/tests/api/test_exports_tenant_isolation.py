"""Tenant-isolation coverage for the ``/api/v1/exports/*`` download endpoints.

This file exists because of a confirmed cross-tenant disclosure in
``export_purchase_order_lines`` (``app/api/endpoints/exports.py``). It was the
only one of the seven exporters with **no** tenant predicate at all:

    query = db.query(PurchaseOrderLine).options(...)
    if start_date or end_date or status:
        query = query.join(PurchaseOrder)      # <-- join was CONDITIONAL
        ...
    lines = query.all()

``GET /api/v1/exports/purchase-orders/lines/export`` with no query parameters
therefore emitted **every** company's PO lines -- po numbers, vendor names, part
numbers, quantities and unit prices -- to any authenticated caller. The endpoint
already declared ``company_id: int = Depends(get_current_company_id)``; it simply
never used it.

Two properties of the fix are load-bearing, and each has its own test below:

1. The ``.join(PurchaseOrder)`` is now UNCONDITIONAL. Had it stayed inside the
   optional-filter block, a no-parameter request would have had no join for a
   predicate on the parent to apply to, and a parent-only tenant filter would
   silently do nothing -- i.e. the leak would survive a fix that looks correct.
2. Both sides are filtered. ``PurchaseOrderLine.company_id`` (TenantMixin) is the
   authoritative scope for the row, and ``PurchaseOrder.company_id`` additionally
   stops a line from being read through a foreign parent -- the exporter renders
   ``line.purchase_order.po_number`` and the parent's vendor name, so the parent
   is disclosed even when the child row is the caller's own.

The assertions read the actual response bytes (CSV text and the XLSX workbook),
not a row count: the failure mode being guarded is foreign *values* appearing in
a file a manager opens, so counting rows would pass on a file that still leaks.

Two sibling exporters (purchase orders, parts) carry the same-shape assertion so
this file is the home for the defect class rather than a single-defect
regression test, and a parameterized sweep at the end asserts the property for
**all seven** exporters -- tenant scoping belongs to the category, not to the one
route that happened to be broken.

All seven now require ADMIN/MANAGER (``docs/RBAC_PERMISSIONS.md`` -> Bulk data
export), so ``make_user`` here defaults to ADMIN. The gate itself is covered in
``test_export_gate_and_audit.py``.

Fixture conventions follow ``test_receiving_compliance.py`` /
``test_inventory_hardening.py``: company A is the seeded default (id=1), company
B is a second tenant (id=2), rows are created directly on the shared
``db_session``, and requests are made with directly-minted tokens.
"""

import io
from datetime import date
from typing import List, Optional

import pytest
from fastapi import status as http_status
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.core.security import create_access_token
from app.models.company import Company
from app.models.part import Part
from app.models.purchasing import POStatus, PurchaseOrder, PurchaseOrderLine, Vendor
from app.models.user import User, UserRole
from app.models.work_order import WorkOrder, WorkOrderStatus

pytestmark = [pytest.mark.api, pytest.mark.requires_db]

COMPANY_A = 1
COMPANY_B = 2

TEST_PASSWORD_HASH = "$2b$12$abcdefghijklmnopqrstuv"  # tokens are minted directly; never used for login

PO_LINES_EXPORT = "/api/v1/exports/purchase-orders/lines/export"
PO_EXPORT = "/api/v1/exports/purchase-orders/export"
PARTS_EXPORT = "/api/v1/exports/parts/export"

# Module-level counter so every fixture row gets a globally unique natural key,
# even across companies and across tests sharing a worker DB under `-n auto`.
_seq = {"n": 0}


def _next() -> int:
    _seq["n"] += 1
    return _seq["n"]


def _ensure_company(db: Session, company_id: int) -> Company:
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        company = Company(id=company_id, name=f"Company {company_id}", slug=f"company-{company_id}", is_active=True)
        db.add(company)
        db.commit()
    return company


def make_user(db: Session, *, company_id: int, role: UserRole = UserRole.ADMIN) -> User:
    _ensure_company(db, company_id)
    n = _next()
    user = User(
        email=f"exp-iso-{n}@co{company_id}.test",
        employee_id=f"EXPI-{n:05d}",
        first_name="Export",
        last_name=f"C{company_id}",
        hashed_password=TEST_PASSWORD_HASH,
        role=role,
        is_active=True,
        is_superuser=False,
        company_id=company_id,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def headers_for(user: User) -> dict:
    token = create_access_token(subject=user.id, company_id=user.company_id)
    return {"Authorization": f"Bearer {token}", "X-Requested-With": "XMLHttpRequest"}


def make_part(db: Session, *, company_id: int) -> Part:
    _ensure_company(db, company_id)
    n = _next()
    part = Part(
        part_number=f"EXPI-PART-{n:05d}",
        name=f"Export fixture part {n}",
        description="exports tenant-isolation fixture part",
        part_type="purchased",
        unit_of_measure="each",
        is_active=True,
        company_id=company_id,
    )
    db.add(part)
    db.commit()
    db.refresh(part)
    return part


def make_vendor(db: Session, *, company_id: int) -> Vendor:
    _ensure_company(db, company_id)
    n = _next()
    vendor = Vendor(
        code=f"EXPI-V-{n:05d}",
        name=f"Vendor EXPI {n:05d}",
        is_active=True,
        is_approved=True,
        company_id=company_id,
    )
    db.add(vendor)
    db.commit()
    db.refresh(vendor)
    return vendor


def make_po(db: Session, *, company_id: int, po_status: POStatus = POStatus.SENT) -> PurchaseOrder:
    vendor = make_vendor(db, company_id=company_id)
    n = _next()
    po = PurchaseOrder(
        po_number=f"EXPI-PO-{n:05d}",
        vendor_id=vendor.id,
        status=po_status,
        order_date=date.today(),
        company_id=company_id,
        is_deleted=False,
    )
    db.add(po)
    db.commit()
    db.refresh(po)
    return po


def make_po_line(
    db: Session,
    *,
    po: PurchaseOrder,
    part: Part,
    unit_price: float,
    company_id: Optional[int] = None,
) -> PurchaseOrderLine:
    """Create a PO line.

    ``company_id`` defaults to the parent PO's company but can be set
    independently on purpose: a line whose own tenant stamp disagrees with its
    parent's is exactly the row that proves *which* predicate excludes it.
    """
    n = _next()
    line = PurchaseOrderLine(
        purchase_order_id=po.id,
        line_number=n,
        part_id=part.id,
        quantity_ordered=5.0,
        quantity_received=0.0,
        unit_price=unit_price,
        line_total=unit_price * 5.0,
        is_closed=False,
        company_id=company_id if company_id is not None else po.company_id,
    )
    db.add(line)
    db.commit()
    db.refresh(line)
    return line


def seed_company_purchasing(db: Session, *, company_id: int, unit_price: float):
    """Vendor + PO + Part + one PO line, all stamped ``company_id``."""
    po = make_po(db, company_id=company_id)
    part = make_part(db, company_id=company_id)
    line = make_po_line(db, po=po, part=part, unit_price=unit_price)
    return po, part, line


def xlsx_cell_text(payload: bytes) -> List[str]:
    """Every non-empty cell of the first worksheet, as strings."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(payload))
    sheet = workbook.active
    return [str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None]


# ===========================================================================
# The defect: GET /exports/purchase-orders/lines/export
# ===========================================================================


def test_po_lines_export_with_no_params_excludes_other_company_csv(client: TestClient, db_session: Session):
    """The exact pre-fix leak: no query params, so no join, so no tenant scope."""
    a_po, a_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_po, b_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_B, unit_price=999.99)
    b_vendor_name = b_po.vendor.name

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_LINES_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text

    # Company A's own line is present (control -- proves the export still works).
    assert a_po.po_number in body
    assert a_part.part_number in body
    assert "11.11" in body

    # Nothing of company B's reaches the file: not the PO number, not the part
    # number, not the vendor identity, not the price.
    assert b_po.po_number not in body
    assert b_part.part_number not in body
    assert b_vendor_name not in body
    assert "999.99" not in body


def test_po_lines_export_with_no_params_excludes_other_company_xlsx(client: TestClient, db_session: Session):
    """Same assertion against the XLSX bytes -- both formats share the query."""
    a_po, a_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_po, b_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_B, unit_price=999.99)
    b_vendor_name = b_po.vendor.name

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_LINES_EXPORT, headers=headers, params={"format": "xlsx"})

    assert response.status_code == http_status.HTTP_200_OK
    cells = xlsx_cell_text(response.content)

    assert a_po.po_number in cells
    assert a_part.part_number in cells

    assert b_po.po_number not in cells
    assert b_part.part_number not in cells
    assert b_vendor_name not in cells
    assert "999.99" not in cells


def test_po_lines_export_stays_scoped_when_filters_are_supplied(client: TestClient, db_session: Session):
    """The filtered path joined before the fix, but still never filtered on tenant."""
    a_po, a_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_po, b_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_B, unit_price=999.99)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(
        PO_LINES_EXPORT,
        headers=headers,
        # Deliberately matches BOTH companies' fixtures: same status, and a date
        # window around today's order_date.
        params={
            "format": "csv",
            "status": POStatus.SENT.value,
            "start_date": date.today().isoformat(),
            "end_date": date.today().isoformat(),
        },
    )

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text
    assert a_po.po_number in body
    assert a_part.part_number in body
    assert b_po.po_number not in body
    assert b_part.part_number not in body


def test_po_lines_export_excludes_foreign_line_hung_off_own_po(client: TestClient, db_session: Session):
    """A line stamped company B, attached to company A's PO, must not export.

    This is the row only the CHILD predicate excludes: the parent PO is company
    A's, so ``PurchaseOrder.company_id == company_id`` passes. ``company_id`` on
    ``purchase_order_lines`` is the TenantMixin column and is authoritative for
    the row.
    """
    a_po, a_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_part = make_part(db_session, company_id=COMPANY_B)
    make_po_line(db_session, po=a_po, part=b_part, unit_price=4242.42, company_id=COMPANY_B)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_LINES_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text
    assert a_part.part_number in body  # control
    assert b_part.part_number not in body
    assert "4242.42" not in body


def test_po_lines_export_excludes_own_line_hung_off_foreign_po(client: TestClient, db_session: Session):
    """A line stamped company A, attached to company B's PO, must not export.

    This is the row only the PARENT predicate excludes -- and it matters because
    the exporter renders the parent's ``po_number`` and the parent vendor's name
    onto every row, so exporting it would disclose company B's purchasing
    identity even though the child row carries company A's stamp.
    """
    _, a_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_po = make_po(db_session, company_id=COMPANY_B)
    b_vendor_name = b_po.vendor.name
    make_po_line(db_session, po=b_po, part=a_part, unit_price=5353.53, company_id=COMPANY_A)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_LINES_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text
    assert "11.11" in body  # control: company A's own line on its own PO
    assert b_po.po_number not in body
    assert b_vendor_name not in body
    assert "5353.53" not in body


def test_po_lines_export_is_empty_for_a_tenant_with_no_purchasing(client: TestClient, db_session: Session):
    """A tenant with no PO lines gets an empty file, never someone else's."""
    b_po, b_part, _ = seed_company_purchasing(db_session, company_id=COMPANY_B, unit_price=999.99)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_LINES_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    assert b_po.po_number not in response.text
    assert b_part.part_number not in response.text


# ===========================================================================
# Sibling exporters -- same shape, so this file owns the defect class
# ===========================================================================


def test_purchase_orders_export_excludes_other_company(client: TestClient, db_session: Session):
    a_po, _, _ = seed_company_purchasing(db_session, company_id=COMPANY_A, unit_price=11.11)
    b_po, _, _ = seed_company_purchasing(db_session, company_id=COMPANY_B, unit_price=999.99)
    b_vendor_name = b_po.vendor.name

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PO_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text
    assert a_po.po_number in body
    assert b_po.po_number not in body
    assert b_vendor_name not in body


def test_parts_export_excludes_other_company(client: TestClient, db_session: Session):
    a_part = make_part(db_session, company_id=COMPANY_A)
    b_part = make_part(db_session, company_id=COMPANY_B)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(PARTS_EXPORT, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, response.text
    body = response.text
    assert a_part.part_number in body
    assert b_part.part_number not in body


# ===========================================================================
# Every exporter, swept
#
# The three tests above grew out of one confirmed defect and cover three of the
# seven routes. Tenant scoping is a property of the CATEGORY, not of the route
# that happened to be broken, so the sweep below asserts it for all seven --
# including work orders, inventory, quotes and the inventory ledger, which had
# no cross-tenant coverage at all.
#
# Each company is seeded with one row of every exported kind, and every one of
# those rows carries the same per-company TAG in a field that appears in the
# exported file. The assertion is then a single, format-agnostic property: the
# other tenant's tag appears NOWHERE in the bytes, and the caller's own tag does
# (so an endpoint that leaks nothing because it returns nothing cannot pass).
# ===========================================================================


def seed_tagged_tenant(db: Session, *, company_id: int) -> str:
    """Seed one row of every exported dataset, all stamped with a unique tag."""
    from app.models.inventory import InventoryItem, InventoryTransaction, TransactionType
    from app.models.quote import Quote, QuoteStatus

    tag = f"TAG{_next():05d}X"
    user = make_user(db, company_id=company_id)

    part = Part(
        part_number=f"EXPI-{tag}-PART",
        name=f"Part {tag}",
        part_type="purchased",
        unit_of_measure="each",
        standard_cost=13.50,
        is_active=True,
        company_id=company_id,
    )
    db.add(part)
    db.flush()

    db.add(
        WorkOrder(
            work_order_number=f"EXPI-{tag}-WO",
            part_id=part.id,
            quantity_ordered=7.0,
            status=WorkOrderStatus.RELEASED,
            priority=3,
            customer_name=f"Customer {tag}",
            customer_po=f"EXPI-{tag}-CPO",
            due_date=date.today(),
            company_id=company_id,
            is_deleted=False,
        )
    )
    db.add(
        InventoryItem(
            part_id=part.id,
            location=f"EXPI-{tag}-LOC",
            warehouse="MAIN",
            quantity_on_hand=4.0,
            quantity_available=4.0,
            lot_number=f"EXPI-{tag}-LOT",
            unit_cost=13.50,
            is_active=True,
            company_id=company_id,
        )
    )
    db.add(
        InventoryTransaction(
            part_id=part.id,
            transaction_type=TransactionType.RECEIVE,
            quantity=4.0,
            reference_type="purchase_order",
            reference_number=f"EXPI-{tag}-TXN",
            lot_number=f"EXPI-{tag}-LOT",
            unit_cost=13.50,
            total_cost=54.0,
            created_by=user.id,
            company_id=company_id,
        )
    )
    db.add(
        Quote(
            quote_number=f"EXPI-{tag}-QUOTE",
            customer_name=f"Customer {tag}",
            customer_contact=f"Contact {tag}",
            customer_email=f"{tag}@expi.test",
            status=QuoteStatus.SENT,
            quote_date=date.today(),
            total=1234.56,
            company_id=company_id,
        )
    )

    vendor = Vendor(
        code=f"EXPI-{tag}-V",
        name=f"Vendor {tag}",
        is_active=True,
        is_approved=True,
        company_id=company_id,
    )
    db.add(vendor)
    db.flush()
    po = PurchaseOrder(
        po_number=f"EXPI-{tag}-PO",
        vendor_id=vendor.id,
        status=POStatus.SENT,
        order_date=date.today(),
        total=54.0,
        company_id=company_id,
        is_deleted=False,
    )
    db.add(po)
    db.flush()
    db.add(
        PurchaseOrderLine(
            purchase_order_id=po.id,
            line_number=1,
            part_id=part.id,
            quantity_ordered=4.0,
            quantity_received=0.0,
            unit_price=13.50,
            line_total=54.0,
            is_closed=False,
            company_id=company_id,
        )
    )
    db.commit()
    return tag


ALL_EXPORT_ROUTES = [
    "/api/v1/exports/work-orders/export",
    "/api/v1/exports/parts/export",
    "/api/v1/exports/inventory/export",
    "/api/v1/exports/purchase-orders/export",
    "/api/v1/exports/purchase-orders/lines/export",
    "/api/v1/exports/quotes/export",
    "/api/v1/exports/inventory/transactions/export",
]


@pytest.mark.parametrize("route", ALL_EXPORT_ROUTES)
def test_no_exporter_emits_another_tenants_rows(client: TestClient, db_session: Session, route: str):
    a_tag = seed_tagged_tenant(db_session, company_id=COMPANY_A)
    b_tag = seed_tagged_tenant(db_session, company_id=COMPANY_B)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(route, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, f"{route}: {response.text}"
    body = response.text
    # Control: the caller's own dataset IS in the file, so "leaks nothing"
    # cannot be satisfied by "returns nothing".
    assert a_tag in body, f"{route} returned none of the caller's own rows"
    assert b_tag not in body, f"{route} leaked company B rows"


@pytest.mark.parametrize("route", ALL_EXPORT_ROUTES)
def test_no_exporter_emits_rows_to_a_tenant_that_has_none(client: TestClient, db_session: Session, route: str):
    """The empty-tenant case: an unscoped query is most obvious when A owns nothing."""
    b_tag = seed_tagged_tenant(db_session, company_id=COMPANY_B)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(route, headers=headers, params={"format": "csv"})

    assert response.status_code == http_status.HTTP_200_OK, f"{route}: {response.text}"
    assert b_tag not in response.text, f"{route} leaked company B rows to an empty tenant"


@pytest.mark.parametrize("route", ALL_EXPORT_ROUTES)
def test_no_exporter_emits_another_tenants_rows_in_xlsx(client: TestClient, db_session: Session, route: str):
    """Same property against the workbook bytes -- both formats share one query."""
    a_tag = seed_tagged_tenant(db_session, company_id=COMPANY_A)
    b_tag = seed_tagged_tenant(db_session, company_id=COMPANY_B)

    headers = headers_for(make_user(db_session, company_id=COMPANY_A))
    response = client.get(route, headers=headers, params={"format": "xlsx"})

    assert response.status_code == http_status.HTTP_200_OK, route
    cells = " ".join(xlsx_cell_text(response.content))
    assert a_tag in cells, f"{route} returned none of the caller's own rows"
    assert b_tag not in cells, f"{route} leaked company B rows"
