import io
import time
from types import SimpleNamespace

import pytest
from fastapi import status
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.api.endpoints.bom import _extract_excel_table
from app.models.audit_log import AuditLog
from app.models.bom import BOM, BOMItem
from app.models.part import Part
from app.models.user import User
from app.models.work_order import WorkOrder, WorkOrderStatus
from app.models.work_order_material import AllocationStatus, WorkOrderMaterialAllocation
from app.services.import_service import MAX_CONSECUTIVE_BLANK_ROWS, MAX_IMPORT_COLUMNS, XLSX_MEDIA_TYPE, ImportFileError
from app.services.pdf_service import extract_text_from_excel


def _make_part(db_session: Session, part_number: str, part_type: str = "manufactured", **kwargs) -> Part:
    part = Part(
        part_number=part_number,
        name=part_number,
        part_type=part_type,
        unit_of_measure="each",
        company_id=1,
        **kwargs,
    )
    db_session.add(part)
    db_session.commit()
    db_session.refresh(part)
    return part


def _commit_payload(assembly_number: str, component_numbers=("COMP-A", "COMP-B")) -> dict:
    return {
        "document_type": "bom",
        "assembly": {"part_number": assembly_number, "name": "Imported Assembly", "revision": "A"},
        "items": [
            {
                "line_number": (idx + 1) * 10,
                "part_number": component_number,
                "description": f"Component {component_number}",
                "quantity": 1,
                "item_type": "make",
                "line_type": "component",
            }
            for idx, component_number in enumerate(component_numbers)
        ],
        "create_missing_parts": True,
    }


def _workbook_bytes(*sheets) -> bytes:
    """Build an in-memory xlsx; each positional arg is one sheet's list of rows."""
    workbook = Workbook()
    for sheet_index, rows in enumerate(sheets):
        sheet = workbook.active if sheet_index == 0 else workbook.create_sheet(f"Sheet{sheet_index + 1}")
        for row in rows:
            sheet.append(row)
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


def _bloated_workbook_bytes() -> bytes:
    """Header + two data rows, plus the production failure mode: a single stray
    whitespace cell in the very last cell of the grid (XFD1048576) bloating the
    declared used range to 16,384 x 1,048,576."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Part Number", "Description", "Qty"])
    sheet.append(["P-1", "Bracket", 2])
    sheet.append(["P-2", "Spacer", 4])
    sheet.cell(row=1_048_576, column=16_384, value=" ")
    out = io.BytesIO()
    workbook.save(out)
    return out.getvalue()


@pytest.mark.api
@pytest.mark.requires_db
class TestBOMImport:
    def test_commit_bom_import_creates_assembly_and_items(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
    ):
        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json={
                "document_type": "bom",
                "assembly": {
                    "part_number": "ASSY-100",
                    "name": "Imported Assembly",
                    "revision": "A",
                    "part_type": "manufactured",
                },
                "items": [
                    {
                        "line_number": 10,
                        "part_number": "COMP-100",
                        "description": "Machined bracket",
                        "quantity": 2,
                        "item_type": "make",
                        "line_type": "component",
                    },
                    {
                        "line_number": 20,
                        "part_number": "BUY-100",
                        "description": "Purchased spacer",
                        "quantity": 4,
                        "item_type": "buy",
                        "line_type": "component",
                    },
                ],
                "create_missing_parts": True,
            },
        )

        assert response.status_code == status.HTTP_201_CREATED
        data = response.json()
        assert data["bom_id"] is not None
        assert data["created_bom_items"] == 2

        assembly = db_session.query(Part).filter(Part.part_number == "ASSY-100").one()
        assert assembly.part_type == "assembly"

        make_component = db_session.query(Part).filter(Part.part_number == "COMP-100").one()
        buy_component = db_session.query(Part).filter(Part.part_number == "BUY-100").one()
        assert make_component.part_type == "manufactured"
        assert buy_component.part_type == "purchased"

        bom = db_session.query(BOM).filter(BOM.id == data["bom_id"]).one()
        items = db_session.query(BOMItem).filter(BOMItem.bom_id == bom.id).all()
        assert bom.part_id == assembly.id
        assert {item.component_part_id for item in items} == {make_component.id, buy_component.id}

        bom_response = client.get(f"/api/v1/bom/by-part/{assembly.id}", headers=auth_headers)
        assert bom_response.status_code == status.HTTP_200_OK

        bom_data = bom_response.json()
        assert bom_data["id"] == bom.id
        assert bom_data["part_id"] == assembly.id
        assert {item["component_part_id"] for item in bom_data["items"]} == {
            make_component.id,
            buy_component.id,
        }


@pytest.mark.api
@pytest.mark.requires_db
class TestBOMImportPreviewBoundedScan:
    """The Excel preview path must never iterate a workbook's full *declared*
    grid: one stray whitespace cell at XFD1048576 used to make a 5 KB upload
    scan ~17B cells (~5 minutes of CPU on the event loop in prod)."""

    def test_bloated_used_range_previews_fast_and_succeeds(
        self,
        client: TestClient,
        auth_headers: dict,
        monkeypatch: pytest.MonkeyPatch,
    ):
        # _extract_excel_table has no kwarg injection, so tighten the
        # scanned-row backstop it reads from module globals to make the
        # cutoff check deterministic: with the per-sheet blank-run cutoff
        # working, the scan stops after ~1k blank rows; if the cutoff ever
        # regresses to unbounded blank scanning, the backstop raises
        # ImportFileError (HTTP 400 here) regardless of runner speed.
        monkeypatch.setattr("app.api.endpoints.bom.MAX_SCANNED_ROWS", 60_000)

        started = time.monotonic()
        response = client.post(
            "/api/v1/bom/import/preview",
            headers=auth_headers,
            files={"file": ("bloated.xlsx", io.BytesIO(_bloated_workbook_bytes()), XLSX_MEDIA_TYPE)},
        )
        elapsed = time.monotonic() - started

        assert response.status_code == status.HTTP_200_OK, response.text
        data = response.json()
        assert data["source_format"] == "excel"
        assert [item["part_number"] for item in data["items"]] == ["P-1", "P-2"]
        assert [item["quantity"] for item in data["items"]] == [2.0, 4.0]
        # Wall clock is only a loose backstop — coverage-traced CI runners are
        # ~60x slower than local; a full-grid regression takes many minutes.
        assert elapsed < 90, f"bloated-dimension preview took {elapsed:.1f}s — grid scan regression"


@pytest.mark.unit
class TestExtractExcelTable:
    def test_multi_sheet_collection(self, tmp_path):
        """Header comes from the first non-empty row anywhere; later non-empty
        rows across ALL sheets are data rows (original semantics preserved)."""
        path = tmp_path / "multi.xlsx"
        path.write_bytes(
            _workbook_bytes(
                [["Part Number", "Description", "Qty"], ["P-1", "Bracket", "2"]],
                [["P-2", "Spacer", "4"], ["P-3", "Shim", "1"]],
            )
        )

        columns, rows = _extract_excel_table(str(path), ".xlsx")

        assert columns == ["Part Number", "Description", "Qty"]
        assert rows == [["P-1", "Bracket", "2"], ["P-2", "Spacer", "4"], ["P-3", "Shim", "1"]]

    def test_wider_than_cap_ignores_extra_columns(self, tmp_path):
        header = ["Part Number"] + [f"extra_{i}" for i in range(MAX_IMPORT_COLUMNS + 10)]
        data = ["P-1"] + ["x"] * (MAX_IMPORT_COLUMNS + 10)
        path = tmp_path / "wide.xlsx"
        path.write_bytes(_workbook_bytes([header, data]))

        columns, rows = _extract_excel_table(str(path), ".xlsx")

        assert len(columns) == MAX_IMPORT_COLUMNS
        assert columns[0] == "Part Number"
        assert len(rows) == 1
        assert len(rows[0]) == MAX_IMPORT_COLUMNS
        assert rows[0][0] == "P-1"

    def test_blank_run_cutoff_is_per_sheet(self, tmp_path):
        """A gap longer than MAX_CONSECUTIVE_BLANK_ROWS ends only THAT sheet's
        scan (treated as used-range bloat); later sheets still contribute rows.
        There is deliberately no loud-refusal look-ahead here — the preview
        shows users exactly which rows parsed before anything is committed."""
        workbook = Workbook()
        first = workbook.active
        first.append(["Part Number", "Qty"])
        first.append(["P-1", "1"])
        first.cell(row=MAX_CONSECUTIVE_BLANK_ROWS + 100, column=1, value="P-DROPPED")
        second = workbook.create_sheet("Second")
        second.append(["P-2", "3"])
        out = io.BytesIO()
        workbook.save(out)
        path = tmp_path / "gap.xlsx"
        path.write_bytes(out.getvalue())

        columns, rows = _extract_excel_table(str(path), ".xlsx")

        assert columns == ["Part Number", "Qty"]
        assert ["P-1", "1"] in rows
        assert ["P-2", "3"] in rows
        assert not any("P-DROPPED" in row for row in rows)

    def test_corrupt_file_raises_import_file_error(self, tmp_path):
        path = tmp_path / "corrupt.xlsx"
        path.write_bytes(b"this is not a spreadsheet")

        with pytest.raises(ImportFileError, match="Could not read the Excel file"):
            _extract_excel_table(str(path), ".xlsx")

    def test_header_padded_to_widest_data_row(self, tmp_path):
        """A data column with no header cell must still surface in the mapping
        UI (rendered as "Col N"), so the header is padded to the widest row."""
        path = tmp_path / "unheadered.xlsx"
        path.write_bytes(
            _workbook_bytes(
                [["Part Number", "Qty"], ["P-1", "2", "vendor-note"]],
            )
        )

        columns, rows = _extract_excel_table(str(path), ".xlsx")

        assert columns == ["Part Number", "Qty", ""]
        assert rows == [["P-1", "2", "vendor-note"]]


@pytest.mark.unit
class TestExtractTextFromExcelBounded:
    def test_bloated_used_range_extracts_quickly(self, tmp_path):
        path = tmp_path / "bloated.xlsx"
        path.write_bytes(_bloated_workbook_bytes())

        started = time.monotonic()
        result = extract_text_from_excel(str(path))
        elapsed = time.monotonic() - started

        assert "P-1" in result.text
        assert "P-2" in result.text
        assert result.confidence == "medium"
        # Loose backstop only (see TestBOMImportPreviewBoundedScan); a
        # full-grid regression takes many minutes, not seconds.
        assert elapsed < 90, f"bloated-dimension text extraction took {elapsed:.1f}s — grid scan regression"

    def test_scan_cap_returns_partial_text_instead_of_failing(self, tmp_path, monkeypatch: pytest.MonkeyPatch):
        """Hitting the workbook-wide scanned-row cap must degrade gracefully:
        stop scanning and return what was extracted so far at medium
        confidence — never raise (text extraction feeds best-effort flows)."""
        monkeypatch.setattr("app.services.pdf_service.MAX_SCANNED_ROWS", 3)
        path = tmp_path / "rows.xlsx"
        path.write_bytes(_workbook_bytes([["h1"], ["r1"], ["r2"], ["r3"], ["r4"]]))

        result = extract_text_from_excel(str(path))

        assert result.confidence == "medium"
        assert "r2" in result.text  # rows inside the cap survive
        assert "r3" not in result.text  # rows past the cap are dropped, not an error
        assert "r4" not in result.text


@pytest.mark.api
@pytest.mark.requires_db
class TestBOMImportAudit:
    """BOM imports create tenant data (parts, a BOM, BOM items) and must leave
    an audit trail: one CREATE row per created part, one CREATE row for the
    BOM with the items summarized in extra_data (the WO-import house pattern),
    and an UPDATE row for the in-place part_type promotion."""

    def test_commit_writes_part_and_bom_audit_rows(self, client: TestClient, auth_headers: dict, db_session: Session):
        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-AUD"),
        )
        assert response.status_code == status.HTTP_201_CREATED, response.text
        data = response.json()

        part_creates = (
            db_session.query(AuditLog).filter(AuditLog.resource_type == "part", AuditLog.action == "CREATE").all()
        )
        assert {row.resource_identifier for row in part_creates} == {"ASSY-AUD", "COMP-A", "COMP-B"}
        assert len(part_creates) == 3  # exactly one CREATE per created part
        assert all(row.extra_data.get("source") == "bom_import" for row in part_creates)

        bom_creates = (
            db_session.query(AuditLog).filter(AuditLog.resource_type == "bom", AuditLog.action == "CREATE").all()
        )
        assert len(bom_creates) == 1
        bom_row = bom_creates[0]
        assert bom_row.resource_id == data["bom_id"]
        # The SAME handle the six BOM header verbs write for resource_type="bom", not the
        # bare part number this importer used to write. One resource_type, one shape.
        assert bom_row.resource_identifier == "ASSY-AUD BOM rev A"
        assert bom_row.extra_data["source"] == "bom_import"
        assert bom_row.extra_data["item_count"] == 2
        assert bom_row.extra_data["component_part_numbers"] == ["COMP-A", "COMP-B"]

        # Every audit row carries the requesting user's company scope.
        assert {row.company_id for row in part_creates + bom_creates} == {1}

    def test_import_born_bom_shares_one_identifier_shape_with_the_header_verbs(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        """One ``resource_type="bom"`` chain, one handle on it.

        The importers wrote the bare assembly part number (``"WRC-1001"``) while all six
        header verbs write ``bom_identifier`` (``"WRC-1001 BOM rev A"``), so an auditor
        pulling the chain for one document saw its CREATE row named differently from every
        row that followed it -- and the revision, which the importer had just written to the
        row, appeared nowhere on the CREATE. This drives the *same* BOM through an import
        and then a header verb and asserts the two rows agree, which is the property that
        actually matters and which neither importer test alone can express.
        """
        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-SHAPE"),
        )
        assert response.status_code == status.HTTP_201_CREATED, response.text
        bom_id = response.json()["bom_id"]

        # A header verb on the same document. Release is the interesting one: it is the
        # approval row an auditor correlates with the CREATE.
        released = client.post(f"/api/v1/bom/{bom_id}/release", headers=auth_headers)
        assert released.status_code == status.HTTP_200_OK, released.text

        rows = db_session.query(AuditLog).filter(AuditLog.resource_type == "bom", AuditLog.resource_id == bom_id).all()
        assert {row.action for row in rows} == {"CREATE", "STATUS_CHANGE"}
        assert {row.resource_identifier for row in rows} == {"ASSY-SHAPE BOM rev A"}

    def test_commit_part_type_promotion_writes_update_audit(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        existing = _make_part(db_session, "ASSY-PROMO", part_type="manufactured")

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-PROMO", component_numbers=("COMP-P",)),
        )
        assert response.status_code == status.HTTP_201_CREATED, response.text

        db_session.refresh(existing)
        assert existing.part_type == "assembly"

        updates = (
            db_session.query(AuditLog)
            .filter(
                AuditLog.resource_type == "part",
                AuditLog.resource_id == existing.id,
                AuditLog.action == "UPDATE",
            )
            .all()
        )
        assert len(updates) == 1
        assert updates[0].old_values["part_type"] == "manufactured"
        assert updates[0].new_values["part_type"] == "assembly"
        assert updates[0].extra_data["source"] == "bom_import"
        # The pre-existing assembly part must NOT get a CREATE row.
        assert (
            db_session.query(AuditLog)
            .filter(
                AuditLog.resource_type == "part",
                AuditLog.resource_id == existing.id,
                AuditLog.action == "CREATE",
            )
            .count()
            == 0
        )

    def test_llm_import_writes_part_and_bom_audit_rows(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        monkeypatch: pytest.MonkeyPatch,
    ):
        """POST /bom/import (one-shot LLM path) must audit like the commit path.
        The LLM/extraction seams are stubbed at the endpoint module."""
        monkeypatch.setattr("app.api.endpoints.bom.save_uploaded_document", lambda content, filename: "/tmp/fake.pdf")
        monkeypatch.setattr(
            "app.api.endpoints.bom.extract_text_from_document",
            lambda path: SimpleNamespace(text="BOM document text " * 10, is_ocr=False),
        )
        monkeypatch.setattr(
            "app.api.endpoints.bom.extract_bom_data_with_llm",
            lambda text, is_ocr=False, company_id=None: {
                "document_type": "bom",
                "assembly": {"part_number": "ASSY-LLM", "name": "LLM Assembly", "revision": "A"},
                "items": [
                    {
                        "line_number": 10,
                        "part_number": "COMP-LLM",
                        "description": "Machined bracket",
                        "quantity": 2,
                        "item_type": "make",
                        "line_type": "component",
                    }
                ],
                "extraction_confidence": "high",
            },
        )

        response = client.post(
            "/api/v1/bom/import",
            headers=auth_headers,
            files={"file": ("bom.pdf", io.BytesIO(b"%PDF-1.4 fake"), "application/pdf")},
        )
        assert response.status_code == status.HTTP_201_CREATED, response.text
        data = response.json()

        part_creates = (
            db_session.query(AuditLog).filter(AuditLog.resource_type == "part", AuditLog.action == "CREATE").all()
        )
        assert {row.resource_identifier for row in part_creates} == {"ASSY-LLM", "COMP-LLM"}
        assert all(row.extra_data.get("source") == "bom_import" for row in part_creates)

        bom_creates = (
            db_session.query(AuditLog).filter(AuditLog.resource_type == "bom", AuditLog.action == "CREATE").all()
        )
        assert len(bom_creates) == 1
        assert bom_creates[0].resource_id == data["bom_id"]
        # Same handle as every other resource_type="bom" row (see the commit-importer test).
        assert bom_creates[0].resource_identifier == "ASSY-LLM BOM rev A"
        assert bom_creates[0].extra_data["item_count"] == 1
        assert bom_creates[0].extra_data["component_part_numbers"] == ["COMP-LLM"]
        assert {row.company_id for row in part_creates + bom_creates} == {1}


@pytest.mark.api
@pytest.mark.requires_db
class TestBOMImportSoftDeleteConflicts:
    """Soft-deleted rows still occupy their unique keys (part_number, BOM.part_id
    has no soft-delete carve-out), so import collisions must fail loudly with an
    actionable 400 — never silently resurrect deleted data or 500 on the unique
    constraint. The whole import is one transaction: nothing may persist."""

    def test_soft_deleted_assembly_part_rejected(
        self, client: TestClient, auth_headers: dict, db_session: Session, test_user: User
    ):
        deleted = _make_part(db_session, "ASSY-DEL")
        deleted.soft_delete(test_user.id)
        db_session.commit()
        audit_before = db_session.query(AuditLog).count()

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-DEL"),
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        detail = response.json()["detail"]
        assert "ASSY-DEL" in detail
        assert "deleted part" in detail
        db_session.refresh(deleted)
        assert deleted.is_deleted is True  # not resurrected
        assert db_session.query(BOM).count() == 0
        assert db_session.query(BOMItem).count() == 0
        assert db_session.query(AuditLog).count() == audit_before

    def test_soft_deleted_component_part_rejected(
        self, client: TestClient, auth_headers: dict, db_session: Session, test_user: User
    ):
        deleted = _make_part(db_session, "COMP-DEL")
        deleted.soft_delete(test_user.id)
        db_session.commit()
        audit_before = db_session.query(AuditLog).count()

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-ROLLBACK", component_numbers=("COMP-DEL",)),
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        detail = response.json()["detail"]
        assert "COMP-DEL" in detail
        assert "deleted part" in detail
        db_session.refresh(deleted)
        assert deleted.is_deleted is True
        # Single transaction: the assembly part flushed before the conflict is rolled back too.
        assert db_session.query(Part).filter(Part.part_number == "ASSY-ROLLBACK").first() is None
        assert db_session.query(BOM).count() == 0
        assert db_session.query(BOMItem).count() == 0
        assert db_session.query(AuditLog).count() == audit_before

    def test_soft_deleted_bom_rejected(
        self, client: TestClient, auth_headers: dict, db_session: Session, test_user: User
    ):
        assembly = _make_part(db_session, "ASSY-BOMDEL", part_type="assembly")
        bom = BOM(part_id=assembly.id, revision="A", status="draft", bom_type="standard", company_id=1)
        bom.soft_delete(test_user.id)
        db_session.add(bom)
        db_session.commit()
        audit_before = db_session.query(AuditLog).count()

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-BOMDEL"),
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        detail = response.json()["detail"]
        assert "deleted BOM exists" in detail
        assert "ASSY-BOMDEL" in detail
        assert db_session.query(BOM).count() == 1  # only the pre-existing soft-deleted row
        assert db_session.query(BOMItem).count() == 0
        assert db_session.query(AuditLog).count() == audit_before

    def test_inactive_bom_rejected_with_400_not_integrity_error(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        """An inactive BOM was invisible to the old is_active==True lookup; the
        import then violated the unique part_id constraint and 500ed."""
        assembly = _make_part(db_session, "ASSY-INACTIVE", part_type="assembly")
        bom = BOM(part_id=assembly.id, revision="A", status="draft", bom_type="standard", is_active=False, company_id=1)
        db_session.add(bom)
        db_session.commit()
        audit_before = db_session.query(AuditLog).count()

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-INACTIVE"),
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        detail = response.json()["detail"]
        assert "inactive BOM exists" in detail
        assert "ASSY-INACTIVE" in detail
        assert db_session.query(BOM).count() == 1
        assert db_session.query(BOMItem).count() == 0
        assert db_session.query(AuditLog).count() == audit_before

    def test_promotion_rolled_back_when_bom_conflict_follows(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        """The part_type promotion (and its UPDATE audit row) is staged BEFORE the
        existing-BOM check; a conflict must discard both — the rollback inside
        _reject_existing_bom, not session teardown, is what guarantees it."""
        assembly = _make_part(db_session, "ASSY-PROMO", part_type="manufactured")
        bom = BOM(part_id=assembly.id, revision="A", status="draft", bom_type="standard", company_id=1)
        db_session.add(bom)
        db_session.commit()
        audit_before = db_session.query(AuditLog).count()

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-PROMO"),
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        detail = response.json()["detail"]
        assert "already exists" in detail
        assert "ASSY-PROMO" in detail
        db_session.refresh(assembly)
        assert assembly.part_type == "manufactured"  # promotion rolled back
        assert db_session.query(AuditLog).count() == audit_before  # incl. no UPDATE row


@pytest.mark.api
@pytest.mark.requires_db
class TestBOMImportAssemblyPromotionRespectsMaterialTies:
    """The importer's assembly promotion is the SECOND conversion door, and it is gated.

    Both importers resolve an EXISTING part by part number (company scope only) and then
    set ``part_type = ASSEMBLY`` whenever the document is a BOM. Before this, importing a
    BOM whose parent number happened to name a ``raw_material`` or ``purchased`` part that
    work orders were still tying as material silently reclassified it — landing exactly the
    state ``material_tie_part_gate`` refuses at tie time, since a tie DEPLETES its part at
    completion and consumption never auto-reverses (invariant 6b).

    **"Still tying" means an OPEN tie on an UNFINISHED work order**, and both halves are
    tested here. Ties are never closed at completion, so a status-only count would warn on
    every BOM whose parent the shop had ever consumed and skip a promotion nothing was
    standing in the way of.

    **It SKIPS and WARNS rather than failing the import**, and that is the deliberate half.
    ``PUT /parts/{id}`` answers the same question with a 409 because it is one record and
    one decision; here a 409 would roll back an entire multi-record import — every
    component part, the BOM header, every line — over a step that is catalog hygiene, not
    a precondition (a BOM attaches by ``part_id`` and is valid on a parent still typed
    ``purchased``). What must never happen is a SILENT promotion, and it cannot: the part
    is either promoted, with an audit row, or reported, through ``warnings``.
    """

    def _tie(
        self,
        db_session: Session,
        part: Part,
        *,
        wo_status: WorkOrderStatus = WorkOrderStatus.IN_PROGRESS,
        qty_consumed: float = 0.0,
        suffix: str = "",
    ) -> None:
        work_order = WorkOrder(
            work_order_number=f"BOMTIE-WO-{part.id}{suffix}",
            customer_name="Acme",
            part_id=part.id,
            quantity_ordered=5,
            status=wo_status,
            priority=5,
            company_id=1,
        )
        db_session.add(work_order)
        db_session.commit()
        db_session.add(
            WorkOrderMaterialAllocation(
                company_id=1,
                work_order_id=work_order.id,
                part_id=part.id,
                source="manual",
                status=AllocationStatus.OPEN,
                qty_planned=3.0,
                unit_of_measure="each",
                qty_consumed=qty_consumed,
            )
        )
        db_session.commit()

    def test_a_tied_material_parent_is_not_promoted_and_the_import_still_succeeds(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        sheet = _make_part(db_session, "ASSY-TIED", part_type="raw_material")
        self._tie(db_session, sheet)

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-TIED", component_numbers=("COMP-T",)),
        )

        assert response.status_code == status.HTTP_201_CREATED, response.text
        body = response.json()
        assert body["bom_id"] is not None, "the BOM still lands — the promotion is not a precondition"

        db_session.refresh(sheet)
        assert sheet.part_type == "raw_material", "the tied part's class must be left alone"

        assert any(
            "ASSY-TIED" in warning and "1 unfinished work order still ties" in warning for warning in body["warnings"]
        ), body["warnings"]

        promotions = (
            db_session.query(AuditLog)
            .filter(
                AuditLog.resource_type == "part",
                AuditLog.resource_id == sheet.id,
                AuditLog.action == "UPDATE",
            )
            .all()
        )
        assert promotions == [], "a refused promotion changed nothing, so it writes no chain row (invariant 2)"

    def test_an_UNTIED_material_parent_is_still_promoted(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        """The negative control: the gate protects TIES, not classes.

        Without this, an importer that had simply stopped promoting anything would satisfy
        the assertions above.
        """
        sheet = _make_part(db_session, "ASSY-UNTIED", part_type="raw_material")

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-UNTIED", component_numbers=("COMP-U",)),
        )

        assert response.status_code == status.HTTP_201_CREATED, response.text
        db_session.refresh(sheet)
        assert sheet.part_type == "assembly"
        assert not any("ASSY-UNTIED" in warning for warning in response.json()["warnings"])

    @pytest.mark.parametrize(
        "finished_status",
        [WorkOrderStatus.COMPLETE, WorkOrderStatus.CLOSED, WorkOrderStatus.CANCELLED],
    )
    def test_a_parent_tied_only_by_FINISHED_work_orders_is_still_promoted(
        self,
        client: TestClient,
        auth_headers: dict,
        db_session: Session,
        finished_status: WorkOrderStatus,
    ):
        """A shipped job's tie is history, not demand — it must not warn and must not skip.

        The gate counts ties on UNFINISHED work orders only, and the importer is where
        getting that wrong would be loudest. Ties are never closed at completion (nothing
        in ``app/`` writes ``AllocationStatus.CLOSED``, and completion neither closes nor
        cancels them), so a status-only count would append this warning on every BOM whose
        parent the shop had ever consumed — and skip a promotion the planner asked for,
        naming a remedy nobody can perform: the untie verbs either 409 on issued material
        or credit a shipped job's material back into stock. A warnings channel that cries
        wolf on routine imports is a channel planners stop reading, which is how the real
        refusals get missed.

        The tie is left OPEN with material consumed against it — the exact state a finished
        job leaves behind.
        """
        part_number = f"ASSY-SHIPPED-{finished_status.value.upper()}"
        sheet = _make_part(db_session, part_number, part_type="raw_material")
        self._tie(db_session, sheet, wo_status=finished_status, qty_consumed=2.0)

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload(part_number, component_numbers=(f"COMP-S-{finished_status.value}",)),
        )

        assert response.status_code == status.HTTP_201_CREATED, response.text
        db_session.refresh(sheet)
        assert sheet.part_type == "assembly", "no live demand stands in the way, so the promotion runs"
        assert not any(part_number in warning for warning in response.json()["warnings"])

        promotions = (
            db_session.query(AuditLog)
            .filter(
                AuditLog.resource_type == "part",
                AuditLog.resource_id == sheet.id,
                AuditLog.action == "UPDATE",
            )
            .all()
        )
        assert len(promotions) == 1, "a promotion that happened writes exactly one chain row (invariant 2)"

    def test_one_LIVE_tie_alongside_finished_ones_still_refuses_the_promotion(
        self, client: TestClient, auth_headers: dict, db_session: Session
    ):
        """The negative control for the test above, and the count the planner is handed.

        Three shipped jobs and one released one is ONE thing to go and untie. Naming four
        would send the planner into three finished work orders looking for ties they must
        not touch.
        """
        sheet = _make_part(db_session, "ASSY-MIXED", part_type="raw_material")
        for index, finished in enumerate((WorkOrderStatus.COMPLETE, WorkOrderStatus.CLOSED, WorkOrderStatus.CANCELLED)):
            self._tie(db_session, sheet, wo_status=finished, qty_consumed=1.0, suffix=f"-F{index}")
        self._tie(db_session, sheet, wo_status=WorkOrderStatus.RELEASED, suffix="-LIVE")

        response = client.post(
            "/api/v1/bom/import/commit",
            headers=auth_headers,
            json=_commit_payload("ASSY-MIXED", component_numbers=("COMP-M",)),
        )

        assert response.status_code == status.HTTP_201_CREATED, response.text
        db_session.refresh(sheet)
        assert sheet.part_type == "raw_material", "live demand still stands, so the class is left alone"
        assert any(
            "ASSY-MIXED" in warning and "1 unfinished work order still ties" in warning
            for warning in response.json()["warnings"]
        ), response.json()["warnings"]
